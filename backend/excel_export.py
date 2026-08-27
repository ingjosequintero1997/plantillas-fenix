from __future__ import annotations

import io
import re
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

try:
    from .validators import to_date_iso
except ImportError:
    from validators import to_date_iso

# Fechas centinela usadas para indicar "sin dato"
DATE_SENTINELS = {"", "1900-01-01", "1800-01-01", "1845-01-01", "SIN DATO", "00:00:00"}


def col_letter(n: int) -> str:
    result = ""
    while n > 0:
        n, r = divmod(n - 1, 26)
        result = chr(65 + r) + result
    return result


def _trimestre_formula(fum_col: int, date_col: int, default: str, limite2: int = 28, limite3: int = 60) -> str:
    """Genera la fórmula de trimestre basada en FUM y una fecha de prueba.
    Usa nombres de función en inglés (formato interno de .xlsx) y comas.
    `default` es la expresión de valor seguro si no se puede calcular."""
    fum = col_letter(fum_col)
    fec = col_letter(date_col)
    return (
        f'=IFERROR(IF(DATEDIF({fum}{{r}},{fec}{{r}},"D")/7<14,"1 Trim",'
        f'IF(DATEDIF({fum}{{r}},{fec}{{r}},"D")/7<{limite2},"2 Trim","3 Trim")),{default})'
    )


def _default_for_type(ttype: str) -> str:
    # Excel no soporta fechas reales antes de 1900; el campo 1845-01-01 se
    # entrega como texto para que quede visible exactamente como se requiere.
    if ttype == "DATE":
        return '"1845-01-01"'
    if ttype in ("INT", "DECIMAL"):
        return "0"
    return '"SIN DATO"'


# Columnas calculadas: índice (1-based) → generador de fórmula (recibe fila)
def build_formulas(types_by_col: dict[int, str] | None = None) -> dict[int, callable]:
    types_by_col = types_by_col or {}
    t = lambda c: types_by_col.get(c, "TEXT")
    d = lambda c: _default_for_type(t(c))
    f = {}

    # 1. EDAD (H) = DATEDIF(G=FechaNac; HOY(); "Y")
    f[8] = lambda r, _d=d(8): f'=IFERROR(IF(G{r}="",{_d},DATEDIF(G{r},TODAY(),"Y")),{_d})'

    # 2. FPP (AF) = FUM(AE) + 280
    f[32] = lambda r, _d=d(32): f'=IFERROR(IF(AE{r}="",{_d},AE{r}+280),{_d})'

    # 3. Días para el parto (AG) = FPP(AF) - HOY()
    f[33] = lambda r, _d=d(33): f'=IFERROR(IF(AF{r}="",{_d},AF{r}-TODAY()),{_d})'

    # 4. Alarma (AH) según días (AG)
    f[34] = lambda r, _d=d(34): (
        f'=IFERROR(IF(AG{r}="",{_d},IF(AG{r}<0,"Nacido",'
        f'IF(AG{r}<=7,"Semana de parto",IF(AG{r}<=28,"Menos 4 sem","Pendiente")))),{_d})'
    )

    # 5. Edad gestacional al ingreso (AI) = (Ingreso(AD) - FUM(AE)) en semanas
    f[35] = lambda r, _d=d(35): f'=IFERROR(IF(OR(AE{r}="",AD{r}=""),{_d},DATEDIF(AE{r},AD{r},"D")/7),{_d})'

    # 6. Trimestre de inicio (AJ) según semanas (AI)
    f[36] = lambda r, _d=d(36): f'=IFERROR(IF(AI{r}="",{_d},IF(AI{r}<14,"1 Trim",IF(AI{r}<28,"2 Trim","3 Trim"))),{_d})'

    # 7. IMC inicial (BB) = peso(AZ) / talla(BA)^2
    f[54] = lambda r, _d=d(54): f'=IFERROR(IF(OR(AZ{r}="",BA{r}=""),{_d},AZ{r}/BA{r}^2),{_d})'

    # 8. Clasificación IMC (BC) según IMC (BB)
    f[55] = lambda r, _d=d(55): (
        f'=IFERROR(IF(BB{r}="",{_d},IF(BB{r}<18.5,"Bajo peso",'
        f'IF(BB{r}<25,"Peso normal",IF(BB{r}<30,"Sobrepeso",'
        f'IF(BB{r}<35,"Obesidad grado 1",IF(BB{r}<40,"Obesidad grado 2","Obesidad grado 3")))))),{_d})'
    )

    # 9. Trimestres de tamizajes VIH / Sífilis (FUM=AE + fecha de la prueba)
    f[68] = _trimestre_formula(31, 67, d(68))    # BP Asesoría VIH  ← BO
    f[71] = _trimestre_formula(31, 69, d(71))    # BS Tamizaje VIH 1 ← BQ
    f[74] = _trimestre_formula(31, 72, d(74))    # BV Tamizaje VIH 2 ← BT
    f[77] = _trimestre_formula(31, 75, d(77))    # BY Tamizaje VIH 3 ← BW
    f[80] = _trimestre_formula(31, 78, d(80))    # CB Sífilis 1 ← BZ
    f[83] = _trimestre_formula(31, 81, d(83))    # CE Sífilis 2 ← CC
    f[86] = _trimestre_formula(31, 84, d(86))    # CH Sífilis 3 ← CF
    f[89] = _trimestre_formula(31, 87, d(89))    # CK Segunda prueba VIH ← CI

    # 10. Trimestre confirmatorio (CM) con límites 13/26 ← CL
    f[91] = _trimestre_formula(31, 90, d(91), limite2=27, limite3=60)

    # 11. Número total de controles (ET) = contar fechas > 0
    controls = [132, 134, 136, 138, 140, 142, 144, 146, 148]
    count = "+".join(f'COUNTIF({col_letter(c)}{{r}},">0")' for c in controls)
    f[150] = lambda r, _count=count, _d=d(150): f"=IFERROR({_count.format(r=r)},{_d})"

    # 12. Último control (EU) = MAX de las fechas de control
    max_refs = ",".join(f"{col_letter(c)}{{r}}" for c in controls)
    f[151] = lambda r, _refs=max_refs, _d=d(151): f'=IF(MAX({_refs.format(r=r)})=0,{_d},MAX({_refs.format(r=r)}))'

    # 13. Edad gestacional actual (EV) = (Último control(EU) - FUM(AE)) en semanas
    f[152] = lambda r, _d=d(152): (
        f'=IF(OR(AE{r}="",EU{r}="1845-01-01"),{_d},IFERROR(DATEDIF(AE{r},EU{r},"D")/7,{_d}))'
    )

    # 14. IMC actual (EY) = peso(EW) / talla(EX)^2
    f[155] = lambda r, _d=d(155): f'=IFERROR(IF(OR(EW{r}="",EX{r}=""),{_d},EW{r}/EX{r}^2),{_d})'

    return f


def parse_corrected(corrected_text: str) -> list[list[str]]:
    rows = []
    for line in corrected_text.replace("\r\n", "\n").split("\n"):
        line = line.rstrip("\n")
        if not line.strip():
            continue
        rows.append(line.split("|"))
    return rows


def build_data_excel(corrected_text: str, template: list[dict]) -> io.BytesIO:
    headers = [t["name"] for t in template]
    types = [t["type"] for t in template]
    types_by_col = {i + 1: t["type"] for i, t in enumerate(template)}
    formulas = build_formulas(types_by_col)
    rows = parse_corrected(corrected_text)

    wb = Workbook()
    wb.calculation.fullCalcOnLoad = True  # Recalcula todas las fórmulas al abrir
    ws = wb.active
    ws.title = "DATA"

    header_fill = PatternFill("solid", fgColor="1B5E20")
    header_font = Font(bold=True, color="FFFFFF", size=10, name="Calibri")
    thin = Border(*[Side(style="thin", color="BDBDBD")] * 4)

    for c, h in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(wrap_text=True, vertical="center")
        cell.border = thin

    for ridx, row in enumerate(rows, start=2):
        for c, (ttype, val) in enumerate(zip(types, row), start=1):
            if c in formulas:
                fdef = formulas[c]
                formula = fdef(ridx) if callable(fdef) else fdef.format(r=ridx)
                cell = ws.cell(row=ridx, column=c, value=formula)
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = thin
                # Las columnas calculadas de tipo fecha/número deben mostrar
                # formato correcto (si no, Excel muestra el número serial).
                if ttype == "DATE":
                    cell.number_format = "yyyy-mm-dd"
                elif ttype == "INT":
                    cell.number_format = "0"
                elif ttype == "DECIMAL":
                    cell.number_format = "0.00"
                continue
            val = (val or "").strip()
            if ttype == "DATE":
                if val in DATE_SENTINELS:
                    # Fecha sin dato: campo visible 1845-01-01 (texto, Excel no
                    # soporta fechas reales antes de 1900).
                    cell = ws.cell(row=ridx, column=c, value="1845-01-01")
                else:
                    iso = to_date_iso(val)
                    if iso:
                        try:
                            cell = ws.cell(row=ridx, column=c, value=datetime.strptime(iso, "%Y-%m-%d"))
                        except Exception:
                            cell = ws.cell(row=ridx, column=c, value=val)
                    else:
                        cell = ws.cell(row=ridx, column=c, value=val)
                cell.number_format = "yyyy-mm-dd"
                cell.border = thin
            elif ttype in ("INT", "DECIMAL"):
                num = re.sub(r"[^0-9.\-]", "", val)
                try:
                    cell = ws.cell(row=ridx, column=c, value=float(num) if "." in num else int(num))
                except Exception:
                    # Número sin dato o no convertible: 0
                    cell = ws.cell(row=ridx, column=c, value=0)
                cell.border = thin
            else:
                # Texto sin dato: SIN DATO
                cell = ws.cell(row=ridx, column=c, value=val if val else "SIN DATO")
                cell.border = thin

    for c in range(1, len(headers) + 1):
        letter = col_letter(c)
        max_len = max(len(str(headers[c - 1])), 10)
        ws.column_dimensions[letter].width = min(max_len + 3, 45)

    ws.freeze_panes = "A2"
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def build_reporte_errores_excel(corrected_text: str, template: list[dict], errors_by_cell: dict) -> io.BytesIO:
    """Genera una UNICA hoja de calculo con la data y una columna final
    'RESULTADO DE VALIDACION' que describe el/los errores de cada fila.
    Las celdas con dato invalido se marcan en rojo para localizarlas."""
    headers = [t["name"] for t in template]
    types = [t["type"] for t in template]
    rows = parse_corrected(corrected_text)

    wb = Workbook()
    ws = wb.active
    ws.title = "DATA"

    header_fill = PatternFill("solid", fgColor="1B5E20")
    header_font = Font(bold=True, color="FFFFFF", size=10, name="Calibri")
    thin = Border(*[Side(style="thin", color="BDBDBD")] * 4)
    error_fill = PatternFill("solid", fgColor="FECACA")  # rojo mas intenso
    error_font = Font(color="B91C1C", bold=True)
    valid_fill = PatternFill("solid", fgColor="DCFCE7")  # verde claro (VALIDADO)

    ncols = len(headers) + 1  # + RESULTADO DE VALIDACION
    for c in range(1, ncols + 1):
        label = headers[c - 1] if c <= len(headers) else "RESULTADO DE VALIDACION"
        cell = ws.cell(row=1, column=c, value=label)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(wrap_text=True, vertical="center")
        cell.border = thin

    # Construir filas con append (rapido)
    for ridx, row in enumerate(rows, start=2):
        out_row = [None] * ncols
        for c, (ttype, val) in enumerate(zip(types, row), start=1):
            val = (val or "").strip()
            if ttype == "DATE":
                if val in DATE_SENTINELS:
                    out_row[c - 1] = "1845-01-01"
                else:
                    iso = to_date_iso(val)
                    out_row[c - 1] = iso if iso else val
            elif ttype in ("INT", "DECIMAL"):
                num = re.sub(r"[^0-9.\-]", "", val)
                try:
                    out_row[c - 1] = float(num) if "." in num else int(num)
                except Exception:
                    out_row[c - 1] = val
            else:
                out_row[c - 1] = val if val else "SIN DATO"
        # RESULTADO DE VALIDACION: describir el error especifico de la fila
        fila_idx = ridx - 2  # 0-based dentro de rows
        errores_fila = [(h, errors_by_cell[(fila_idx, h)]) for h in headers if (fila_idx, h) in errors_by_cell]
        if errores_fila:
            desc = " - ".join(f"{h}: {msg}" for h, msg in errores_fila)
        else:
            desc = "VALIDADO"
        out_row[ncols - 1] = desc
        ws.append(out_row)

    # Marcar en rojo las celdas con dato invalido y la celda RESULTADO
    col_idx = {h: i + 1 for i, h in enumerate(headers)}
    # Rellenar RESULTADO DE VALIDACION por fila
    for ridx in range(len(rows)):
        rc = ws.cell(row=ridx + 2, column=ncols)
        rc.border = thin
        tiene = any((ridx, h) in errors_by_cell for h in headers)
        if tiene:
            rc.fill = error_fill
            rc.font = error_font
        else:
            rc.fill = valid_fill
            rc.font = Font(color="166534", bold=True)
    # Celdas con dato invalido
    for (fila_idx, h) in errors_by_cell.keys():
        if fila_idx >= len(rows):
            continue
        c = col_idx.get(h)
        if c is None:
            continue
        cell = ws.cell(row=fila_idx + 2, column=c)
        cell.fill = error_fill
        cell.font = error_font

    for c in range(1, ncols + 1):
        letter = col_letter(c)
        if c <= len(headers):
            max_len = max(len(str(headers[c - 1])), 10)
            ws.column_dimensions[letter].width = min(max_len + 3, 45)
        else:
            ws.column_dimensions[letter].width = 80  # RESULTADO DE VALIDACION
    ws.freeze_panes = "A2"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf
