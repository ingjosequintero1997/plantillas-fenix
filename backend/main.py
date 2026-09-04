from fastapi import FastAPI, UploadFile, File, HTTPException, Form, Query, Depends, Request
from fastapi.responses import StreamingResponse, JSONResponse
from fastapi.middleware.cors import CORSMiddleware
import pandas as pd
import io
import re
import gzip
import base64
import json
from datetime import datetime
from pydantic import BaseModel
from sqlalchemy.exc import OperationalError
try:
	from .utils import fuzzy_map, normalize_text
	from .templates_registry import get_template_by_key, list_templates_meta
	from .validators import validate_and_correct, to_date_iso
	from .evaluator import evaluate, build_evaluation_excel
except ImportError:
	from utils import fuzzy_map, normalize_text
	from templates_registry import get_template_by_key, list_templates_meta
	from validators import validate_and_correct, to_date_iso
	from evaluator import evaluate, build_evaluation_excel

import os

API_ROOT_PATH = os.environ.get("API_ROOT_PATH", "")
app = FastAPI(title="Validador IPS", root_path=API_ROOT_PATH)

class RevalidatePayload(BaseModel):
	raw_text: str
	mapping: dict[str, str | None]
	template_key: str = "gestante"
	mode: str = "limpiador"

def template_names(template: list[dict]):
	return [t['name'] for t in template]

# CORS config
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# ─── Auth ────────────────────────────────────────────────────────────────
try:
    from .auth_utils import (
        create_token,
        get_current_user,
        hash_password,
        require_admin,
        verify_credentials,
    )
    from .database import init_db as db_init_db, SessionLocal, Prestador, User, Cargue, HistoriaClinica, PrestadorPlantilla, crear_tabla_gestantes, GESTANTE_COLUMNS
    from . import gcs_storage
    from . import corporate_db
except ImportError:
    from auth_utils import (
        create_token,
        get_current_user,
        hash_password,
        require_admin,
        verify_credentials,
    )
    from database import init_db as db_init_db, SessionLocal, Prestador, User, Cargue, HistoriaClinica, PrestadorPlantilla, crear_tabla_gestantes, GESTANTE_COLUMNS
    import gcs_storage
    import corporate_db

class LoginPayload(BaseModel):
    username: str
    password: str

@app.post("/auth/login")
async def auth_login(payload: LoginPayload):
    ensure_db_ready()
    user = verify_credentials(payload.username, payload.password)
    if user is None:
        raise HTTPException(status_code=401, detail="Usuario o contraseña incorrectos")
    token = create_token(user)
    return {
        "token": token,
        "user": {"id": user.id, "username": user.username, "name": user.name, "role": user.role},
    }

@app.get("/auth/me")
async def auth_me(current_user: User = Depends(get_current_user)):
    prestador_id = None
    prestador_nombre = None
    prestador_ips = None
    try:
        db = SessionLocal()
        try:
            prestador = db.query(Prestador).filter(Prestador.user_id == current_user.id).first()
            prestador_id = prestador.id if prestador else None
            prestador_nombre = prestador.nombre if prestador else None
            prestador_ips = prestador.ips if prestador else None
        finally:
            db.close()
    except Exception:
        pass
    return {
        "user": {
            "id": current_user.id,
            "username": current_user.username,
            "name": current_user.name,
            "role": current_user.role,
            "prestador_id": prestador_id,
            "prestador_nombre": prestador_nombre,
            "prestador_ips": prestador_ips,
        }
    }


def seed_admin():
    try:
        db = SessionLocal()
        try:
            existing = db.query(User).filter(User.username == "admin").first()
            if existing is None:
                db.add(
                    User(
                        username="admin",
                        password_hash=hash_password("admin123"),
                        name="Administrador",
                        role="admin",
                        active=True,
                    )
                )
                db.commit()
        finally:
            db.close()
    except Exception:
        pass  # Sin BD persistente: se usa el admin de respaldo


_db_ready = False

def ensure_db_ready():
    # Inicializa la BD de forma perezosa (solo al primer login), para no
    # bloquear el cold start de las funciones serverless con la conexión a BD.
    global _db_ready
    if _db_ready:
        return
    try:
        db_init_db()
    except Exception:
        pass
    try:
        seed_admin()
    except Exception:
        pass
    _db_ready = True

@app.get("/health")
async def health():
	return {"status": "ok"}

@app.get("/debug-db")
async def debug_db():
	"""Diagnostica el estado de la conexion a la base de datos."""
	try:
		from .database import engine, DB_AVAILABLE, DATABASE_URL
	except ImportError:
		from database import engine, DB_AVAILABLE, DATABASE_URL
	info = {
		"db_available": DB_AVAILABLE,
		"engine": str(engine.url).split("@")[-1] if "@" in str(engine.url) else str(engine.url),
		"es_postgres": str(engine.url).startswith("postgresql"),
		"es_sqlite": str(engine.url).startswith("sqlite"),
	}
	try:
		from sqlalchemy import inspect, text
		with engine.connect() as conn:
			conn.execute(text("SELECT 1"))
		insp = inspect(engine)
		tables = insp.get_table_names()
		info["conectado"] = True
		info["tablas"] = tables[:100]
		# Conteo de cargues y columnas de la tabla cargues
		try:
			with engine.connect() as conn:
				cnt = conn.execute(text("SELECT COUNT(*) FROM cargues")).scalar()
			info["cargues_count"] = cnt
			try:
				rows = conn.execute(text("SELECT id, template_key, user_id, original_filename, status, created_at FROM cargues ORDER BY id DESC LIMIT 5")).fetchall()
				info["cargues_recientes"] = [list(r) for r in rows]
			except Exception as e:
				info["cargues_recientes"] = f"error: {str(e)[:150]}"
		except Exception as e:
			info["cargues_count"] = f"error: {str(e)[:150]}"
		try:
			cols = [c['name'] for c in inspect(engine).get_columns('cargues')]
			info["cargues_columnas"] = cols
		except Exception as e:
			info["cargues_columnas"] = f"error: {str(e)[:150]}"
	except Exception as e:
		info["conectado"] = False
		info["error_conexion"] = str(e)[:300]
	return info

@app.get("/debug-corporate-db")
async def debug_corporate_db():
	"""Diagnostica la conexión con la BD corporativa Dusakawi."""
	try:
		conectado = corporate_db.test_conexion_corporativa()
		if conectado:
			# Intentar leer tabla de afiliados
			engine = corporate_db.get_corporate_connection()
			if engine:
				from sqlalchemy import text
				conn = engine.connect()
				try:
					result = conn.execute(text('SELECT COUNT(*) FROM administrativo."af_afiliado"')).scalar()
					host = os.environ.get("CORP_DB_HOST", "")
					port = os.environ.get("CORP_DB_PORT", "")
					dbname = os.environ.get("CORP_DB_NAME", "")
					return {
						"conectado": True,
						"db_corporativa": f"postgresql://***@{host}:{port}/{dbname}",
						"tabla_afiliados": "administrativo.af_afiliado",
						"total_afiliados": int(result),
						"error": None
					}
				finally:
					conn.close()
		host = os.environ.get("CORP_DB_HOST", "")
		port = os.environ.get("CORP_DB_PORT", "")
		dbname = os.environ.get("CORP_DB_NAME", "")
		return {
			"conectado": False,
			"db_corporativa": f"postgresql://***@{host}:{port}/{dbname}",
			"tabla_afiliados": "administrativo.af_afiliado",
			"error": "No se pudo establecer conexión con BD corporativa"
		}
	except Exception as e:
		return {
			"conectado": False,
			"error": str(e)[:300]
		}

@app.post("/setup-gestantes")
async def setup_gestantes(current_user: User = Depends(require_admin)):
	"""Crea el esquema public y la tabla gestantes con todos los encabezados."""
	try:
		ncols = crear_tabla_gestantes()
		return {"ok": True, "tabla": "public.gestantes", "columnas": ncols}
	except Exception as e:
		raise HTTPException(status_code=500, detail=f"No se pudo crear la tabla: {e}")

@app.get("/debug-oidc")
async def debug_oidc(request: Request):
	token = request.headers.get("x-vercel-oidc-token", "")
	return {
		"oidc_header_present": bool(token),
		"oidc_header_len": len(token),
		"oidc_header_preview": token[:40],
		"all_headers": [h for h in request.headers.keys()],
	}

@app.get("/template")
async def get_template(template_key: str = Query(default="gestante")):
	meta = get_template_by_key(template_key)
	return {
		"template_key": meta["key"],
		"label": meta["label"],
		"template": meta["template"],
	}

@app.get("/templates")
async def get_templates(current_user: User = Depends(get_current_user)):
	all_templates = list_templates_meta()
	if current_user.role == "admin":
		return {"templates": all_templates}
	# Prestador: solo las plantillas que tiene asignadas
	db = SessionLocal()
	try:
		prestador = db.query(Prestador).filter(Prestador.user_id == current_user.id).first()
		if prestador is None:
			return {"templates": []}
		assigned = [p.template_key for p in prestador.plantillas]
	except Exception:
		assigned = []
	finally:
		db.close()
	if not assigned:
		return {"templates": []}
	filtered = [t for t in all_templates if t["key"] in assigned]
	return {"templates": filtered}


# ─── Detección automática de plantilla ────────────────────────────────────
def _template_names_of(key: str) -> list[str]:
	try:
		from .templates_registry import TEMPLATE_REGISTRY
	except ImportError:
		from templates_registry import TEMPLATE_REGISTRY
	entry = TEMPLATE_REGISTRY.get(key)
	if not entry:
		return []
	return [t["name"] for t in entry["template_factory"]()]


def union_template_names() -> list[str]:
	names = set()
	for key in ("gestante", "citologia", "mamografia", "penta"):
		for name in _template_names_of(key):
			names.add(name)
	return sorted(names)


def detect_best_template(headers: list[str]) -> tuple[str | None, dict]:
	best_key = None
	best_score = -1.0
	scores: dict[str, float] = {}
	for key in ("gestante", "citologia", "mamografia", "penta"):
		names = _template_names_of(key)
		if not names:
			continue
		row_map = fuzzy_map(headers, names, score_cutoff=76)
		hits = sum(1 for value in row_map.values() if value is not None)
		coverage = hits / max(1, len(names))
		score = hits + coverage
		scores[key] = round(score, 3)
		if score > best_score:
			best_score = score
			best_key = key
	return best_key, scores

def parse_pipe_text(text: str) -> pd.DataFrame:
	return pd.read_csv(
		io.StringIO(text),
		sep='|',
		header=None,
		dtype=str,
		engine='python',
		keep_default_na=False,
	)

def parse_excel_bytes(contents: bytes) -> pd.DataFrame:
	return pd.read_excel(
		io.BytesIO(contents),
		header=None,
		dtype=str,
		engine='openpyxl',
	)

def clean_string(value) -> str:
	if value is None or pd.isna(value):
		return ""
	return str(value).strip()

def make_unique_headers(headers: list[str]) -> list[str]:
	seen = {}
	result = []
	for index, name in enumerate(headers, start=1):
		base = clean_string(name) or f"C{index}"
		if base not in seen:
			seen[base] = 1
			result.append(base)
			continue
		seen[base] += 1
		result.append(f"{base}_{seen[base]}")
	return result

def detect_header_row(raw_df: pd.DataFrame, expected_names: list[str]) -> int | None:
	if raw_df.empty:
		return None
	max_scan = min(len(raw_df), 30)
	best_index = None
	best_hits = 0
	best_coverage = 0.0
	for ridx in range(max_scan):
		cells_raw = [clean_string(v) for v in raw_df.iloc[ridx].tolist()]
		cells = [c for c in cells_raw if normalize_text(c)]
		if not cells:
			continue

		# Mide qué tan bien la fila se parece a encabezados de la plantilla.
		row_map = fuzzy_map(cells, expected_names, score_cutoff=76)
		hits = sum(1 for value in row_map.values() if value is not None)
		coverage = hits / max(1, len(cells))

		if hits > best_hits or (hits == best_hits and coverage > best_coverage):
			best_hits = hits
			best_coverage = coverage
			best_index = ridx

	if best_index is None:
		return None

	# Para archivos con pocas columnas, el umbral fijo de 4 hits impide detectar encabezado.
	if best_hits <= 0:
		return None
	# Usa el ancho real de la fila para calcular umbral mínimo razonable.
	best_cells = [clean_string(v) for v in raw_df.iloc[best_index].tolist()]
	non_empty_cells = [value for value in best_cells if normalize_text(value)]
	cell_count = max(1, len(non_empty_cells))
	min_hits = 1 if cell_count <= 2 else (2 if cell_count <= 5 else 4)
	coverage_ok = best_coverage >= (0.50 if cell_count <= 2 else 0.40 if cell_count <= 5 else 0.20)
	return best_index if (best_hits >= min_hits and coverage_ok) else None

def normalize_source_dataframe(raw_df: pd.DataFrame, expected_names: list[str]) -> pd.DataFrame:
	raw_df = raw_df.copy()
	raw_df = raw_df.dropna(axis=0, how='all')
	if raw_df.empty:
		return raw_df

	header_idx = detect_header_row(raw_df, expected_names)
	if header_idx is not None:
		headers = make_unique_headers([clean_string(v) for v in raw_df.iloc[header_idx].tolist()])
		df = raw_df.iloc[header_idx + 1 :].reset_index(drop=True).copy()
		df.columns = headers
	else:
		df = raw_df.reset_index(drop=True).copy()
		df.columns = [f"C{i + 1}" for i in range(df.shape[1])]

	# Si el archivo trae la columna "RESULTADO DE VALIDACION" (de una descarga
	# previa del reporte), se elimina para poder re-validar la data corregida.
	col_validacion = [c for c in df.columns if "RESULTADO DE VALIDACION" in str(c).upper() or str(c).upper() == "ERRORES"]
	if col_validacion:
		df = df.drop(columns=col_validacion[0])

	# Compatibilidad con pandas 2.x/3.x sin usar applymap.
	df = df.apply(lambda col: col.map(clean_string))

	# Descarta únicamente columnas completamente vacías al final del bloque.
	if df.shape[1] > 0:
		non_empty_cols = df.ne("").any(axis=0)
		last_idx = df.shape[1] - 1
		while last_idx >= 0 and not bool(non_empty_cols.iloc[last_idx]):
			last_idx -= 1
		if last_idx >= 0:
			df = df.iloc[:, : last_idx + 1]

	# Elimina filas completamente vacías.
	df = df[df.ne("").any(axis=1)]
	return df.reset_index(drop=True)


def headers_are_generic(headers: list[str]) -> bool:
	if not headers:
		return False
	for header in headers:
		n = normalize_text(header)
		if not re.fullmatch(r"C\s*\d+", n):
			return False
	return True

def infer_mapping(headers: list[str], active_template: list[dict]) -> dict[str, str | None]:
	active_names = template_names(active_template)
	# Solo aplica mapeo posicional cuando encabezados son genéricos (C1..Cn).
	if headers_are_generic(headers):
		mapping = {header: None for header in headers}
		# Si no se detectó encabezado real, usa posición para no perder datos al exportar.
		limit = min(len(headers), len(active_names))
		for idx in range(limit):
			mapping[headers[idx]] = active_names[idx]
		return mapping
	map_suggest = fuzzy_map(headers, active_names)
	return map_suggest


def build_mapping_stats(headers: list[str], mapping: dict[str, str | None], template_cols: int) -> dict:
	mapped_headers = [key for key, value in mapping.items() if value is not None]
	unmapped_headers = [key for key, value in mapping.items() if value is None]
	mapped_template_names = {value for value in mapping.values() if value is not None}
	header_coverage = round((len(mapped_headers) / max(1, len(headers))) * 100, 2)
	template_coverage = round((len(mapped_template_names) / max(1, template_cols)) * 100, 2)
	return {
		"mapped_headers": len(mapped_headers),
		"total_headers": len(headers),
		"coverage_percent": header_coverage,
		"template_coverage_percent": template_coverage,
		"unmapped_headers": unmapped_headers,
	}


def build_structure_validation(headers: list[str], template_cols: int, row_count: int) -> dict:
	return {
		"input_columns": len(headers),
		"template_columns": template_cols,
		"column_diff": len(headers) - template_cols,
		"row_count": row_count,
	}

def _gz_compress(text: str) -> str:
	# Comprime y codifica en base64 para reducir el tamaño de la respuesta
	# (las respuestas grandes son lentas/inestables en serverless de Vercel).
	return base64.b64encode(gzip.compress(text.encode("utf-8"))).decode("ascii")

def _correccion_excel(tipo, tdef, val_str):
	"""Mensaje de correccion claro para el comentario de una celda con error."""
	try:
		from .validators import _mensaje_esperado
	except ImportError:
		from validators import _mensaje_esperado
	msg = _mensaje_esperado(tipo, tdef, tdef.get("name", ""), val_str)
	# Prefix amigable segun el tipo
	if tipo == "SET":
		return "Corrige: " + msg
	if tipo == "INT":
		return "Numero invalido. " + msg
	if tipo == "DECIMAL":
		return "Numero invalido. " + msg
	if tipo == "DATE":
		return "Fecha invalida. " + msg
	return "Dato invalido. " + msg


def build_response_payload(df: pd.DataFrame, mapping: dict, raw_text: str, template_key: str, active_template: list[dict]):
	corrected_df, _corrector_logs, stats = validate_and_correct(df, mapping, active_template)
	try:
		from .validators import normalizar_fechas_df, limpiar_celdas_export, validate_cross_fields
	except ImportError:
		from validators import normalizar_fechas_df, limpiar_celdas_export, validate_cross_fields
	corrected_df = normalizar_fechas_df(corrected_df, active_template)
	corrected_df = limpiar_celdas_export(corrected_df)
	corrected_df = corrected_df.fillna("SIN DATO").astype(str)
	for col in corrected_df.columns:
		corrected_df[col] = corrected_df[col].str.replace(r'[\r\n]+', ' ', regex=True)
	buf = io.StringIO()
	corrected_df.to_csv(buf, sep='|', index=False, header=False, na_rep='SIN DATO')
	corrected_text = buf.getvalue()
	# Log de errores por celda: usar _errores_rapidos sobre la data corregida
	# (misma fuente que el Excel de errores para consistencia total).
	errors_by_cell = _errores_rapidos(corrected_text, template_key)
	logs_sample = []
	tmap2 = {t["name"]: t for t in active_template}
	for (row_idx, col_name), msg in errors_by_cell.items():
		val = ""
		if row_idx < len(corrected_df) and col_name in corrected_df.columns:
			val = str(corrected_df.iloc[row_idx][col_name]) if col_name in corrected_df.columns else ""
		logs_sample.append({
			"row": row_idx + 1,
			"column": col_name,
			"original": val,
			"corrected": msg,
			"status": "error",
		})
	# Cross-field validation for gestante template
	cross_field_errors = []
	if template_key == "gestante":
		try:
			cross_field_errors = validate_cross_fields(corrected_df)
		except Exception:
			pass
	all_logs = logs_sample + [
		{"row": e["row"], "column": e["column"], "original": "", "corrected": e["message"], "status": e["severity"]}
		for e in cross_field_errors
	]
	stats["cross_field_errors"] = len(cross_field_errors)
	preview_rows = corrected_df.head(30).to_dict(orient='records')
	return {
		"success": True,
		"template_key": template_key,
		"mapping_suggested": mapping,
		"mapping": mapping,
		"summary": stats,
		"logs_sample": all_logs[:50000],
		"cross_field_errors": cross_field_errors,
		"corrected_text": _gz_compress(corrected_text),
		"preview_rows": preview_rows,
		"raw_text": _gz_compress(raw_text),
		"template_names": [t['name'] for t in active_template],
		"compressed": True,
	}

@app.post("/upload")
async def upload_file(
	file: UploadFile = File(...),
	template_key: str = Form(default="auto"),
	strict_mode: bool = Form(default=False),
	min_template_coverage: float = Form(default=95.0),
	require_exact_columns: bool = Form(default=True),
	mode: str = Form(default="limpiador"),
):
	filename = (file.filename or '').lower()
	if not (filename.endswith('.txt') or filename.endswith('.xlsx') or filename.endswith('.xls')):
		raise HTTPException(status_code=400, detail="Solo se permiten .txt, .xlsx o .xls")

	try:
		contents = await file.read()
		# Detectar y descomprimir gzip (enviado desde frontend para archivos > 1 MB)
		if len(contents) >= 2 and contents[:2] == b'\x1f\x8b':
			try:
				contents = gzip.decompress(contents)
			except Exception:
				pass

		raw_df = parse_pipe_text(contents.decode(errors='replace')) if filename.endswith('.txt') else parse_excel_bytes(contents)

		# Detección automática de plantilla a partir de la fila de encabezados
		union = union_template_names()
		header_idx = detect_header_row(raw_df, union)
		if header_idx is not None:
			detect_headers = make_unique_headers([clean_string(v) for v in raw_df.iloc[header_idx].tolist()])
		else:
			detect_headers = [f"C{i + 1}" for i in range(raw_df.shape[1])]

		if not template_key or template_key == "auto":
			detected_key, detect_scores = detect_best_template(detect_headers)
			best_score = detect_scores.get(detected_key, 0) if detected_key else 0
			if detected_key is None or best_score < 2:
				raise HTTPException(status_code=400, detail="No se pudo identificar la plantilla del archivo. Asegúrate de usar la plantilla descargada de la aplicación.")
			template_key = detected_key

		meta = get_template_by_key(template_key)
		active_template = meta["template"]
		active_names = template_names(active_template)
		template_key = meta["key"]

		df = normalize_source_dataframe(raw_df, active_names)

		if len(df) == 0:
			raise HTTPException(status_code=400, detail="Archivo vacío")

		# Regla de integridad: en modo LIMPIADOR ningun dato puede quedar vacio.
		# Se rellena con el valor por tipo segun el instructivo: texto->SIN DATO,
		# numerico->0, fecha->1845-01-01, SET->SIN DATO.
		# En modo VALIDADOR NO se rellena: los campos vacios son ERROR y el
		# prestador debe corregirlos (validacion estricta contra el instructivo).
		if mode != "validador":
			try:
				from .validators import rellenar_vacios
			except ImportError:
				from validators import rellenar_vacios
			df = rellenar_vacios(df, active_template)

		orig_headers = list(df.columns)
		map_suggest = infer_mapping(orig_headers, active_template)
		mapping_stats = build_mapping_stats(orig_headers, map_suggest, len(active_names))
		structure_validation = build_structure_validation(orig_headers, len(active_names), len(df))
		strict_validation = {
			"strict_mode": bool(strict_mode),
			"min_template_coverage": float(min_template_coverage),
			"require_exact_columns": bool(require_exact_columns),
		}
		strict_reasons = []

		if strict_mode:
			reasons = []
			if mapping_stats["template_coverage_percent"] < float(min_template_coverage):
				reasons.append(
					f"Cobertura de plantilla insuficiente ({mapping_stats['template_coverage_percent']}% < {float(min_template_coverage)}%)."
				)
			if require_exact_columns and structure_validation["input_columns"] != structure_validation["template_columns"]:
				reasons.append(
					f"Estructura invalida: columnas de archivo ({structure_validation['input_columns']}) distintas a columnas de plantilla ({structure_validation['template_columns']})."
				)
			if mapping_stats["unmapped_headers"]:
				reasons.append(
					f"Existen encabezados no mapeados: {', '.join(mapping_stats['unmapped_headers'][:20])}"
				)
			strict_reasons = reasons

		strict_validation["status"] = "warning" if strict_reasons else "ok"
		strict_validation["reasons"] = strict_reasons
		if strict_mode and strict_reasons:
			raise HTTPException(status_code=400, detail={
				"message": "El archivo no cumple con los requisitos del modo estricto",
				"reasons": strict_reasons,
			})

# Reordenar el df al orden de la plantilla (evita corrimiento de datos
		# cuando el archivo trae las columnas en otro orden).
		try:
			from .validators import reordenar_a_template
		except ImportError:
			from validators import reordenar_a_template
		df = reordenar_a_template(df, map_suggest, active_template)

		# Normalizar todas las fechas a AAAA-MM-DD antes de exportar cualquier texto
		try:
			from .validators import normalizar_fechas_df
		except ImportError:
			from validators import normalizar_fechas_df
		df = normalizar_fechas_df(df, active_template)

		canonical_raw_text = df.to_csv(sep='|', index=False, header=False)

		if mode == "validador":
			# En modo validador, normalizar y limpiar la data ANTES de validar,
			# para que el log de errores y el Excel usen la misma data canonicalizada.
			try:
				from .validators import normalizar_fechas_df, limpiar_celdas_export, validate_cross_fields
			except ImportError:
				from validators import normalizar_fechas_df, limpiar_celdas_export, validate_cross_fields
			df = normalizar_fechas_df(df, active_template)
			df = limpiar_celdas_export(df)
			df_safe = df.fillna("SIN DATO").astype(str)
			canonical_raw_text = df_safe.to_csv(sep='|', index=False, header=False)
			# Usar _errores_rapidos: misma fuente que el Excel de errores
			errors_by_cell = _errores_rapidos(canonical_raw_text, template_key)
			logs_sample = []
			for (row_idx, col_name), msg in errors_by_cell.items():
				val = ""
				if row_idx < len(df_safe) and col_name in df_safe.columns:
					val = str(df_safe.iloc[row_idx][col_name])
				logs_sample.append({
					"row": row_idx + 1,
					"column": col_name,
					"original": val,
					"corrected": msg,
					"status": "error",
				})
			cross_field_errors = []
			if template_key == "gestante":
				try:
					cross_field_errors = validate_cross_fields(df_safe)
				except Exception:
					pass
			all_logs = logs_sample + [
				{"row": e["row"], "column": e["column"], "original": "", "corrected": e["message"], "status": e["severity"]}
				for e in cross_field_errors
			]
			n = len(df_safe)
			n_errors = len(set(lr["row"] for lr in all_logs if lr["status"] == "error"))
			stats = {
				"total": n,
				"rows_with_errors": n_errors,
				"rows_ok": n - n_errors,
				"errors": len([lr for lr in all_logs if lr["status"] == "error"]),
				"cross_field_errors": len(cross_field_errors),
			}
			raw_text_compressed = _gz_compress(canonical_raw_text)
			if stats["rows_with_errors"] == 0:
				try:
					from .formulas import aplicar_formulas_df
				except ImportError:
					from formulas import aplicar_formulas_df
				df_safe = aplicar_formulas_df(df_safe)
				canonical_raw_text = df_safe.to_csv(sep='|', index=False, header=False)
				raw_text_compressed = _gz_compress(canonical_raw_text)
			return JSONResponse({
				"success": True,
				"template_key": template_key,
				"mode": "validador",
				"mapping_suggested": map_suggest,
				"mapping": map_suggest,
				"summary": stats,
				"logs_sample": all_logs[:50000],
				"corrected_text": raw_text_compressed,
				"raw_text": raw_text_compressed,
				"preview_rows": df_safe.head(30).to_dict(orient='records'),
				"template_names": template_names(active_template),
				"mapping_stats": mapping_stats,
				"structure_validation": structure_validation,
				"strict_validation": strict_validation,
				"compressed": True,
			})

		payload = build_response_payload(df, map_suggest, canonical_raw_text, template_key, active_template)
		return JSONResponse({
			**payload,
			"original_headers": orig_headers,
			"template_names": template_names(active_template),
			"mapping_stats": mapping_stats,
			"structure_validation": structure_validation,
			"strict_validation": strict_validation,
		})
	except HTTPException:
		raise
	except Exception as e:
		raise HTTPException(status_code=500, detail=str(e))

import tempfile
import os as _os

@app.post("/upload-chunk")
async def upload_chunk(
	chunk: UploadFile = File(...),
	upload_id: str = Form(...),
	chunk_index: int = Form(...),
	total_chunks: int = Form(...),
	original_name: str = Form(...),
	template_key: str = Form(default="gestante"),
	strict_mode: bool = Form(default=False),
	min_template_coverage: float = Form(default=95.0),
	require_exact_columns: bool = Form(default=True),
):
	tmp_dir = tempfile.gettempdir()
	chunk_path = _os.path.join(tmp_dir, f"fenix_{upload_id}_{chunk_index}")
	data = await chunk.read()
	with open(chunk_path, "wb") as f:
		f.write(data)

	# Si no es el último chunk, responder "done: false"
	if chunk_index < total_chunks - 1:
		return {"done": False}

	# Último chunk: reensamblar y procesar
	try:
		parts = []
		for i in range(total_chunks):
			p = _os.path.join(tmp_dir, f"fenix_{upload_id}_{i}")
			with open(p, "rb") as f:
				parts.append(f.read())
		contents = b"".join(parts)

		# Limpiar chunks temporales
		for i in range(total_chunks):
			p = _os.path.join(tmp_dir, f"fenix_{upload_id}_{i}")
			try:
				_os.remove(p)
			except Exception:
				pass

		# Procesar como upload normal
		meta = get_template_by_key(template_key)
		active_template = meta["template"]
		active_names = template_names(active_template)
		template_key = meta["key"]
		filename = original_name.lower()

		if not (filename.endswith('.txt') or filename.endswith('.xlsx') or filename.endswith('.xls')):
			raise HTTPException(status_code=400, detail="Solo se permiten .txt, .xlsx o .xls")

		if len(contents) >= 2 and contents[:2] == b'\x1f\x8b':
			try:
				contents = gzip.decompress(contents)
			except Exception:
				pass

		if filename.endswith('.txt'):
			text = contents.decode(errors='replace')
			df = parse_pipe_text(text)
			df = normalize_source_dataframe(df, active_names)
		else:
			df = parse_excel_bytes(contents)
			df = normalize_source_dataframe(df, active_names)

		if len(df) == 0:
			raise HTTPException(status_code=400, detail="Archivo vacío")

		orig_headers = list(df.columns)
		map_suggest = infer_mapping(orig_headers, active_template)
		mapping_stats = build_mapping_stats(orig_headers, map_suggest, len(active_names))
		structure_validation = build_structure_validation(orig_headers, len(active_names), len(df))
		strict_validation = {
			"strict_mode": bool(strict_mode),
			"min_template_coverage": float(min_template_coverage),
			"require_exact_columns": bool(require_exact_columns),
		}
		strict_reasons = []

		if strict_mode:
			reasons = []
			if mapping_stats["template_coverage_percent"] < float(min_template_coverage):
				reasons.append(f"Cobertura de plantilla insuficiente ({mapping_stats['template_coverage_percent']}% < {float(min_template_coverage)}%).")
			if require_exact_columns and structure_validation["input_columns"] != structure_validation["template_columns"]:
				reasons.append(f"Estructura invalida: columnas de archivo ({structure_validation['input_columns']}) distintas a columnas de plantilla ({structure_validation['template_columns']}).")
			if mapping_stats["unmapped_headers"]:
				reasons.append(f"Existen encabezados no mapeados: {', '.join(mapping_stats['unmapped_headers'][:20])}")
			strict_reasons = reasons

		strict_validation["status"] = "warning" if strict_reasons else "ok"
		strict_validation["reasons"] = strict_reasons
		if strict_mode and strict_reasons:
			raise HTTPException(status_code=400, detail={
				"message": "El archivo no cumple con los requisitos del modo estricto",
				"reasons": strict_reasons,
			})

		canonical_raw_text = df.to_csv(sep='|', index=False, header=True)
		payload = build_response_payload(df, map_suggest, canonical_raw_text, template_key, active_template)
		return {"done": True, **payload, **{
			"original_headers": orig_headers,
			"template_names": template_names(active_template),
			"mapping_stats": mapping_stats,
			"structure_validation": structure_validation,
			"strict_validation": strict_validation,
		}}
	except HTTPException:
		raise
	except Exception as e:
		raise HTTPException(status_code=500, detail=str(e))

@app.post("/revalidate")
async def revalidate(payload: RevalidatePayload):
	try:
		meta = get_template_by_key(payload.template_key)
		active_template = meta["template"]
		template_key = meta["key"]
		# En modo validador la data llega sin fila de encabezado (solo datos en
		# orden de plantilla). En limpiador llega con encabezado.
		if payload.mode == "validador":
			df = pd.read_csv(io.StringIO(payload.raw_text), sep='|', dtype=str, engine='python', header=None, keep_default_na=False)
			df = df.fillna('').astype(str)
			if len(df) == 0:
				raise HTTPException(status_code=400, detail="Archivo vacío")
			try:
				from .validators import validate_only, validate_cross_fields, reordenar_a_template
			except ImportError:
				from validators import validate_only, validate_cross_fields, reordenar_a_template
			validation_result = validate_only(df, payload.mapping, active_template)
			# Cross-field validation for gestante template
			cross_field_errors = []
			if template_key == "gestante":
				try:
					df_ordered = reordenar_a_template(df, payload.mapping, active_template)
					cross_field_errors = validate_cross_fields(df_ordered)
				except Exception:
					pass
			# Merge cross-field errors into logs
			all_logs = validation_result["logs"] + [
				{"row": e["row"], "column": e["column"], "original": "", "corrected": e["message"], "status": e["severity"]}
				for e in cross_field_errors
			]
			# Update stats
			validation_result["stats"]["cross_field_errors"] = len(cross_field_errors)
			validation_result["stats"]["errors"] += len(cross_field_errors)
			raw_text_compressed = _gz_compress(payload.raw_text)
			return JSONResponse({
				"success": True,
				"template_key": template_key,
				"mode": "validador",
				"mapping_suggested": payload.mapping,
				"mapping": payload.mapping,
				"summary": validation_result["stats"],
				"logs_sample": all_logs[:5000],
				"cross_field_errors": cross_field_errors,
				"corrected_text": raw_text_compressed,
				"raw_text": raw_text_compressed,
				"template_names": [t['name'] for t in active_template],
				"compressed": True,
			})

		df = pd.read_csv(io.StringIO(payload.raw_text), sep='|', dtype=str, engine='python', keep_default_na=False)
		df = df.fillna('').astype(str)
		if len(df) == 0:
			raise HTTPException(status_code=400, detail="Archivo vacío")
		return JSONResponse(build_response_payload(df, payload.mapping, payload.raw_text, template_key, active_template))
	except HTTPException:
		raise
	except Exception as e:
		raise HTTPException(status_code=500, detail=str(e))

@app.post("/export")
async def export_file(payload: dict):
	ct = payload.get('corrected_text')
	if not ct:
		raise HTTPException(status_code=400, detail="Se requiere corrected_text")
	filename = payload.get('filename','export_corrigido.txt')
	# Normalizar line endings a Windows y codificar sin BOM
	ct_normalized = ct.replace('\r\n', '\n').replace('\n', '\r\n')
	return StreamingResponse(
		io.BytesIO(ct_normalized.encode('utf-8')),
		media_type='text/plain; charset=utf-8',
		headers={"Content-Disposition": f"attachment; filename={filename}"}
	)

@app.post("/export-excel")
async def export_excel(payload: dict):
	try:
		from .excel_export import build_data_excel
	except ImportError:
		from excel_export import build_data_excel
	ct = payload.get('corrected_text')
	template_key = payload.get('template_key', 'gestante')
	if not ct:
		raise HTTPException(status_code=400, detail="Se requiere corrected_text")
	meta = get_template_by_key(template_key)
	buf = build_data_excel(ct, meta["template"])
	filename = payload.get('filename', f'data_corregida_{template_key}.xlsx')
	return StreamingResponse(
		buf,
		media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
		headers={"Content-Disposition": f"attachment; filename={filename}"},
	)

# ─── Cargues (historial mensual) ──────────────────────────────────────────

class CarguePayload(BaseModel):
	corrected_text: str = ""
	raw_text: str = ""
	template_key: str = "gestante"
	filename: str = "cargue.xlsx"
	mes: str = ""
	compressed: bool = False
	summary: dict | None = None
	logs: list | None = None
	row_count: int = 0
	errors_count: int = 0
	corrected_count: int = 0
	quality_percent: float = 0.0
	status: str = "validado"


class RegistroFormularioPayload(BaseModel):
	registro: dict
	template_key: str = "gestante"


def _current_month() -> str:
	now = datetime.utcnow()
	return now.strftime("%Y-%m")


@app.post("/cargues")
async def create_cargue(payload: CarguePayload, current_user: User = Depends(get_current_user)):
	ensure_db_ready()
	db = SessionLocal()
	try:
		prestador = db.query(Prestador).filter(Prestador.user_id == current_user.id).first()

		# Validar IPS: si el prestador tiene IPS asignada, verificar que todas las filas pertenezcan a ella
		# ── Validar cantidad de variables (200 para gestante) ─────────────
		if payload.template_key == "gestante":
			ESPERA_NUM_VARS = 200
			texto_val = payload.corrected_text or payload.raw_text or ""
			if payload.compressed:
				try:
					import base64, gzip
					texto_val = gzip.decompress(base64.b64decode(texto_val)).decode("utf-8", errors="replace")
				except Exception:
					pass
			lineas_val = [l for l in texto_val.strip().split("\n") if l.strip()]
			if lineas_val:
				# Primera linea = header
				header_cols = lineas_val[0].split("|")
				num_vars = len(header_cols)
				if num_vars != ESPERA_NUM_VARS:
					raise HTTPException(
						status_code=400,
						detail=f"El archivo tiene {num_vars} variables pero el instructivo de gestantes requiere exactamente {ESPERA_NUM_VARS}. "
							   f"Verifique que el archivo Excel sea la versión correcta del formato."
					)
				# Validar tambien las filas de datos
				errores_vars = []
				for i, linea in enumerate(lineas_val[1:], start=2):
					cols = linea.split("|")
					if len(cols) != ESPERA_NUM_VARS:
						errores_vars.append(f"Fila {i}: {len(cols)} vars")
						if len(errores_vars) >= 5:
							break
				if errores_vars:
					raise HTTPException(
						status_code=400,
						detail=f"Algunas filas no tienen {ESPERA_NUM_VARS} variables: {'; '.join(errores_vars)}"
					)

		# ── Validar IPS ──────────────────────────────────────────────────
		if prestador and prestador.ips and payload.template_key == "gestante":
			ips_prestador_code = str(prestador.ips).strip()
			ips_prestador_nombre = None
			try:
				from sqlalchemy import text
				with db.bind.connect() as conn:
					row = conn.execute(
						text('SELECT "razon_social" FROM "administrativo"."ct_ips" WHERE "ips" = :cod LIMIT 1'),
						{"cod": ips_prestador_code},
					).fetchone()
					if row:
						ips_prestador_nombre = str(row[0]).strip().upper()
			except Exception:
				pass

			if ips_prestador_nombre:
				IPS_COL_INDEX = 28  # Indice de "Nombre de la IPS Primaria" en instructivo (0-based)
				texto_raw = payload.corrected_text or payload.raw_text or ""
				if payload.compressed:
					try:
						import base64, gzip
						texto_raw = gzip.decompress(base64.b64decode(texto_raw)).decode("utf-8", errors="replace")
					except Exception:
						pass
				lineas = [l for l in texto_raw.strip().split("\n") if l.strip()]
				if len(lineas) > 1:
					errores_ips = []
					for i, linea in enumerate(lineas[1:], start=2):
						cols = linea.split("|")
						if len(cols) > IPS_COL_INDEX:
							ips_fila = cols[IPS_COL_INDEX].strip().upper()
							if ips_fila and ips_fila != ips_prestador_nombre and ips_fila != "NA":
								errores_ips.append(f"Fila {i}: {ips_fila}")
					if errores_ips:
						raise HTTPException(
							status_code=400,
							detail=f"No se puede guardar: {len(errores_ips)} usuaria(s) pertenecen a otra IPS. "
								   f"Tu IPS asignada es: {ips_prestador_nombre}. "
								   f"Usuarias con otra IPS: {'; '.join(errores_ips[:5])}"
								   f"{'...' if len(errores_ips) > 5 else ''}"
						)

		cargue = Cargue(
			prestador_id=prestador.id if prestador else None,
			user_id=current_user.id,
			template_key=payload.template_key,
			mes=payload.mes or _current_month(),
			original_filename=payload.filename,
			raw_text=payload.raw_text,
			corrected_text=payload.corrected_text,
			compressed=bool(payload.compressed),
			summary_json=json.dumps(payload.summary) if payload.summary else None,
			logs_json=json.dumps(payload.logs) if payload.logs else None,
			row_count=payload.row_count,
			errors_count=payload.errors_count,
			corrected_count=payload.corrected_count,
			quality_percent=payload.quality_percent,
			status=payload.status if payload.status in ("validado", "con_errores") else "validado",
		)
		db.add(cargue)
		db.commit()
		db.refresh(cargue)

		# ── Poblar tabla gestantes ────────────────────────────────────────
		gestantes_insertadas = 0
		gestantes_errores = []
		if payload.template_key == "gestante" and payload.corrected_text:
			try:
				# Asegurar que la tabla gestantes existe
				try:
					crear_tabla_gestantes()
				except Exception:
					pass

				texto_cargue = payload.corrected_text
				# Descomprimir si viene comprimido (base64+gzip)
				if payload.compressed:
					try:
						import base64, gzip
						raw_bytes = base64.b64decode(texto_cargue)
						texto_cargue = gzip.decompress(raw_bytes).decode("utf-8", errors="replace")
					except Exception as e_decomp:
						gestantes_errores.append(f"Error descomprimiendo: {str(e_decomp)[:200]}")

				lineas = [l for l in texto_cargue.strip().split("\n") if l.strip()]
				gestantes_errores.append(f"DEBUG: {len(lineas)} lineas, compressed={payload.compressed}, texto_len={len(texto_cargue)}")

				if len(lineas) > 0:
					db_cols = GESTANTE_COLUMNS

					# Obtener columnas reales de la tabla gestantes
					try:
						real_cols_row = db.execute(text('SELECT * FROM gestantes WHERE 1=0')).cursor.description
						real_cols = [c.name for c in real_cols_row]
					except Exception as e_cols:
						gestantes_errores.append(f"Error leyendo columnas de gestantes: {str(e_cols)[:200]}")
						real_cols = []

					for idx, linea in enumerate(lineas):
						cols = linea.split("|")
						if len(cols) < 3:
							continue

						# Construir dict mapeando cada columna del template a su valor
						registro = {}
						for ci in range(min(len(db_cols), len(cols) - 1)):
							registro[db_cols[ci]] = cols[ci + 1].strip()

						# Asignar metadata del cargue
						registro["prestador_id"] = str(prestador.id) if prestador else ""
						registro["user_id"] = str(current_user.id)
						registro["mes"] = cargue.mes
						registro["original_filename"] = cargue.original_filename

						try:
							# Usar SOLO columnas que existen en la tabla real
							if real_cols:
								cols_validas = [c for c in real_cols if c in registro and c != "id"]
							else:
								cols_validas = [c for c in db_cols if c in registro]
							if not cols_validas:
								gestantes_errores.append(f"Fila {idx+1}: no hay columnas validas")
								continue

							placeholders = ", ".join(f':{c}' for c in cols_validas)
							col_names_sql = ", ".join(f'"{c}"' for c in cols_validas)
							params = {c: str(registro.get(c, "")) for c in cols_validas}
							insert_sql = f'INSERT INTO gestantes ({col_names_sql}) VALUES ({placeholders})'
							db.execute(text(insert_sql), params)
							gestantes_insertadas += 1
						except Exception as e_insert:
							gestantes_errores.append(f"Fila {idx+1}: {str(e_insert)[:150]}")

					db.commit()
			except Exception as e_parse:
				gestantes_errores.append(f"Error general: {str(e_parse)[:300]}")

		return {
			"id": cargue.id,
			"mes": cargue.mes,
			"template_key": cargue.template_key,
			"gestantes_insertadas": gestantes_insertadas,
			"gestantes_errores": gestantes_errores[:20],
			"total_errores": len(gestantes_errores),
		}
	except OperationalError:
		raise HTTPException(status_code=503, detail="No se pudo conectar a la base de datos para guardar el cargue. Verifica la conexión al servidor PostgreSQL.")
	finally:
		db.close()


UNIFICADO_MES = "UNIFICADO"


# Bloques del formulario: nombre + rango de indices (0-based) en el template gestante
FORMULARIO_BLOQUES = [
    {"id": "datos", "titulo": "Datos personales", "descripcion": "Identificación y datos generales de la gestante", "inicio": 0, "fin": 28},
    {"id": "control", "titulo": "Control prenatal y antecedentes", "descripcion": "Fechas de control, FUM y fórmula obstétrica", "inicio": 29, "fin": 43},
    {"id": "riesgos", "titulo": "Antecedentes y riesgos", "descripcion": "Enfermedades, eventos obstétricos, peso, talla y riesgos psicosociales", "inicio": 44, "fin": 71},
    {"id": "vih", "titulo": "Tamizajes VIH y Sífilis", "descripcion": "Asesoría, fechas y resultados de pruebas de VIH y sífilis", "inicio": 72, "fin": 98},
    {"id": "laboratorio", "titulo": "Laboratorios", "descripcion": "Urocultivo, glicemia, hemoglobina, grupo RH, hepatitis, toxoplasma y otros", "inicio": 99, "fin": 128},
    {"id": "vacunas", "titulo": "Vacunas y ecografías", "descripcion": "Aplicación de vacunas, ecografías y suministro de micronutrientes", "inicio": 129, "fin": 143},
    {"id": "controles", "titulo": "Controles prenatales", "descripcion": "Fecha de cada control y quién lo realizó, más datos del último control", "inicio": 144, "fin": 170},
    {"id": "especialistas", "titulo": "Atención especializada", "descripcion": "Consultas de ginecología, nutrición, psicología y otros especialistas", "inicio": 171, "fin": 177},
    {"id": "eventos", "titulo": "Eventos obstétricos", "descripcion": "Aborto, parto, complicaciones, defunción y planificación familiar", "inicio": 178, "fin": 197},
]


@app.get("/cargue-unificado/estructura")
async def estructura_formulario(
	template_key: str = Query("gestante"),
	current_user: User = Depends(get_current_user),
):
	"""Devuelve la estructura del formulario por bloques con los campos y sus opciones."""
	meta = get_template_by_key(template_key)
	tmpl = meta["template"]
	bloques = []
	for b in FORMULARIO_BLOQUES:
		campos = []
		for i in range(b["inicio"], min(b["fin"], len(tmpl)) + 1):
			if i >= len(tmpl):
				continue
			x = tmpl[i]
			campos.append({
				"name": x["name"],
				"type": x["type"],
				"allowed": x.get("allowed"),
			})
		bloques.append({
			"id": b["id"],
			"titulo": b["titulo"],
			"descripcion": b["descripcion"],
			"campos": campos,
		})
	return {"template_key": template_key, "bloques": bloques}


def _get_or_create_cargue_unificado(db, current_user: User, template_key: str) -> Cargue:
	"""Retorna el cargue unico del prestador para la plantilla, creandolo si no existe."""
	prestador = db.query(Prestador).filter(Prestador.user_id == current_user.id).first()
	cargue = (
		db.query(Cargue)
		.filter(
			Cargue.user_id == current_user.id,
			Cargue.template_key == template_key,
			Cargue.mes == UNIFICADO_MES,
		)
		.first()
	)
	if cargue is None:
		cargue = Cargue(
			prestador_id=prestador.id if prestador else None,
			user_id=current_user.id,
			template_key=template_key,
			mes=UNIFICADO_MES,
			original_filename="cargue_unificado.txt",
			raw_text="",
			corrected_text="",
			compressed=False,
			summary_json=None,
			logs_json=None,
			row_count=0,
			errors_count=0,
			corrected_count=0,
			quality_percent=100.0,
			status="validado",
		)
		db.add(cargue)
		db.commit()
		db.refresh(cargue)
	return cargue


def _filas_cargue(cargue: Cargue) -> list[str]:
	"""Devuelve las lineas (filas) pipe-delimited del cargue."""
	text = _decompress_cargue(cargue) if cargue.compressed else (cargue.corrected_text or "")
	lines = []
	for line in text.replace("\r\n", "\n").split("\n"):
		if line.strip():
			lines.append(line)
	return lines


def _guardar_filas(db, cargue: Cargue, filas: list[str]):
	text = "\n".join(filas)
	cargue.corrected_text = text
	cargue.raw_text = text
	cargue.compressed = False
	cargue.row_count = len(filas)
	cargue.quality_percent = 100.0
	cargue.status = "validado"
	db.commit()


@app.get("/cargue-unificado")
async def get_cargue_unificado(
	template_key: str = Query("gestante"),
	current_user: User = Depends(get_current_user),
):
	"""Devuelve el cargue unificado del prestador con sus registros en formato filas."""
	ensure_db_ready()
	db = SessionLocal()
	try:
		cargue = _get_or_create_cargue_unificado(db, current_user, template_key)
		filas = _filas_cargue(cargue)
		# Devolver las filas como lista de dicts usando los nombres del template
		meta = get_template_by_key(template_key)
		tmpl = meta["template"]
		nombres = [t["name"] for t in tmpl]
		registros = []
		for line in filas:
			cells = line.split("|")
			rec = {}
			for i, name in enumerate(nombres):
				rec[name] = cells[i] if i < len(cells) else ""
			registros.append(rec)
		return {
			"cargue_id": cargue.id,
			"template_key": template_key,
			"registros": registros,
			"row_count": len(registros),
		}
	except OperationalError:
		raise HTTPException(status_code=503, detail="No se pudo conectar a la base de datos para consultar el cargue unificado.")
	finally:
		db.close()


@app.post("/cargue-unificado/registro")
async def agregar_registro_unificado(
	payload: RegistroFormularioPayload,
	current_user: User = Depends(get_current_user),
):
	"""Agrega un registro (fila) al cargue unificado del prestador. Los campos
	se validan contra el template del instructivo antes de guardar."""
	ensure_db_ready()
	db = SessionLocal()
	try:
		meta = get_template_by_key(payload.template_key)
		tmpl = meta["template"]
		nombres = [t["name"] for t in tmpl]

		# Construir la fila en orden de plantilla
		celdas = []
		errores = []
		for i, tdef in enumerate(tmpl):
			name = tdef["name"]
			val = payload.registro.get(name, "")
			val = "" if val is None else str(val).strip()
			# Formula: se recalcula, se guarda vacio si no viene
			if tdef["type"] == "FORMULA":
				celdas.append("")
				continue
			# Campos obligatorios minimos
			if i == 0 and not val:
				errores.append("El campo 'No' (consecutivo) es obligatorio")
			if i == 2 and not val:
				errores.append("El campo 'No. De Identificación' es obligatorio")
			if i == 3 and not val:
				errores.append("El campo 'Apellido_1' es obligatorio")
			if i == 5 and not val:
				errores.append("El campo 'Nombre_1' es obligatorio")
			celdas.append(val)

		if errores:
			raise HTTPException(status_code=400, detail="; ".join(errores))

		# Los campos vacios se dejan vacios (el formulario solo exige los obligatorios).
		# Se validan solo los campos que el prestador realmente ingreso.

		# Validar cada campo con valor contra el template (mismos criterios del instructivo)
		errores_val = []
		for i, tdef in enumerate(tmpl):
			val = celdas[i]
			if tdef["type"] == "FORMULA" or val == "":
				continue
			df_one = pd.DataFrame([[val]])
			df_one.columns = [nombres[i]]
			mapping_one = {nombres[i]: nombres[i]}
			try:
				from .validators import validate_only
			except ImportError:
				from validators import validate_only
			res = validate_only(df_one, mapping_one, [tdef])
			if res["stats"]["rows_with_errors"] > 0:
				detalles = [l["corrected"] for l in res["logs"][:3]]
				errores_val.append(f"{tdef['name']}: {'; '.join(detalles)}")
		if errores_val:
			raise HTTPException(status_code=400, detail="El registro tiene errores de validación: " + " | ".join(errores_val[:10]))

		# Validar IPS del registro contra IPS del prestador
		prestador = db.query(Prestador).filter(Prestador.user_id == current_user.id).first()
		if prestador and prestador.ips and payload.template_key == "gestante":
			ips_prestador_code = str(prestador.ips).strip()
			ips_prestador_nombre = None
			try:
				from sqlalchemy import text as sa_text
				with db.bind.connect() as conn:
					row = conn.execute(
						sa_text('SELECT "razon_social" FROM "administrativo"."ct_ips" WHERE "ips" = :cod LIMIT 1'),
						{"cod": ips_prestador_code},
					).fetchone()
					if row:
						ips_prestador_nombre = str(row[0]).strip().upper()
			except Exception:
				pass
			if ips_prestador_nombre:
				ips_registro = celdas[28].strip().upper() if len(celdas) > 28 else ""
				if ips_registro and ips_registro != ips_prestador_nombre and ips_registro != "NA":
					raise HTTPException(
						status_code=400,
						detail=f"No se puede guardar: esta gestante pertenece a otra IPS ({celdas[28]}). "
							   f"Tu IPS asignada es: {ips_prestador_nombre}."
					)

		# Aplicar formulas al registro
		df = pd.DataFrame([celdas])
		df.columns = nombres
		try:
			from .formulas import aplicar_formulas_df
		except ImportError:
			from formulas import aplicar_formulas_df
		try:
			df = aplicar_formulas_df(df)
		except Exception:
			pass

		# Guardar como fila en el cargue unificado
		cargue = _get_or_create_cargue_unificado(db, current_user, payload.template_key)
		filas = _filas_cargue(cargue)
		line = df.iloc[0].astype(str).str.replace("|", " ").str.replace("\r", " ").str.replace("\n", " ").tolist()
		filas.append("|".join(line))
		_guardar_filas(db, cargue, filas)

		return {
			"success": True,
			"cargue_id": cargue.id,
			"row_count": len(filas),
			"registro": {n: v for n, v in zip(nombres, line)},
		}
	except OperationalError:
		raise HTTPException(status_code=503, detail="No se pudo conectar a la base de datos para guardar el registro.")
	finally:
		db.close()


@app.delete("/cargue-unificado/registro/{indice}")
async def eliminar_registro_unificado(
	indice: int,
	template_key: str = Query("gestante"),
	current_user: User = Depends(get_current_user),
):
	"""Elimina un registro (fila) del cargue unificado por su indice (1-based)."""
	ensure_db_ready()
	db = SessionLocal()
	try:
		cargue = _get_or_create_cargue_unificado(db, current_user, template_key)
		filas = _filas_cargue(cargue)
		if indice < 1 or indice > len(filas):
			raise HTTPException(status_code=404, detail="Registro no encontrado")
		del filas[indice - 1]
		_guardar_filas(db, cargue, filas)
		return {"success": True, "row_count": len(filas)}
	except OperationalError:
		raise HTTPException(status_code=503, detail="No se pudo conectar a la base de datos para eliminar el registro.")
	finally:
		db.close()


@app.get("/cargues")
async def list_cargues(
    template_key: str = Query(""),
    current_user: User = Depends(get_current_user),
):
	ensure_db_ready()
	db = SessionLocal()
	try:
		q = db.query(Cargue)
		if current_user.role == "prestador":
			q = q.filter(Cargue.user_id == current_user.id)
		elif current_user.role == "lider":
			# El líder ve todos los cargues de su plantilla asignada
			prestador = db.query(Prestador).filter(Prestador.user_id == current_user.id).first()
			assigned = [p.template_key for p in (prestador.plantillas if prestador else [])]
			if assigned:
				q = q.filter(Cargue.template_key.in_(assigned))
			else:
				q = q.filter(Cargue.id == -1)
		if template_key:
			q = q.filter(Cargue.template_key == template_key)
		# Regla "todo o nada": solo mostrar cargues validados (sin errores).
		q = q.filter(Cargue.status == "validado")
		q = q.order_by(Cargue.created_at.desc())
		items = q.limit(100).all()
		result = []
		for c in items:
			summary = {}
			try:
				summary = json.loads(c.summary_json) if c.summary_json else {}
			except Exception:
				pass
			result.append({
				"id": c.id,
				"template_key": c.template_key,
				"mes": c.mes,
				"original_filename": c.original_filename,
				"prestador": (c.prestador.nombre if c.prestador else None) or (c.user.name if c.user else None),
				"row_count": c.row_count,
				"errors_count": c.errors_count,
				"corrected_count": c.corrected_count,
				"quality_percent": c.quality_percent,
				"status": c.status,
				"created_at": c.created_at.isoformat() if c.created_at else None,
				"summary": summary,
			})
		return {"cargues": result}
	except OperationalError:
		raise HTTPException(status_code=503, detail="No se pudo conectar a la base de datos para consultar el historial. Verifica la conexión al servidor PostgreSQL.")
	finally:
		db.close()


@app.get("/cargues/{cargue_id}")
async def get_cargue(cargue_id: int, current_user: User = Depends(get_current_user)):
	ensure_db_ready()
	db = SessionLocal()
	try:
		cargue = db.get(Cargue, cargue_id)
		if cargue is None:
			raise HTTPException(status_code=404, detail="Cargue no encontrado")
		if current_user.role == "prestador" and cargue.user_id != current_user.id:
			raise HTTPException(status_code=403, detail="No autorizado")
		if current_user.role == "lider":
			_prest = db.query(Prestador).filter(Prestador.user_id == current_user.id).first()
			_assigned = [p.template_key for p in (_prest.plantillas if _prest else [])]
			if cargue.template_key not in _assigned:
				raise HTTPException(status_code=403, detail="No autorizado")
		def _parse_json(value):
			try:
				return json.loads(value) if value else {}
			except Exception:
				return {}
		return {
			"id": cargue.id,
			"template_key": cargue.template_key,
			"mes": cargue.mes,
			"original_filename": cargue.original_filename,
			"corrected_text": cargue.corrected_text,
			"raw_text": cargue.raw_text,
			"compressed": cargue.compressed,
			"row_count": cargue.row_count,
			"errors_count": cargue.errors_count,
			"corrected_count": cargue.corrected_count,
			"quality_percent": cargue.quality_percent,
			"summary": _parse_json(cargue.summary_json),
			"logs_sample": _parse_json(cargue.logs_json),
			"created_at": cargue.created_at.isoformat() if cargue.created_at else None,
		}
	except OperationalError:
		raise HTTPException(status_code=503, detail="No se pudo conectar a la base de datos. Verifica la conexión al servidor PostgreSQL.")
	finally:
		db.close()


def _decompress_cargue(cargue) -> str:
	ct = cargue.corrected_text or ""
	if cargue.compressed and ct:
		try:
			return gzip.decompress(base64.b64decode(ct)).decode("utf-8")
		except Exception:
			return ct
	return ct


def _normalizar_fechas_texto(text: str, template_key: str) -> str:
	"""Parsea un texto pipe-delimited y normaliza las fechas de las columnas
	tipo DATE a AAAA-MM-DD (ej: 3/04/2000 -> 2000-04-03). Usado al descargar
	para que cualquier cargue guardado salga siempre con fechas ISO y sin
	caracteres que rompan la estructura pipe-delimited."""
	from validators import to_date_iso
	try:
		meta = get_template_by_key(template_key)
		tmpl = meta["template"]
		# Indices de columnas DATE del template (para normalizar SOLO fechas)
		date_idx = {i for i, t in enumerate(tmpl) if t["type"] == "DATE"}
		lines = text.replace("\r\n", "\n").split("\n")
		out = []
		for line in lines:
			if not line.strip():
				continue
			parts = line.split("|")
			for i in range(len(parts)):
				val = parts[i].strip()
				# Limpiar caracteres que rompen el formato pipe-delimited
				val = val.replace("|", " ").replace("\r", " ").replace("\n", " ").replace("\t", " ")
				val = val.replace("_x000D_", " ").replace("_x000B_", " ")
				val = re.sub(r"\s+", " ", val).strip()
				# Normalizar SOLO columnas de fecha (nunca numeros/EDAD)
				if i in date_idx and val and val.upper() not in ("SIN DATO", "SIN DATOS", "N/A", "NONE", "NAN", "NULL", "NA"):
					iso = to_date_iso(val)
					if iso:
						val = iso
				parts[i] = val
			out.append("|".join(parts))
		return "\r\n".join(out)
	except Exception:
		return text.replace('\r\n', '\n').replace('\n', '\r\n')


@app.get("/cargues/{cargue_id}/download-txt")
async def download_cargue_txt(cargue_id: int, current_user: User = Depends(get_current_user)):
	ensure_db_ready()
	db = SessionLocal()
	try:
		cargue = db.get(Cargue, cargue_id)
		if cargue is None:
			raise HTTPException(status_code=404, detail="Cargue no encontrado")
		if current_user.role == "prestador" and cargue.user_id != current_user.id:
			raise HTTPException(status_code=403, detail="No autorizado")
		if current_user.role == "lider":
			_prest = db.query(Prestador).filter(Prestador.user_id == current_user.id).first()
			_assigned = [p.template_key for p in (_prest.plantillas if _prest else [])]
			if cargue.template_key not in _assigned:
				raise HTTPException(status_code=403, detail="No autorizado")
		text = _decompress_cargue(cargue)
		text = _normalizar_fechas_texto(text, cargue.template_key or "gestante")
		filename = cargue.original_filename.replace(".xlsx", "_corregido.txt").replace(".xls", "_corregido.txt")
		text_normalized = text.replace('\r\n', '\n').replace('\n', '\r\n')
		return StreamingResponse(
			io.BytesIO(text_normalized.encode('utf-8')),
			media_type='text/plain; charset=utf-8',
			headers={"Content-Disposition": f"attachment; filename={filename}"},
		)
	except OperationalError:
		raise HTTPException(status_code=503, detail="No se pudo conectar a la base de datos. Verifica la conexión al servidor PostgreSQL.")
	finally:
		db.close()


@app.get("/cargues/{cargue_id}/download-excel")
async def download_cargue_excel(cargue_id: int, current_user: User = Depends(get_current_user)):
	try:
		from .excel_export import build_data_excel
	except ImportError:
		from excel_export import build_data_excel
	ensure_db_ready()
	db = SessionLocal()
	try:
		cargue = db.get(Cargue, cargue_id)
		if cargue is None:
			raise HTTPException(status_code=404, detail="Cargue no encontrado")
		if current_user.role == "prestador" and cargue.user_id != current_user.id:
			raise HTTPException(status_code=403, detail="No autorizado")
		if current_user.role == "lider":
			_prest = db.query(Prestador).filter(Prestador.user_id == current_user.id).first()
			_assigned = [p.template_key for p in (_prest.plantillas if _prest else [])]
			if cargue.template_key not in _assigned:
				raise HTTPException(status_code=403, detail="No autorizado")
		text = _decompress_cargue(cargue)
		meta = get_template_by_key(cargue.template_key)
		buf = build_data_excel(text, meta["template"])
		filename = cargue.original_filename.replace(".xlsx", "_ajustada.xlsx").replace(".xls", "_ajustada.xlsx")
		return StreamingResponse(
			buf,
			media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
			headers={"Content-Disposition": f"attachment; filename={filename}"},
		)
	except OperationalError:
		raise HTTPException(status_code=503, detail="No se pudo conectar a la base de datos. Verifica la conexión al servidor PostgreSQL.")
	finally:
		db.close()


# ─── Historias clínicas (PDF) ─────────────────────────────────────────────

MAX_PDF_MB = 20


@app.post("/historias")
async def upload_historia(
    request: Request,
    file: UploadFile = File(...),
    paciente_documento: str = Form(""),
    paciente_nombre: str = Form(""),
    template_key: str = Form("gestante"),
    current_user: User = Depends(get_current_user),
):
    ensure_db_ready()
    filename = (file.filename or "historia.pdf").strip()
    content = file.file.read()
    if len(content) == 0:
        raise HTTPException(status_code=400, detail="El archivo está vacío.")
    if not filename.lower().endswith(".pdf"):
        raise HTTPException(status_code=400, detail="Solo se permiten archivos PDF.")
    if len(content) > MAX_PDF_MB * 1024 * 1024:
        raise HTTPException(
            status_code=400,
            detail=f"El archivo supera {MAX_PDF_MB} MB. En el entorno desplegado el límite de subida es de ~4.5 MB; divide la historia en partes menores.",
        )
    oidc_token = request.headers.get("x-vercel-oidc-token", "")
    db = SessionLocal()
    try:
        prestador = db.query(Prestador).filter(Prestador.user_id == current_user.id).first()
        historia = HistoriaClinica(
            prestador_id=prestador.id if prestador else None,
            user_id=current_user.id,
            template_key=template_key.strip() or "gestante",
            paciente_documento=paciente_documento.strip(),
            paciente_nombre=paciente_nombre.strip(),
            filename=filename,
            content_type=file.content_type or "application/pdf",
            file_size=len(content),
        )
        db.add(historia)
        db.flush()  # genera el id
        if gcs_storage.gcs_enabled():
            safe_name = re.sub(r"[^A-Za-z0-9._-]", "_", filename)
            blob_path = f"historias/{historia.id}/{safe_name}"
            gcs_storage.upload_pdf(oidc_token, blob_path, content, historia.content_type)
            historia.pdf_path = blob_path
        else:
            historia.pdf_data = content
        db.commit()
        db.refresh(historia)
        storage_used = "gcs" if historia.pdf_path else "db"
        return {
            "id": historia.id,
            "filename": historia.filename,
            "paciente_nombre": historia.paciente_nombre,
            "storage": storage_used,
        }
    except OperationalError:
        raise HTTPException(status_code=503, detail="No se pudo conectar a la base de datos para guardar la historia clínica. Verifica la conexión al servidor PostgreSQL.")
    except Exception as exc:
        db.rollback()
        detail = "No se pudo guardar el PDF en el almacenamiento. Verifica la configuración de Google Cloud Storage."
        if os.environ.get("DEBUG_GCS", ""):
            detail += f" Detalle: {exc} | oidc_len={len(oidc_token)}"
        raise HTTPException(status_code=500, detail=detail)
    finally:
        db.close()


@app.get("/historias")
async def list_historias(
    q: str = Query(""),
    template_key: str = Query(""),
    current_user: User = Depends(get_current_user),
):
    ensure_db_ready()
    db = SessionLocal()
    try:
        query = db.query(HistoriaClinica)
        if current_user.role == "prestador":
            query = query.filter(HistoriaClinica.user_id == current_user.id)
        elif current_user.role == "lider":
            prestador = db.query(Prestador).filter(Prestador.user_id == current_user.id).first()
            assigned = [p.template_key for p in (prestador.plantillas if prestador else [])]
            if assigned:
                query = query.filter(HistoriaClinica.template_key.in_(assigned))
            else:
                query = query.filter(HistoriaClinica.id == -1)
        if template_key:
            query = query.filter(HistoriaClinica.template_key == template_key)
        query = query.order_by(HistoriaClinica.created_at.desc())
        items = query.limit(500).all()
        term = q.strip().lower()
        result = []
        for h in items:
            if term and term not in f"{h.filename} {h.paciente_nombre or ''} {h.paciente_documento or ''}".lower():
                continue
            result.append({
                "id": h.id,
                "template_key": h.template_key or "gestante",
                "prestador": (h.prestador.nombre if h.prestador else None) or (h.user.name if h.user else None),
                "paciente_nombre": h.paciente_nombre,
                "paciente_documento": h.paciente_documento,
                "filename": h.filename,
                "file_size": h.file_size,
                "created_at": h.created_at.isoformat() if h.created_at else None,
            })
        return {"historias": result}
    except OperationalError:
        raise HTTPException(status_code=503, detail="No se pudo conectar a la base de datos para consultar las historias clínicas. Verifica la conexión al servidor PostgreSQL.")
    finally:
        db.close()


@app.get("/historias/{historia_id}")
async def get_historia(request: Request, historia_id: int, current_user: User = Depends(get_current_user)):
    ensure_db_ready()
    db = SessionLocal()
    try:
        h = db.get(HistoriaClinica, historia_id)
        if h is None:
            raise HTTPException(status_code=404, detail="Historia clínica no encontrada")
        if current_user.role == "prestador" and h.user_id != current_user.id:
            raise HTTPException(status_code=403, detail="No autorizado")
        if current_user.role == "lider":
            _prest = db.query(Prestador).filter(Prestador.user_id == current_user.id).first()
            _assigned = [p.template_key for p in (_prest.plantillas if _prest else [])]
            if (h.template_key or "gestante") not in _assigned:
                raise HTTPException(status_code=403, detail="No autorizado")
        filename = h.filename or f"historia_{h.id}.pdf"
        if h.pdf_path:
            oidc_token = request.headers.get("x-vercel-oidc-token", "")
            try:
                data = gcs_storage.download_pdf(oidc_token, h.pdf_path)
            except Exception:
                raise HTTPException(status_code=502, detail="No se pudo leer el PDF desde el almacenamiento. Verifica la configuración de Google Cloud Storage.")
        elif h.pdf_data:
            data = bytes(h.pdf_data)
        else:
            raise HTTPException(status_code=404, detail="El archivo PDF no está disponible")
        return StreamingResponse(
            io.BytesIO(data),
            media_type=h.content_type or "application/pdf",
            headers={"Content-Disposition": f'inline; filename="{filename}"'},
        )
    except OperationalError:
        raise HTTPException(status_code=503, detail="No se pudo conectar a la base de datos. Verifica la conexión al servidor PostgreSQL.")
    finally:
        db.close()


@app.delete("/historias/{historia_id}")
async def delete_historia(request: Request, historia_id: int, current_user: User = Depends(get_current_user)):
    ensure_db_ready()
    db = SessionLocal()
    try:
        h = db.get(HistoriaClinica, historia_id)
        if h is None:
            raise HTTPException(status_code=404, detail="Historia clínica no encontrada")
        if current_user.role == "prestador" and h.user_id != current_user.id:
            raise HTTPException(status_code=403, detail="No autorizado")
        if current_user.role == "lider":
            _prest = db.query(Prestador).filter(Prestador.user_id == current_user.id).first()
            _assigned = [p.template_key for p in (_prest.plantillas if _prest else [])]
            if (h.template_key or "gestante") not in _assigned:
                raise HTTPException(status_code=403, detail="No autorizado")
        if h.pdf_path:
            try:
                oidc_token = request.headers.get("x-vercel-oidc-token", "")
                gcs_storage.delete_pdf(oidc_token, h.pdf_path)
            except Exception:
                pass  # no bloquear el borrado si el objeto no existe
        db.delete(h)
        db.commit()
        return {"ok": True}
    except OperationalError:
        raise HTTPException(status_code=503, detail="No se pudo conectar a la base de datos. Verifica la conexión al servidor PostgreSQL.")
    finally:
        db.close()


# ─── Gestión de prestadores (admin / EPS) ─────────────────────────────────

class PrestadorPayload(BaseModel):
	nombre: str
	nit: str = ""
	ips: str = ""
	municipio: str = ""
	username: str
	password: str
	template_key: str = "gestante"
	role: str = "prestador"  # "prestador" | "lider"


@app.post("/admin/prestadores")
async def create_prestador(payload: PrestadorPayload, admin: User = Depends(require_admin)):
	ensure_db_ready()
	db = SessionLocal()
	try:
		exists = db.query(User).filter(User.username == payload.username.strip()).first()
		if exists:
			raise HTTPException(status_code=400, detail="El nombre de usuario ya existe")
		role = payload.role if payload.role in ("prestador", "lider") else "prestador"
		user = User(
			username=payload.username.strip(),
			password_hash=hash_password(payload.password),
			name=payload.nombre,
			role=role,
			active=True,
		)
		db.add(user)
		db.flush()
		prestador = Prestador(
			user_id=user.id,
			nombre=payload.nombre,
			nit=payload.nit,
			ips=payload.ips,
			municipio=payload.municipio,
		)
		db.add(prestador)
		db.flush()
		pp = PrestadorPlantilla(prestador_id=prestador.id, template_key=payload.template_key)
		db.add(pp)
		db.commit()
		return {"id": prestador.id, "username": user.username, "nombre": prestador.nombre, "ips": prestador.ips, "template_key": payload.template_key, "role": role}
	except OperationalError:
		raise HTTPException(status_code=503, detail="No se pudo conectar a la base de datos. Verifica la conexión al servidor PostgreSQL.")
	except HTTPException:
		raise
	except Exception as e:
		raise HTTPException(status_code=500, detail=f"Error al crear el usuario: {str(e)[:300]}")
	finally:
		db.close()


@app.get("/admin/prestadores")
async def list_prestadores(admin: User = Depends(require_admin)):
	ensure_db_ready()
	db = SessionLocal()
	try:
		items = db.query(Prestador).order_by(Prestador.nombre.asc()).all()
		result = []
		for p in items:
			cargues_count = db.query(Cargue).filter(Cargue.prestador_id == p.id).count()
			plantillas = [pp.template_key for pp in p.plantillas]
			result.append({
				"id": p.id,
				"nombre": p.nombre,
				"nit": p.nit,
				"ips": p.ips,
				"municipio": p.municipio,
				"username": p.user.username if p.user else None,
				"cargues_count": cargues_count,
				"template_key": plantillas[0] if plantillas else "gestante",
				"role": p.user.role if p.user else "prestador",
				"permissions": p.permissions or {},
			})
		return {"prestadores": result}
	except OperationalError:
		raise HTTPException(status_code=503, detail="No se pudo conectar a la base de datos. Verifica la conexión al servidor PostgreSQL.")
	finally:
		db.close()


@app.put("/admin/prestadores/{prestador_id}/permissions")
async def update_prestador_permissions(prestador_id: int, payload: dict, admin: User = Depends(require_admin)):
	"""Actualiza los permisos de un prestador."""
	ensure_db_ready()
	db = SessionLocal()
	try:
		prestador = db.query(Prestador).filter(Prestador.id == prestador_id).first()
		if not prestador:
			raise HTTPException(status_code=404, detail="Prestador no encontrado")
		prestador.permissions = payload
		db.commit()
		return {"success": True, "permissions": prestador.permissions}
	except OperationalError:
		raise HTTPException(status_code=503, detail="No se pudo conectar a la base de datos.")
	finally:
		db.close()


@app.put("/admin/prestadores/{prestador_id}")
async def update_prestador(prestador_id: int, payload: dict, admin: User = Depends(require_admin)):
	"""Actualiza datos de un prestador (nombre, IPS, NIT, municipio, etc)."""
	ensure_db_ready()
	db = SessionLocal()
	try:
		prestador = db.query(Prestador).filter(Prestador.id == prestador_id).first()
		if not prestador:
			raise HTTPException(status_code=404, detail="Prestador no encontrado")
		if "nombre" in payload:
			prestador.nombre = payload["nombre"]
			if prestador.user:
				prestador.user.name = payload["nombre"]
		if "nit" in payload:
			prestador.nit = payload["nit"]
		if "ips" in payload:
			prestador.ips = payload["ips"]
		if "municipio" in payload:
			prestador.municipio = payload["municipio"]
		if "role" in payload and prestador.user:
			if payload["role"] in ("prestador", "lider"):
				prestador.user.role = payload["role"]
		db.commit()
		return {"success": True}
	except OperationalError:
		raise HTTPException(status_code=503, detail="No se pudo conectar a la base de datos.")
	finally:
		db.close()


@app.get("/auth/permissions")
async def get_my_permissions(current_user: User = Depends(get_current_user)):
	"""Devuelve los permisos del prestador actual."""
	DEFAULT_PERMISSIONS = {
		"cargue_masivo": True,
		"historias_clinicas": True,
		"ver_historial": True,
		"verificar_afiliado": True,
		"formulario_registro": True,
	}
	try:
		db = SessionLocal()
		try:
			prestador = db.query(Prestador).filter(Prestador.user_id == current_user.id).first()
			if current_user.role == "admin":
				return {"permissions": {k: True for k in DEFAULT_PERMISSIONS}}
			if prestador and prestador.permissions:
				merged = {**DEFAULT_PERMISSIONS, **prestador.permissions}
				return {"permissions": merged}
			return {"permissions": DEFAULT_PERMISSIONS}
		finally:
			db.close()
	except Exception:
		return {"permissions": DEFAULT_PERMISSIONS}


# ─── Consolidación de cargues (admin / EPS) ───────────────────────────────

@app.post("/consolidate")
async def consolidate_cargues(payload: dict, current_user: User = Depends(get_current_user)):
	if current_user.role not in ("admin", "lider"):
		raise HTTPException(status_code=403, detail="Solo el EPS o el líder del programa puede consolidar la data")
	try:
		from .excel_export import build_data_excel
	except ImportError:
		from excel_export import build_data_excel
	template_key = payload.get("template_key", "gestante")
	mes = payload.get("mes", "")
	ensure_db_ready()
	db = SessionLocal()
	try:
		# El líder solo puede consolidar su plantilla asignada
		if current_user.role == "lider":
			prestador = db.query(Prestador).filter(Prestador.user_id == current_user.id).first()
			assigned = [p.template_key for p in (prestador.plantillas if prestador else [])]
			if template_key not in assigned:
				raise HTTPException(status_code=403, detail="No tienes acceso a esta plantilla")
		q = db.query(Cargue).filter(Cargue.template_key == template_key)
		if mes:
			q = q.filter(Cargue.mes == mes)
		q = q.order_by(Cargue.created_at.asc())
		cargues = q.all()
		if not cargues:
			raise HTTPException(status_code=404, detail="No hay cargues para consolidar")
		all_rows = []
		total = 0
		errores = 0
		for c in cargues:
			text = _decompress_cargue(c)
			for line in text.replace("\r\n", "\n").split("\n"):
				if line.strip():
					all_rows.append(line)
			total += c.row_count or 0
			errores += c.errors_count or 0
		corrected_text = "\n".join(all_rows)
		meta = get_template_by_key(template_key)
		buf = build_data_excel(corrected_text, meta["template"])
		filename = f"consolidada_{template_key}{'_'+mes if mes else ''}.xlsx"
		return StreamingResponse(
			buf,
			media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
			headers={"Content-Disposition": f"attachment; filename={filename}"},
		)
	except OperationalError:
		raise HTTPException(status_code=503, detail="No se pudo conectar a la base de datos. Verifica la conexión al servidor PostgreSQL.")
	finally:
		db.close()


@app.post("/evaluate")
async def evaluate_endpoint(payload: dict, format: str = Query(default="json")):
	ct = payload.get('corrected_text', '')
	template_names = payload.get('template_names', [])
	template_key = payload.get('template_key', 'gestante')
	if not ct or not ct.strip():
		raise HTTPException(
			status_code=400,
			detail="No hay datos corregidos para evaluar. Primero carga y valida un archivo."
		)

	try:
		from io import StringIO
		df = pd.read_csv(StringIO(ct), sep='|', header=None, dtype=str, engine='python', keep_default_na=False)
		df = df.fillna('').astype(str)
	except Exception as e:
		raise HTTPException(status_code=400, detail=f"Error al parsear datos: {e}")

	if df.empty:
		raise HTTPException(status_code=400, detail="No hay datos para evaluar")

	# Asignar nombres de columna según template
	if template_names and len(df.columns) == len(template_names):
		df.columns = template_names
	else:
		meta = get_template_by_key(template_key)
		tmpl_names = [t['name'] for t in meta['template']]
		if len(df.columns) == len(tmpl_names):
			df.columns = tmpl_names

	indicators, patients = evaluate(df, template_key)

	if format == "xlsx":
		filename = payload.get('filename', f'evaluacion_{template_key}_{datetime.now().strftime("%Y%m%d")}.xlsx')
		buf = build_evaluation_excel(indicators, patients)
		return StreamingResponse(
			buf,
			media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
			headers={"Content-Disposition": f"attachment; filename={filename}"},
		)

	# JSON response for dashboard
	eval_cols = [c for c in patients.columns if c.startswith("_")]
	data_cols = [c for c in patients.columns if not c.startswith("_")]

	return {
		"indicators": indicators.replace({pd.NA: None}).to_dict(orient="records"),
		"patients": patients.replace({pd.NA: None}).to_dict(orient="records"),
		"eval_columns": eval_cols,
		"data_columns": data_cols,
		"total_patients": len(patients),
		"template_key": template_key,
	}

@app.get("/download-template/{template_key}")
async def download_template(template_key: str):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.worksheet.datavalidation import DataValidation

    meta = get_template_by_key(template_key)
    tmpl = meta["template"]
    headers = [t['name'] for t in tmpl]

    wb = Workbook()
    ws = wb.active
    ws.title = meta["label"]

    # Hidden sheet for long dropdown lists
    helper = wb.create_sheet(title="_Listas")
    helper.sheet_state = "hidden"

    type_fills = {
        "SET":    PatternFill("solid", fgColor="FFF3E0"),
        "INT":    PatternFill("solid", fgColor="E3F2FD"),
        "DECIMAL": PatternFill("solid", fgColor="E8F5E9"),
        "DATE":   PatternFill("solid", fgColor="F3E5F5"),
        "TEXT":   PatternFill("solid", fgColor="F5F5F5"),
    }
    header_fill = PatternFill("solid", fgColor="1B5E20")
    header_font = Font(bold=True, color="FFFFFF", size=10, name="Calibri")
    thin_border = Border(
        left=Side(style="thin", color="BDBDBD"),
        right=Side(style="thin", color="BDBDBD"),
        top=Side(style="thin", color="BDBDBD"),
        bottom=Side(style="thin", color="BDBDBD"),
    )

    # Register unique allowed lists on helper sheet
    list_registry: dict[str, str] = {}
    helper_col = 0

    def _col_letter(n: int) -> str:
        result = ""
        while n > 0:
            n, r = divmod(n - 1, 26)
            result = chr(65 + r) + result
        return result

    def register_list(allowed: list[str]) -> str:
        nonlocal helper_col
        key = ",".join(allowed)
        if key in list_registry:
            return list_registry[key]
        helper_col += 1
        col_letter = _col_letter(helper_col)
        for row_idx, val in enumerate(allowed, start=1):
            helper.cell(row=row_idx, column=helper_col, value=val)
        ref = f"='_Listas'!${col_letter}$1:${col_letter}${len(allowed)}"
        list_registry[key] = ref
        return ref

    for col_idx, f in enumerate(tmpl, start=1):
        col_letter = ws.cell(row=1, column=col_idx).column_letter

        cell = ws.cell(row=1, column=col_idx, value=f['name'])
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(wrap_text=True, vertical="center")
        cell.border = thin_border

        # Fila 2: leyenda del tipo de dato (coloreada), también sirve de área de datos
        cell2 = ws.cell(row=2, column=col_idx)
        cell2.fill = type_fills.get(f['type'], PatternFill())
        cell2.border = thin_border
        cell2.alignment = Alignment(vertical="center")

        # Formato por tipo aplicado a toda el área de datos (filas 2-102)
        if f['type'] == 'DATE':
            for r in range(2, 102):
                c = ws.cell(row=r, column=col_idx)
                c.number_format = 'yyyy-mm-dd'
                c.border = thin_border
        elif f['type'] == 'INT':
            for r in range(2, 102):
                c = ws.cell(row=r, column=col_idx)
                c.number_format = '0'
                c.border = thin_border
        elif f['type'] == 'DECIMAL':
            for r in range(2, 102):
                c = ws.cell(row=r, column=col_idx)
                c.number_format = '0.00'
                c.border = thin_border
        else:
            for r in range(2, 102):
                c = ws.cell(row=r, column=col_idx)
                c.border = thin_border

        # Lista desplegable con los valores permitidos, aplicada al área de datos
        allowed = f.get('allowed', [])
        if allowed:
            formula = register_list(allowed)
            dv = DataValidation(type="list", formula1=formula, allow_blank=True)
            dv.error = f"Valor no permitido. Use uno de: {', '.join(allowed)}"
            dv.errorTitle = "Valor inválido"
            ws.add_data_validation(dv)
            dv.add(f"{col_letter}2:{col_letter}101")

    # Column widths
    ws.column_dimensions['A'].width = 4
    for col_idx in range(1, len(headers) + 1):
        col_letter = ws.cell(row=1, column=col_idx).column_letter
        max_len = max(len(str(headers[col_idx - 1])), 12)
        ws.column_dimensions[col_letter].width = min(max_len + 3, 55)

    ws.freeze_panes = "A2"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=plantilla_{template_key}.xlsx"},
    )

# ─── Eliminar cargue ─────────────────────────────────────────────────────

@app.delete("/cargues/{cargue_id}")
async def delete_cargue(cargue_id: int, current_user: User = Depends(get_current_user)):
	ensure_db_ready()
	db = SessionLocal()
	try:
		cargue = db.get(Cargue, cargue_id)
		if cargue is None:
			raise HTTPException(status_code=404, detail="Cargue no encontrado")
		if current_user.role == "prestador" and cargue.user_id != current_user.id:
			raise HTTPException(status_code=403, detail="No autorizado")
		if current_user.role == "lider":
			_prest = db.query(Prestador).filter(Prestador.user_id == current_user.id).first()
			_assigned = [p.template_key for p in (_prest.plantillas if _prest else [])]
			if cargue.template_key not in _assigned:
				raise HTTPException(status_code=403, detail="No autorizado")
		db.delete(cargue)
		db.commit()
		return {"ok": True, "id": cargue_id}
	except OperationalError:
		db.rollback()
		raise HTTPException(status_code=503, detail="No se pudo conectar a la base de datos. Verifica la conexion al servidor PostgreSQL.")
	except HTTPException:
		db.rollback()
		raise
	except Exception as exc:
		db.rollback()
		raise HTTPException(status_code=500, detail=f"Error al eliminar el cargue: {exc}")
	finally:
		db.close()


# ─── Validacion sin correccion (validador estricto) ──────────────────────

@app.post("/validate-data")
async def validate_data(payload: dict):
	template_key = payload.get("template_key", "gestante")
	corrected_text = payload.get("corrected_text", "")
	template_names_list = payload.get("template_names", [])

	if not corrected_text or not corrected_text.strip():
		raise HTTPException(status_code=400, detail="No hay datos para validar")

	try:
		df = pd.read_csv(io.StringIO(corrected_text), sep='|', header=None, dtype=str, engine='python', keep_default_na=False)
		df = df.fillna('').astype(str)
	except Exception as e:
		raise HTTPException(status_code=400, detail=f"Error al parsear datos: {e}")

	if df.empty:
		raise HTTPException(status_code=400, detail="No hay datos para validar")

	meta = get_template_by_key(template_key)
	tmpl = meta["template"]

	if template_names_list and len(df.columns) == len(template_names_list):
		df.columns = template_names_list
	else:
		tmpl_names = [t['name'] for t in tmpl]
		if len(df.columns) == len(tmpl_names):
			df.columns = tmpl_names

	# Usar validate_only() de validators.py — UN SOLO validador para todo
	try:
		from .validators import validate_only, normalize_text
	except ImportError:
		from validators import validate_only, normalize_text

	mapping = {c: c for c in df.columns if c in [t['name'] for t in tmpl]}
	validation_result = validate_only(df, mapping, tmpl)

	# Convertir resultado de validate_only al formato que espera el frontend
	n = len(df)
	tmap = {t["name"]: t for t in tmpl}
	template_cols = [t["name"] for t in tmpl]

	row_errors: dict[int, list[str]] = {}
	errors_by_cell: dict[tuple, str] = {}
	stats = {"total": n, "rows_with_errors": 0, "total_error_cells": 0, "by_column": {}}

	for log_entry in validation_result.get("logs", []):
		ridx = int(log_entry["row"]) - 1
		col = log_entry["column"]
		orig = log_entry.get("original", "")
		tdef = tmap.get(col)
		tipo = tdef.get("type") if tdef else "TEXT"

		if ridx not in row_errors:
			row_errors[ridx] = []

		if tipo == "SET":
			allowed = [str(a).strip() for a in (tdef.get("allowed") or [])]
			opciones = ", ".join(allowed[:8]) if allowed else "los valores del instructivo"
			if not orig.strip():
				msg = f"[{col}] Campo vacio. Opciones: {pciones}"
			else:
				msg = f"[{col}] Valor '{orig}' no valido. Opciones: {pciones}"
		elif tipo == "INT":
			if not orig.strip():
				msg = f"[{col}] Campo vacio. Solo enteros (ej: 25)"
			else:
				msg = f"[{col}] '{orig}' no es entero valido (ej: 25)"
		elif tipo == "DECIMAL":
			if not orig.strip():
				msg = f"[{col}] Campo vacio. Numero decimal (ej: 60.5)"
			else:
				msg = f"[{col}] '{orig}' no es decimal valido (ej: 60.5)"
		elif tipo == "DATE":
			if not orig.strip():
				msg = f"[{col}] Campo vacio. Fecha (AAAA-MM-DD)"
			else:
				msg = f"[{col}] '{orig}' no es fecha valida (AAAA-MM-DD)"
		else:
			msg = f"[{col}] {log_entry.get('corrected', 'Dato invalido')}"

		row_errors[ridx].append(msg)
		errors_by_cell[(ridx, col)] = log_entry.get("corrected", "Dato invalido")
		stats["total_error_cells"] += 1

	# Contar errores por columna
	for ridx, errs in row_errors.items():
		for e in errs:
			import re as _re
			m = _re.match(r'\[([^\]]+)\]', e)
			if m:
				cn = m.group(1)
				if cn not in stats["by_column"]:
					stats["by_column"][cn] = {"type": tmap.get(cn, {}).get("type", "?"), "errors": 0}
				stats["by_column"][cn]["errors"] += 1

	# Cross-field validation for gestante template
	cross_field_errors = []
	if template_key == "gestante":
		try:
			from .validators import validate_cross_fields
		except ImportError:
			from validators import validate_cross_fields
		try:
			cross_field_errors = validate_cross_fields(df)
			for e in cross_field_errors:
				ridx = e["row"] - 1
				if ridx not in row_errors:
					row_errors[ridx] = []
				row_errors[ridx].append(f"[{e['column']}] {e['message']}")
				stats["total_error_cells"] += 1
		except Exception:
			pass

	stats["rows_with_errors"] = len(row_errors)

	tipos_por_col = {t["name"]: t["type"] for t in tmpl}

	def normalizar_celda(valor, nombre_col):
		s = str(valor) if valor is not None else ""
		s = s.strip()
		s = s.replace("|", " ").replace("\r", " ").replace("\n", " ").replace("\t", " ")
		s = s.replace("_x000D_", " ").replace("_x000B_", " ")
		s = re.sub(r"\s+", " ", s).strip()
		if not s or s.upper() in ("SIN DATO", "NO APLICA", "N/A", "NONE"):
			return s
		if tipos_por_col.get(nombre_col) == "DATE":
			iso = to_date_iso(s)
			if iso:
				return iso
		return s

	header_line = "|".join(template_cols) + "|RESULTADO DE VALIDACION"
	output_lines = [header_line]
	for ridx in range(n):
		row_vals = [normalizar_celda(df.iloc[ridx, ci], template_cols[ci]) if ci < len(df.columns) else "" for ci in range(len(template_cols))]
		errs = row_errors.get(ridx, [])
		err_col = "; ".join(errs) if errs else "VALIDADO"
		output_lines.append("|".join(row_vals) + "|" + err_col)

	report_text = "\ufeff" + "\r\n".join(output_lines)

	logs_sample = []
	for (r, c), msg in errors_by_cell.items():
		val = ""
		if r < len(template_rows) and c < len(template_cols):
			val = str(template_rows[r][c]) if c < len(template_rows[r]) else ""
		# Skip false-positive entries: empty, SIN DATO, or formula markers
		val_norm = val.upper().replace(" ", "").strip()
		if not val or val_norm in ("", "SINDATO", "N/A", "NA"):
			continue
		if msg == "OBLIGATORIO":
			continue
		# Check for formula marker that may leak into corrector output
		if "<CALCULAR:" in msg.upper() or "_CALCULADO" in val_norm:
			continue
		logs_sample.append({
			"row": r + 1,
			"column": c,
			"original": val,
			"corrected": msg,
			"status": "error",
		})

	return {
		"valid": stats["rows_with_errors"] == 0,
		"stats": {
			"total_rows": n,
			"rows_with_errors": stats["rows_with_errors"],
			"rows_ok": n - stats["rows_with_errors"],
			"total_error_cells": stats["total_error_cells"],
			"by_column": stats["by_column"],
			"cross_field_errors": len(cross_field_errors),
		},
		"row_errors": {str(k): v for k, v in row_errors.items()},
		"errors_by_cell": {f"{r}|{c}": m for (r, c), m in errors_by_cell.items()},
		"cross_field_errors": cross_field_errors,
		"logs_sample": logs_sample,
		"report_text": base64.b64encode(report_text.encode("utf-8")).decode("ascii"),
		"template_key": template_key,
	}


@app.post("/indicadores")
async def indicadores_endpoint(payload: dict):
	template_key = payload.get("template_key", "gestante")
	corrected_text = payload.get("corrected_text", "")

	if not corrected_text or not corrected_text.strip():
		raise HTTPException(status_code=400, detail="No hay datos para generar indicadores")

	try:
		df = pd.read_csv(io.StringIO(corrected_text), sep='|', header=None, dtype=str, engine='python', keep_default_na=False)
		df = df.fillna('').astype(str)
	except Exception as e:
		raise HTTPException(status_code=400, detail=f"Error al parsear datos: {e}")

	if df.empty:
		raise HTTPException(status_code=400, detail="No hay datos")

	meta = get_template_by_key(template_key)
	tmpl = meta["template"]
	tmpl_names = [t['name'] for t in tmpl]

	if len(df.columns) == len(tmpl_names):
		df.columns = tmpl_names

	tmap = {t["name"]: t for t in tmpl}

	def safe_col(name):
		if name in df.columns:
			return df[name]
		return pd.Series([""] * len(df), index=df.index)

	def count_values(series, allowed_map=None):
		counts = {}
		for v in series:
			vn = normalize_text(str(v).strip()) if v else ""
			if not vn or vn in ("SIN DATO", "NONE", "NAN", ""):
				continue
			if allowed_map and vn in allowed_map:
				vn = allowed_map[vn]
			counts[vn] = counts.get(vn, 0) + 1
		return counts

	def pct(part, total):
		return round(part / max(total, 1) * 100, 1)

	totalregistros = len(df)
	registrosok = 0
	registrosconerror = 0

	# Stats base
	result = {
		"template_key": template_key,
		"total_registros": totalregistros,
		"indicadores": {},
	}

	if template_key == "gestante":
		try:
			from .indicadores_pare import calcular_indicadores_gestante
		except ImportError:
			from indicadores_pare import calcular_indicadores_gestante
		pare = calcular_indicadores_gestante(df)
		# Descriptivos complementarios (conteos)
		sexo = count_values(safe_col("SEXO"))
		regimen = count_values(safe_col("REGIMEN DE AFILIACION"))
		etnia = count_values(safe_col("ETNIA"))
		zona = count_values(safe_col("ZONA"))
		riesgo = count_values(safe_col("CLASIFICACION DEL RIESGO"))
		vih = count_values(safe_col("RESULTADO PRIMER TAMIZAJE PRUEBA DE VIH"))
		sifilis = count_values(safe_col("RESULTADO PRIMERA PRUEBA TREPONEMICA RAPIDA SIFILIS"))
		hipertension = count_values(safe_col("HIPERTENSION ARTERIAL"))
		diabetes = count_values(safe_col("DIABETES"))
		trimestre = count_values(safe_col("TRIMESTRE INICIO CONTROL"))
		parto = count_values(safe_col("CARACTERISTICAS DEL PARTO"))
		condicion_rn = count_values(safe_col("CONDICION DEL RECIEN NACIDO"))
		vacuna_bcg = count_values(safe_col("VACUNACION CON BCG"))
		vacuna_hepb = count_values(safe_col("VACUNACION ANTIHEPATITIS B"))

		result["indicadores"] = {
			"pare_mm": {
				"label": "Indicadores PARE MM (Cohorte de Gestantes)",
				"type": "pare",
				"total_gestantes": pare["total_gestantes"],
				"fecha_referencia": pare["fecha_referencia"],
				"lista": pare["indicadores"],
				"por_municipio": pare.get("por_municipio", []),
			},
			"descriptivos": {
				"label": "Distribucion de la cohorte",
				"type": "descriptivos",
				"data": {
					"sexo": sexo,
					"regimen": regimen,
					"etnia": etnia,
					"zona": zona,
					"riesgo": riesgo,
					"vih": vih,
					"sifilis": sifilis,
					"hipertension": hipertension,
					"diabetes": diabetes,
					"trimestre": trimestre,
					"tipo_parto": parto,
					"condicion_rn": condicion_rn,
					"vacuna_bcg": vacuna_bcg,
					"vacuna_hepb": vacuna_hepb,
				},
			},
		}

	elif template_key == "citologia":
		resultado_citologia = count_values(safe_col("RESULTADO CITOLOGIA CERVICOUTERINA"))
		tipo_doc = count_values(safe_col("TIPO DE DOCUMENTO DE IDENTIDAD"))
		result["indicadores"] = {
			"resultado_citologia": {"label": "Resultado citologia", "data": resultado_citologia, "total": totalregistros},
			"tipo_documento": {"label": "Tipo de documento", "data": tipo_doc, "total": totalregistros},
		}

	elif template_key == "mamografia":
		resultado_mamo = count_values(safe_col("RESULTADO MAMOGRAFIA"))
		result["indicadores"] = {
			"resultado_mamografia": {"label": "Resultado mamografia", "data": resultado_mamo, "total": totalregistros},
		}

	elif template_key == "penta":
		vacuna_penta = count_values(safe_col("VACUNACION PENTAVALENTE"))
		result["indicadores"] = {
			"vacuna_penta": {"label": "Vacunacion pentavalente", "data": vacuna_penta, "total": totalregistros},
		}

	else:
		result["indicadores"] = {}

	return result


@app.post("/indicadores-de-cargue/{cargue_id}")
async def indicadores_de_cargue(cargue_id: int, current_user: User = Depends(get_current_user)):
	"""Calcula los indicadores directamente desde un cargue guardado en la BD."""
	ensure_db_ready()
	db = SessionLocal()
	try:
		cargue = db.get(Cargue, cargue_id)
		if cargue is None:
			raise HTTPException(status_code=404, detail="Cargue no encontrado")
		if current_user.role == "prestador" and cargue.user_id != current_user.id:
			raise HTTPException(status_code=403, detail="No autorizado")
		text = cargue.corrected_text or cargue.raw_text or ""
		if cargue.compressed and text:
			try:
				text = gzip.decompress(base64.b64decode(text)).decode("utf-8")
			except Exception:
				pass
		if not text or not text.strip():
			raise HTTPException(status_code=400, detail="El cargue no tiene datos validos")
		# Reutilizar la logica del endpoint /indicadores con el texto del cargue
		return await indicadores_endpoint({
			"template_key": cargue.template_key or "gestante",
			"corrected_text": text,
		})
	finally:
		db.close()


@app.post("/indicadores-excel/{cargue_id}")
async def indicadores_excel(cargue_id: int, current_user: User = Depends(get_current_user)):
	"""Exporta los indicadores PARE MM de un cargue a Excel."""
	ensure_db_ready()
	db = SessionLocal()
	try:
		cargue = db.get(Cargue, cargue_id)
		if cargue is None:
			raise HTTPException(status_code=404, detail="Cargue no encontrado")
		if current_user.role == "prestador" and cargue.user_id != current_user.id:
			raise HTTPException(status_code=403, detail="No autorizado")
		text = cargue.corrected_text or cargue.raw_text or ""
		if cargue.compressed and text:
			try:
				text = gzip.decompress(base64.b64decode(text)).decode("utf-8")
			except Exception:
				pass
		if not text or not text.strip():
			raise HTTPException(status_code=400, detail="El cargue no tiene datos validos")
		data = await indicadores_endpoint({
			"template_key": cargue.template_key or "gestante",
			"corrected_text": text,
		})
		pare = (data.get("indicadores") or {}).get("pare_mm") or {}
		descriptivos = (data.get("indicadores") or {}).get("descriptivos")
		try:
			from .indicadores_excel import build_indicadores_excel
		except ImportError:
			from indicadores_excel import build_indicadores_excel
		buf = build_indicadores_excel(pare, descriptivos, cargue.original_filename or "")
		filename = f"indicadores_pare_mm_{cargue_id}.xlsx"
		return StreamingResponse(
			buf,
			media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
			headers={"Content-Disposition": f"attachment; filename={filename}"},
		)
	finally:
		db.close()


@app.post("/cargues/{cargue_id}/reporte-errores")
async def descargar_reporte_errores(cargue_id: int, current_user: User = Depends(get_current_user)):
	"""Genera el TXT de errores directamente desde el cargue en BD (descarga rapida)."""
	ensure_db_ready()
	db = SessionLocal()
	try:
		cargue = db.get(Cargue, cargue_id)
		if cargue is None:
			raise HTTPException(status_code=404, detail="Cargue no encontrado")
		if current_user.role == "prestador" and cargue.user_id != current_user.id:
			raise HTTPException(status_code=403, detail="No autorizado")
		text = cargue.corrected_text or cargue.raw_text or ""
		if cargue.compressed and text:
			try:
				text = gzip.decompress(base64.b64decode(text)).decode("utf-8")
			except Exception:
				pass
		if not text or not text.strip():
			raise HTTPException(status_code=400, detail="El cargue no tiene datos validos")
		# Reutilizar la validacion de /validate-data para obtener el reporte
		resp = await validate_data({
			"template_key": cargue.template_key or "gestante",
			"corrected_text": text,
		})
		report_text = resp["report_text"]
		# Decodificar base64 a bytes y agregar BOM UTF-8
		bin_bytes = base64.b64decode(report_text)
		if not (bin_bytes[:3] == b"\xef\xbb\xbf"):
			bin_bytes = b"\xef\xbb\xbf" + bin_bytes
		filename = f"reporte_errores_{cargue_id}.txt"
		return StreamingResponse(
			io.BytesIO(bin_bytes),
			media_type="text/plain; charset=utf-8",
			headers={"Content-Disposition": f"attachment; filename={filename}"},
		)
	finally:
		db.close()


@app.post("/cargues/{cargue_id}/reporte-errores-excel")
async def descargar_reporte_errores_excel(cargue_id: int, current_user: User = Depends(get_current_user)):
	"""Genera un Excel de la data con errores marcados en ROJO y comentarios
	de correccion. El prestador corrige y re-subve el archivo."""
	ensure_db_ready()
	db = SessionLocal()
	try:
		cargue = db.get(Cargue, cargue_id)
		if cargue is None:
			raise HTTPException(status_code=404, detail="Cargue no encontrado")
		if current_user.role == "prestador" and cargue.user_id != current_user.id:
			raise HTTPException(status_code=403, detail="No autorizado")
		text = cargue.corrected_text or cargue.raw_text or ""
		if cargue.compressed and text:
			try:
				text = gzip.decompress(base64.b64decode(text)).decode("utf-8")
			except Exception:
				pass
		if not text or not text.strip():
			raise HTTPException(status_code=400, detail="El cargue no tiene datos validos")
		errors_by_cell = _errores_rapidos(text, cargue.template_key or "gestante")
		meta = get_template_by_key(cargue.template_key or "gestante")
		tmpl = meta["template"]
		try:
			from .excel_export import build_reporte_errores_excel
		except ImportError:
			from excel_export import build_reporte_errores_excel
		buf = build_reporte_errores_excel(text, tmpl, errors_by_cell)
		filename = cargue.original_filename.replace(".xlsx", "_errores.xlsx").replace(".xls", "_errores.xlsx")
		return StreamingResponse(
			buf,
			media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
			headers={"Content-Disposition": f"attachment; filename={filename}"},
		)
	finally:
		db.close()


@app.post("/reporte-errores-excel-data")
async def reporte_errores_excel_data(payload: dict):
	"""Genera el Excel de errores directamente desde un texto (sin cargue en BD).
	Usado desde la pantalla de validacion."""
	template_key = payload.get("template_key", "gestante")
	corrected_text = payload.get("corrected_text", "")
	if not corrected_text or not corrected_text.strip():
		raise HTTPException(status_code=400, detail="No hay datos para validar")
	errors_by_cell = _errores_rapidos(corrected_text, template_key)
	meta = get_template_by_key(template_key)
	tmpl = meta["template"]
	try:
		from .excel_export import build_reporte_errores_excel
	except ImportError:
		from excel_export import build_reporte_errores_excel
	buf = build_reporte_errores_excel(corrected_text, tmpl, errors_by_cell)
	filename = f"reporte_errores_{template_key}.xlsx"
	return StreamingResponse(
		buf,
		media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
		headers={"Content-Disposition": f"attachment; filename={filename}"},
	)


def _errores_rapidos(corrected_text: str, template_key: str) -> dict:
	"""Calcula los errores por celda usando el validador VECTORIZADO (rapido).
	Devuelve {(row_idx, col_name): mensaje_correccion}."""
	try:
		from .validators import validate_only, normalize_text
	except ImportError:
		from validators import validate_only, normalize_text
	import pandas as _pd
	df = _pd.read_csv(io.StringIO(corrected_text), sep='|', header=None, dtype=str, engine='python', keep_default_na=False)
	df = df.fillna('').astype(str)
	meta = get_template_by_key(template_key)
	tmpl = meta["template"]
	tmpl_names = [t['name'] for t in tmpl]
	if len(df.columns) == len(tmpl_names):
		df.columns = tmpl_names
	# Sin rellenar vacios: el validador estricto marca vacios y tipos incorrectos
	mapping = {c: c for c in df.columns if c in tmpl_names}
	res = validate_only(df, mapping, tmpl)
	errors = {}
	tmap2 = {t["name"]: t for t in tmpl}
	for l in res["logs"]:
		row = int(l["row"]) - 1
		col = l["column"]
		corr = l["corrected"] or "Dato invalido"
		orig = l["original"] or ""
		es_vacio = orig.strip() == ""
		# Comentario explicito: que se encontro y que debe ingresar
		tdef2 = tmap2.get(col)
		tipo2 = tdef2.get("type") if tdef2 else None
		if tipo2 == "SET":
			allowed = [str(a).strip() for a in (tdef2.get("allowed") or [])]
			opciones = ", ".join(allowed) if allowed else "los valores del instructivo"
			if es_vacio:
				msg = f"Este campo esta VACIO. Debe escribir uno de estos valores: {opciones}."
			else:
				msg = f"El valor '{orig}' no es valido. Debe escribir uno de estos valores: {opciones}."
		elif tipo2 == "INT":
			if es_vacio:
				msg = f"Este campo esta VACIO y debe llevar SOLO NUMEROS (entero). Ej: 25, 1103567890."
			else:
				msg = f"El valor '{orig}' no es un numero entero valido. En este campo solo van NUMEROS (ej: 25, 1103567890)."
		elif tipo2 == "DECIMAL":
			if es_vacio:
				msg = f"Este campo esta VACIO y debe llevar UN NUMERO (puede tener decimales). Ej: 60.5, 1.60."
			else:
				msg = f"El valor '{orig}' no es un numero valido. En este campo va un NUMERO (ej: 60.5, 1.60)."
		elif tipo2 == "DATE":
			if es_vacio:
				msg = f"Este campo esta VACIO y debe llevar UNA FECHA en formato AAAA-MM-DD. Ej: 2025-10-10."
			else:
				msg = f"El valor '{orig}' no es una fecha valida. Debe escribir la fecha como AAAA-MM-DD (ej: 2025-10-10)."
		elif tipo2 == "TEXT":
			if es_vacio:
				msg = f"Este campo esta VACIO. Debe escribir el dato en TEXTO, o 'SIN DATO' si no lo tiene."
			else:
				msg = f"El valor '{orig}' no es valido. Este campo va en TEXTO (letras), o 'SIN DATO' si no tiene el dato."
		else:
			msg = f"El dato '{orig}' no es valido. {corr}"
		errors[(row, col)] = msg
	return errors


# ─── Verificar afiliado ─────────────────────────────────────────────────
# Consulta la tabla administrativo.af_afiliado por numero de documento y
# extrae los datos demograficos de la usuaria.

AFILIADO_ESQUEMA = "administrativo"
AFILIADO_TABLA = "af_afiliado"

# Codigos DANE de departamentos
DEPARTAMENTOS = {
    "05": "ANTIOQUIA", "08": "ATLANTICO", "11": "BOGOTA D.C.", "13": "BOLIVAR",
    "15": "BOYACA", "17": "CALDAS", "18": "CAQUETÁ", "19": "CASANARE",
    "20": "CAUCA", "23": "CESAR", "25": "CUNDINAMARCA", "27": "CHOCO",
    "41": "HUILA", "44": "LA GUAJIRA", "47": "MAGDALENA", "50": "META",
    "52": "NARIÑO", "54": "NORTE DE SANTANDER", "63": "QUINDIO", "66": "RISARALDA",
    "68": "SANTANDER", "70": "SUCRE", "73": "TOLIMA", "76": "VALLE DEL CAUCA",
    "81": "ARAUCA", "85": "CASANARE", "86": "PUTUMAYO", "88": "SAN ANDRÉS",
    "91": "AMAZONAS", "94": "GUAINÍA", "95": "VAUPÉS", "97": "VICHADA",
}

# Codigos DANE de municipios principales (departamento + 3 digitos)
MUNICIPIOS = {
    "05001": "MEDELLIN", "05002": "ABEJORRAL", "05004": "ABRIAQUI",
    "08001": "BARRANQUILLA", "08078": "BARANOA", "08137": "CAMPO DE LA CRUZ",
    "08296": "GALAPA", "08372": "JUAN DE ACOSTA", "08421": "LURUACO",
    "08433": "MALAMBO", "08440": "MANATI", "08549": "PALMAR DE VARELA",
    "08558": "PIOJO", "08560": "POLONUEVO", "08573": "PONEDERA",
    "08606": "PUERTO COLOMBIA", "08634": "REPELON", "08638": "SABANAGRANDE",
    "08675": "SABANALARGA", "08685": "SANTA LUCIA", "08770": "SANTO TOMAS",
    "08758": "SOLEDAD", "08832": "SUAN", "08849": "TUBARA",
    "08885": "USIACURI", "11001": "BOGOTA D.C.",
    "13001": "CARTAGENA", "13006": "ACHI", "13030": "ALTOS DEL ROSARIO",
    "13042": "ARJONA", "13052": "ARROYOHONDO", "13062": "BARRANCO DE LOBA",
    "13074": "CALAMAR", "13160": "CICUCO", "13188": "CORDOBA",
    "13212": "CLEMENCIA", "13244": "HATILLO DE LOBA", "13248": "MAGANGUE",
    "13268": "MAHATES", "13300": "MARGARITA", "13322": "MARIA LA BAJA",
    "13440": "MONTECRISTO", "13473": "MORALES", "13549": "PINILLOS",
    "13580": "REGIDOR", "13600": "RIO VIEJO", "13620": "SAN CRISTOBAL",
    "13647": "SAN ESTANISLAO", "13650": "SAN FERNANDO", "13654": "SAN JACINTO",
    "13655": "SAN JACINTO DEL CAUCA", "13657": "SAN JUAN NEPOMUCENO",
    "13667": "SAN MARTIN DE LOBA", "13670": "SAN PABLO", "13673": "SANTA CATALINA",
    "13683": "SANTA ROSA", "13688": "SANTO DOMINGO", "13744": "SIPEHI",
    "13760": "SOPLAVIENTO", "13780": "TALAIGUA NUEVO", "13810": "TIQUISIO",
    "13836": "TURBACO", "13838": "TURBANA", "13873": "VILLANUEVA",
    "13894": "ZAMBRANO",
    "15001": "TUNJA", "15022": "ALMEIDA", "15047": "AQUITANIA",
    "15051": "ARCABUCO", "15087": "BELEN", "15090": "BERBEO",
    "15092": "BETEITIVA", "15097": "BOAVITA", "15104": "BOYACA",
    "15106": "BUENAVISTA", "15109": "BUSTUBANTE", "15114": "CALDAS",
    "15131": "CAMPOHERMOSO", "15162": "CERINZA", "15172": "CHINAVITA",
    "15176": "CHIQUINQUIRA", "15180": "CHIQUIZA", "15183": "CHIVATA",
    "15185": "CIENEGA", "15187": "COMBITA", "15204": "COPER",
    "15212": "CORRALES", "15215": "COVARACHIA", "15223": "CUBARA",
    "15226": "CUCAITA", "15232": "CUITIVA", "15236": "CHIVOR",
    "15238": "DUITAMA", "15244": "EL COCUY", "15248": "EL ESPINO",
    "15264": "FIRAVITOBA", "15272": "FLORESTA", "15276": "GACHANTIVA",
    "15293": "GAMEZA", "15296": "GARAGOA", "15317": "GUACAMAYAS",
    "15322": "GUATEQUE", "15325": "GUAYATA", "15332": "GARAGOA",
    "15362": "IZA", "15367": "JENESANO", "15377": "JORDAN",
    "15380": "LA CAPILLA", "15401": "LENDA", "15425": "MARIPI",
    "15442": "MIRAFLORES", "15455": "MONGUA", "15464": "MONGUI",
    "15466": "MONIQUIRA", "15476": "MOTAVITA", "15480": "MUZO",
    "15491": "NOBSA", "15494": "NUEVO COLON", "15500": "OICATA",
    "15507": "OTANCHE", "15511": "PACHAVITA", "15516": "PAEZ",
    "15522": "PAIPA", "15531": "PAJARITO", "15533": "PANQUEBA",
    "15537": "PAUNA", "15542": "PAYA", "15550": "PAZ DE RIO",
    "15572": "PESCA", "15580": "PISBA", "15599": "PUERTO BOYACA",
    "15600": "QUIPAMA", "15621": "RAMIRIQUI", "15632": "RAQUIRA",
    "15638": "RONDON", "15646": "SABOYA", "15660": "SACHICA",
    "15664": "SAMACA", "15667": "SAN EDUARDO", "15673": "SAN JOSE DE PARE",
    "15676": "SAN LUIS DE SACO", "15681": "SANTANA", "15686": "SANTA MARIA",
    "15690": "SANTA SOFIA", "15693": "SANTANDER", "15696": "SATIVANORTE",
    "15697": "SATIVASUR", "15707": "SOACHA", "15740": "SOGAMOSO",
    "15753": "SOMONDOCO", "15755": "SORA", "15757": "SOTAQUIRA",
    "15759": "SOTARA", "15761": "SUSACON", "15762": "SUTAMARCHAN",
    "15764": "SUTATENZA", "15774": "TASCO", "15776": "TENZA",
    "15781": "TIBANA", "15783": "TIBASOSA", "15785": "TINJACA",
    "15787": "TIPACOQUE", "15793": "TOCA", "15797": "TOGUI",
    "15804": "TOPAGA", "15806": "TOTA", "15808": "TUNUNGUA",
    "15810": "TURMEQUE", "15814": "TUTA", "15816": "TUTAZA",
    "15820": "UMBITA", "15832": "VENTAQUEMADA", "15835": "VIRACACHA",
    "15842": "ZETAQUIRA", "17001": "MANIZALES", "17013": "ANSERMA",
    "17042": "ARANZAZU", "17050": "BELALCAZAR", "17088": "CHINCHINA",
    "17174": "FILADELFIA", "17272": "LA DORADA", "17380": "LA MERCED",
    "17433": "MONTENEGRO", "17442": "PENSILVANIA", "17497": "RIOSUCIO",
    "17524": "RISARALDA", "17541": "SANTA ROSA DE CABAL", "17614": "SANTUARIO",
    "18001": "FLORENCIA", "18029": "ALBANIA", "18094": "BELEN DE LOS ANDAQUIES",
    "18150": "CARTAGENA DEL CHAIRA", "18205": "CURILOCO", "18247": "EL DONCELLO",
    "18256": "EL PAUJIL", "18410": "LA MONTAÑITA", "18460": "MILAN",
    "18479": "MORELIA", "18592": "PUERTO RICO", "18610": "SAN JOSE DE LA MONTAÑA",
    "18653": "SAN MIGUEL DEL DUDA", "18656": "SAN MARTIN DE LOS LLANOS",
    "18753": "SOLANO", "18756": "SOLITA", "18860": "VALPARAISO",
    "19001": "POPAYAN", "19022": "ALMAGUER", "19050": "ARGELIA",
    "19075": "BALBOA", "19100": "BOLIVAR", "19110": "BUENOS AIRES",
    "19130": "CAJIBIO", "19137": "CALDONO", "19142": "CALOTO",
    "19215": "CORINTO", "19256": "EL TAMBO", "19290": "FLORENCIA",
    "19300": "GUACHENE", "19310": "GUAPI", "19355": "INZA",
    "19364": "JAMBALO", "19392": "LA SIERRA", "19397": "LA VEGA",
    "19418": "MERCADERES", "19450": "MISAGA", "19455": "MONDOMO",
    "19473": "MORALES", "19513": "PADILLA", "19517": "PAEZ",
    "19532": "PATIA (EL BORDO)", "19533": "PIAMONTE", "19548": "PIENDAMO",
    "19573": "PUERTO TEJADA", "19585": "PURACE (COCONUCO)", "19622": "ROSAS",
    "19693": "SAN SEBASTIAN", "19698": "SANTANDER DE QUILICHAO",
    "19701": "SANTA ROSA", "19743": "SILVIA", "19760": "SOTARA (PAISPAMBA)",
    "19780": "SUAREZ", "19785": "SUCRE", "19807": "TIMBIO", "19809": "TIMBIQUI",
    "19821": "TORIBIO", "19824": "TOTORO", "19845": "VILLA RICA",
    "20001": "POPAYAN", "20011": "ALMEIDA", "20013": "BARRANQUILLA",
    "20032": "AVENIDA", "20045": "CALIFORNIA", "20060": "EL COCUY",
    "20099": "GARZON", "20110": "GIGANTE", "20124": "GUADALUPE",
    "20130": "HOBO", "20136": "ICACAS", "20149": "LA ARGENTINA",
    "20156": "LA PLATA", "20167": "NATAGA", "20170": "OPORAPA",
    "20174": "PAICOL", "20177": "PALMIRA", "20183": "PALESTINA",
    "20196": "PITAL", "20215": "PITALITO", "20260": "RIVERA",
    "20272": "SALADOBLANCO", "20274": "SAN AGUSTIN", "20283": "SANTA MARIA",
    "20296": "SUAZA", "20310": "TESALIA", "20319": "TULUA",
    "20349": "VILLAVIEJA", "20357": "YAGUARA", "20001": "GAIRA",
    "23001": "VALLEDUPAR", "23001": "VALLEDUPAR", "23001": "VALLEDUPAR",
    "25001": "AGUA DE DIOS", "25001": "AGUA DE DIOS", "25001": "AGUA DE DIOS",
    "41001": "NEIVA", "41001": "NEIVA", "41001": "NEIVA",
    "44001": "RIOHACHA", "44001": "RIOHACHA", "44001": "RIOHACHA",
    "47001": "SANTA MARTA", "47001": "SANTA MARTA", "47001": "SANTA MARTA",
    "50001": "VILLAVICENCIO", "50001": "VILLAVICENCIO", "50001": "VILLAVICENCIO",
    "52001": "PASTO", "52001": "PASTO", "52001": "PASTO",
    "54001": "CUCUTA", "54001": "CUCUTA", "54001": "CUCUTA",
    "63001": "ARMENIA", "63001": "ARMENIA", "63001": "ARMENIA",
    "66001": "PEREIRA", "66001": "PEREIRA", "66001": "PEREIRA",
    "68001": "BUCARAMANGA", "68001": "BUCARAMANGA", "68001": "BUCARAMANGA",
    "70001": "SINCELEJO", "70001": "SINCELEJO", "70001": "SINCELEJO",
    "73001": "IBAGUE", "73001": "IBAGUE", "73001": "IBAGUE",
    "76001": "CALI", "76001": "CALI", "76001": "CALI",
    "81001": "ARAUCA", "81001": "ARAUCA", "81001": "ARAUCA",
    "85001": "YOPAL", "85001": "YOPAL", "85001": "YOPAL",
    "86001": "MOCOA", "86001": "MOCOA", "86001": "MOCOA",
    "88001": "SAN ANDRES", "88001": "SAN ANDRES", "88001": "SAN ANDRES",
    "91001": "LETICIA", "91001": "LETICIA", "91001": "LETICIA",
    "94001": "INIRIDA", "94001": "INIRIDA", "94001": "INIRIDA",
    "95001": "MITU", "95001": "MITU", "95001": "MITU",
    "97001": "PUERTO CARRENO", "97001": "PUERTO CARRENO", "97001": "PUERTO CARRENO",
}


def resolver_departamento(codigo) -> str:
    """Resuelve codigo de departamento a nombre."""
    if codigo is None:
        return None
    cod = str(codigo).strip().zfill(2)
    return DEPARTAMENTOS.get(cod, str(codigo).strip())


def resolver_municipio(codigo) -> str:
    """Resuelve codigo de municipio a nombre."""
    if codigo is None:
        return None
    cod = str(codigo).strip().zfill(5)
    nombre = MUNICIPIOS.get(cod)
    if nombre:
        return nombre
    # Si no esta en el diccionario, intentar con el codigo de depto
    depto = cod[:2]
    nombre_depto = DEPARTAMENTOS.get(depto, "")
    if nombre_depto:
        return f"Municipio {cod} ({nombre_depto})"
    return str(codigo).strip()

# Campos demograficos solicitados y las palabras clave para mapear la columna real
AFILIADO_CAMPOS = [
    ("primer_nombre", ["PRIMER NOMBRE", "PRIMERNOMBRE", "NOMBRE1", "NOMBRE 1"]),
    ("segundo_nombre", ["SEGUNDO NOMBRE", "SEGUNDONOMBRE", "NOMBRE2", "NOMBRE 2"]),
    ("primer_apellido", ["PRIMER APELLIDO", "PRIMERAPELLIDO", "APELLIDO1", "APELLIDO 1"]),
    ("segundo_apellido", ["SEGUNDO APELLIDO", "SEGUNDOAPELLIDO", "APELLIDO2", "APELLIDO 2"]),
    ("fecha_nacimiento", ["FECHA NACIMIENTO", "FECHA DE NACIMIENTO", "FECHANACIMIENTO", "FECHA_NAC"]),
    ("sexo", ["SEXO"]),
    ("direccion", ["DIRECCION", "DIRECCIÓN"]),
    ("correo", ["CORREO", "EMAIL", "CORREO ELECTRONICO", "CORREO ELECTRÓNICO"]),
    ("categoria", ["CATEGORIA", "CATEGORÍA", "NIVEL SISBEN", "NIVELSISBEN"]),
    ("estado_afiliado", ["ESTADO AFILIADO", "ESTADOAFILIADO", "ESTADO"]),
    ("tipo_afiliado", ["TIPO AFILIADO", "TIPOAFILIADO", "TIPO"]),
    ("fecha_inicio_cobertura", ["FECHA INICIO COBERTURA", "FECHAINICIOC", "INICIO COBERTURA", "FECHA AFILIACION ENTIDAD", "FECHAAFILIACIONENTIDAD", "FECHA AFILIACION"]),
    ("municipio_afiliacion", ["MUNICIPIO AFILIACION", "MUNICIPIOAFILIACION", "MUNICIPIO"]),
    ("departamento_afiliacion", ["DEPARTAMENTO AFILIACION", "DEPARTAMENTOAFILIACION", "DEPARTAMENTO"]),
    ("discapacidad", ["DISCAPACIDAD"]),
    ("telefono", ["TELEFONO", "TELÉFONO", "CELULAR"]),
    ("telefono_2", ["TELEFONO 2", "TELÉFONO 2", "TELEFONO2"]),
    ("celular", ["CELULAR"]),
    ("celular_2", ["CELULAR 2", "CELULAR2"]),
    ("barrio", ["BARRIO"]),
    ("ips_primaria", ["IPS PRIMARIA", "IPS"]),
]

# Palabras clave para localizar la columna de numero de documento
AFILIADO_DOC_KEYWORDS = ["NUMERO DE DOCUMENTO", "NUMERODOCUMENTO", "NUMERO DOCUMENTO", "N DOCUMENTO", "NUMERO IDENTIFICACION", "NUMEROIDENTIFICACION", "NO IDENTIFICACION", "IDENTIFICACION", "DOCUMENTO"]


def _descubrir_columnas_afiliado() -> list[str] | None:
	"""Devuelve las columnas reales de administrativo.af_afiliado o None si no accesible."""
	try:
		from .database import engine
	except ImportError:
		from database import engine
	try:
		from sqlalchemy import inspect
		insp = inspect(engine)
		cols = insp.get_columns(AFILIADO_TABLA, schema=AFILIADO_ESQUEMA)
		return [c["name"] for c in cols]
	except Exception:
		pass
	try:
		from sqlalchemy import text
		with engine.connect() as conn:
			if str(engine.url).startswith("sqlite"):
				rows = conn.execute(text(f'PRAGMA table_info("{AFILIADO_TABLA}")')).fetchall()
				return [r[1] for r in rows]
			rows = conn.execute(text(
				f"SELECT column_name FROM information_schema.columns "
				f"WHERE table_schema = '{AFILIADO_ESQUEMA}' AND table_name = '{AFILIADO_TABLA}'"
			)).fetchall()
			return [r[0] for r in rows]
	except Exception:
		return None


def _normalizar_columna(s) -> str:
	import unicodedata
	t = str(s or "").strip().upper()
	t = ''.join(ch for ch in unicodedata.normalize('NFD', t) if unicodedata.category(ch) != 'Mn')
	t = t.replace("_", " ").replace("-", " ").replace(".", " ")
	t = ' '.join(t.split())
	return t


def _mapear_columnas_afiliado(cols: list[str]) -> tuple[dict, str | None]:
	"""Mapea columnas reales a los campos demograficos y localiza la columna de documento.
	La columna de documento se busca con prioridad (coincidencia de palabras completas)
	para no confundirla con otras columnas (ej: 'CC' dentro de 'direccion')."""
	mapping = {}
	norm_cols = {c: _normalizar_columna(c) for c in cols}

	# 1) Columna de documento: coincidencia de TODAS las palabras de la keyword
	doc_candidates = []
	for c, nc in norm_cols.items():
		n_words = set(nc.split())
		for kw in AFILIADO_DOC_KEYWORDS:
			kw_words = set(kw.split())
			if nc == kw or (kw_words and kw_words.issubset(n_words)):
				doc_candidates.append(c)
				break
	def doc_score(c):
		nc = norm_cols[c]
		words = nc.split()
		return (
			0 if "NUMERO" in words else 1,
			0 if ("DOCUMENTO" in words or "IDENTIFICACION" in words) else 1,
		)
	doc_candidates.sort(key=doc_score)
	doc_col = doc_candidates[0] if doc_candidates else None

	# 2) Campos demograficos
	for c, nc in norm_cols.items():
		for campo, keywords in AFILIADO_CAMPOS:
			if campo in mapping:
				continue
			for kw in keywords:
				if kw == nc or (len(kw) >= 6 and kw in nc):
					mapping[campo] = c
					break
	return mapping, doc_col


def _buscar_afiliado(documento: str):
	"""Busca un afiliado por numero de documento en administrativo.af_afiliado."""
	cols = _descubrir_columnas_afiliado()
	if not cols:
		return None, "No se pudo acceder a la tabla administrativo.af_afiliado (verifica conexión y permisos)."

	mapping, doc_col = _mapear_columnas_afiliado(cols)
	if not doc_col:
		# Fallback: asumir columna documento
		cands = [c for c in cols if _normalizar_columna(c) in AFILIADO_DOC_KEYWORDS]
		doc_col = cands[0] if cands else None
	if not doc_col:
		return None, "No se encontró la columna de número de documento en la tabla af_afiliado."

	try:
		from .database import engine
	except ImportError:
		from database import engine
	try:
		from sqlalchemy import text
		doc = str(documento).strip()
		doc_limpio = doc.replace(" ", "").replace("-", "")
		with engine.connect() as conn:
			# Buscar afiliado y nombre de IPS en una sola consulta
			query = f'''
				SELECT a.*, i."razon_social" as ips_nombre
				FROM "{AFILIADO_ESQUEMA}"."{AFILIADO_TABLA}" a
				LEFT JOIN "{AFILIADO_ESQUEMA}"."ct_ips" i ON a."ips" = i."ips"
				WHERE a."{doc_col}" = :doc
				LIMIT 1
			'''
			row = conn.execute(text(query), {"doc": doc_limpio}).fetchone()
			if row is None and doc_limpio.isdigit():
				query = f'''
					SELECT a.*, i."razon_social" as ips_nombre
					FROM "{AFILIADO_ESQUEMA}"."{AFILIADO_TABLA}" a
					LEFT JOIN "{AFILIADO_ESQUEMA}"."ct_ips" i ON a."ips" = i."ips"
					WHERE a."{doc_col}" = :num
					LIMIT 1
				'''
				row = conn.execute(text(query), {"num": int(doc_limpio)}).fetchone()
			if row is None:
				return None, None
			columnas = list(row._mapping.keys())
			valores = list(row)
			return dict(zip(columnas, valores)), None
	except Exception as e:
		return None, f"Error al consultar el afiliado: {str(e)[:200]}"


def _serializar_afiliado(data: dict, mapping: dict) -> dict:
	"""Extrae los campos demograficos solicitados de la fila consultada.
	Algunos campos (telefono, correo) pueden tener varias columnas de origen;
	se usa el primer valor no vacio."""
	def get(campo, extra_cols=()):
		for col in ([mapping.get(campo)] + list(extra_cols)):
			if not col:
				continue
			v = data.get(col)
			if v is None:
				continue
			s = str(v).strip()
			if s:
				return s
		return None
	return {
		"primer_nombre": get("primer_nombre"),
		"segundo_nombre": get("segundo_nombre"),
		"primer_apellido": get("primer_apellido"),
		"segundo_apellido": get("segundo_apellido"),
		"fecha_nacimiento": get("fecha_nacimiento"),
		"sexo": get("sexo"),
		"direccion": get("direccion"),
		"correo": get("correo"),
		"categoria": get("categoria"),
		"estado_afiliado": get("estado_afiliado"),
		"tipo_afiliado": get("tipo_afiliado"),
		"fecha_inicio_cobertura": get("fecha_inicio_cobertura"),
		"municipio_afiliacion": resolver_municipio(get("municipio_afiliacion")),
		"departamento_afiliacion": resolver_departamento(get("departamento_afiliacion")),
		"discapacidad": get("discapacidad"),
		"telefono": get("telefono", extra_cols=(mapping.get("telefono_2"), mapping.get("celular"), mapping.get("celular_2"))),
		"barrio": get("barrio"),
		"ips_primaria": data.get("ips_nombre") or get("ips_primaria"),
	}


@app.get("/verificar-afiliado/{documento}")
async def verificar_afiliado(documento: str, current_user: User = Depends(get_current_user)):
	"""Consulta datos demograficos de una usuaria por numero de documento."""
	cols = _descubrir_columnas_afiliado()
	extra = {"columnas_tabla": cols}
	data, err = _buscar_afiliado(documento)
	if err:
		return {"encontrado": False, "error": err, **extra}
	if data is None:
		return {"encontrado": False, "documento": documento, **extra}
	mapping, _ = _mapear_columnas_afiliado(cols)
	afiliado = _serializar_afiliado(data, mapping)

	# Validar IPS: si el prestador tiene IPS asignada, verificar que el afiliado pertenezca a ella
	if current_user.role != "admin":
		try:
			db = SessionLocal()
			try:
				prestador = db.query(Prestador).filter(Prestador.user_id == current_user.id).first()
				if prestador and prestador.ips:
					ips_prestador_code = str(prestador.ips).strip()
					# Obtener codigo IPS del afiliado
					ips_col = mapping.get("ips_primaria")
					ips_afiliado_code = str(data.get(ips_col, "")).strip() if ips_col else ""
					if ips_afiliado_code and ips_prestador_code != ips_afiliado_code:
						return {
							"encontrado": True,
							"documento": documento,
							"afiliado": afiliado,
							"restriction": "ips_no_coincide",
							"message": f"Esta usuaria pertenece a otra IPS ({afiliado.get('ips_primaria', 'desconocida')}). Tu IPS asignada tiene código {ips_prestador_code}.",
							**extra,
						}
			finally:
				db.close()
		except Exception:
			pass

	return {"encontrado": True, "documento": documento, "afiliado": afiliado, **extra}


@app.post("/validate-affiliation")
async def validate_affiliation(payload: dict, current_user: User = Depends(get_current_user)):
	"""Valida afiliación institucional: tipo+numero vs administrativo.af_afiliado.
	Si no se envía corrected_text, lee del último cargue en la BD."""
	import base64 as _b64, gzip as _gzip
	ensure_db_ready()
	
	corrected_text = payload.get("corrected_text", "")
	
	# Si no viene texto, leer del último cargue
	if not corrected_text or not corrected_text.strip():
		try:
			db = SessionLocal()
			from sqlalchemy import text as sa_text
			cargues = db.query(Cargue).filter(Cargue.template_key == "gestante").order_by(Cargue.id.desc()).limit(1).all()
			if cargues:
				c = cargues[0]
				texto = c.corrected_text or c.raw_text or ""
				if c.compressed and texto:
					try:
						texto = _gzip.decompress(_b64.b64decode(texto)).decode("utf-8", errors="replace")
					except Exception:
						pass
				corrected_text = texto
			db.close()
		except Exception:
			pass
	
	if not corrected_text or not corrected_text.strip():
		return {"success": True, "encontrados": 0, "no_encontrados": 0, "errors": [], "valid_users": [], "ips_groups": {}, "info": "No hay datos para validar"}
	
	try:
		df = pd.read_csv(io.StringIO(corrected_text), sep='|', header=None, dtype=str, engine='python', keep_default_na=False)
		df = df.fillna('').astype(str)
	except Exception as e:
		return {"success": True, "encontrados": 0, "no_encontrados": 0, "errors": [], "valid_users": [], "ips_groups": {}, "info": f"Error parseando datos: {str(e)[:200]}"}
	
	if df.empty:
		return {"success": True, "encontrados": 0, "no_encontrados": 0, "errors": [], "valid_users": [], "ips_groups": {}, "info": "Sin datos"}
	
	template_key = payload.get("template_key", "gestante")
	meta = get_template_by_key(template_key)
	tmpl = meta["template"]
	tmpl_names = [t['name'] for t in tmpl]
	has_cols = len(df.columns) == len(tmpl_names)
	if has_cols:
		df.columns = tmpl_names
	
	if has_cols:
		tipo_col = tmpl_names[1]
		num_col = tmpl_names[2]
		nombre1_col = tmpl_names[5]
		apellido1_col = tmpl_names[3]
		nombre2_col = tmpl_names[6]
		apellido2_col = tmpl_names[4]
	else:
		tipo_col = 1
		num_col = 2
		nombre1_col = 5
		apellido1_col = 3
		nombre2_col = 6
		apellido2_col = 4
	
	usuarios = []
	for idx, row in df.iterrows():
		tipo_id = str(row.get(tipo_col, "")).strip().upper()
		num_id = str(row.get(num_col, "")).strip()
		nombre1 = str(row.get(nombre1_col, "")).strip()
		apellido1 = str(row.get(apellido1_col, "")).strip()
		nombre2 = str(row.get(nombre2_col, "")).strip()
		apellido2 = str(row.get(apellido2_col, "")).strip()
		if tipo_id and num_id and tipo_id != "SIN DATO" and num_id != "0":
			usuarios.append({
				"row_idx": idx,
				"tipo_id": tipo_id,
				"numero_id": num_id,
				"nombre1": nombre1,
				"nombre2": nombre2,
				"apellido1": apellido1,
				"apellido2": apellido2,
			})
	
	if not usuarios:
		return {"success": True, "encontrados": 0, "no_encontrados": 0, "errors": [], "valid_users": [], "ips_groups": {}}
	
	# Validar contra af_afiliado (consultas por lotes)
	try:
		from .corporate_db import validar_afiliados_lote, obtener_nombres_ips
	except ImportError:
		try:
			from corporate_db import validar_afiliados_lote, obtener_nombres_ips
		except ImportError:
			return {"success": True, "encontrados": 0, "no_encontrados": 0, "errors": [], "valid_users": [], "ips_groups": {}, "info": "Modulo corporativo no disponible"}
	
	try:
		lote_input = [{"tipo_id": u["tipo_id"], "numero_id": u["numero_id"]} for u in usuarios]
		resultado_lote = validar_afiliados_lote(lote_input)
	except Exception as e:
		return {"success": True, "encontrados": 0, "no_encontrados": len(usuarios), "errors": [{"row": 0, "column": "DB", "original": "", "corrected": f"Error consultando BD corporativa: {str(e)[:200]}", "status": "error"}], "valid_users": [], "ips_groups": {}, "info": f"Error BD corporativa: {str(e)[:200]}"}
	
	if resultado_lote.get("error"):
		return {"success": True, "encontrados": 0, "no_encontrados": len(usuarios), "errors": [{"row": 0, "column": "DB", "original": "", "corrected": f"Error BD corporativa: {resultado_lote['error'][:200]}", "status": "error"}], "valid_users": [], "ips_groups": {}, "info": f"Error BD: {resultado_lote['error'][:200]}"}
	
	# Indexar encontrados por (tipo, numero) -> ips_code
	indx_encontrados = {}
	for enc in resultado_lote["encontrados"]:
		key = (enc["tipo_id"], enc["numero_id"])
		indx_encontrados[key] = enc.get("ips")
	
	# Separar encontrados y no encontrados
	encontrados = []
	no_encontrados = []
	errors = []
	
	for u in usuarios:
		key = (u["tipo_id"], u["numero_id"])
		if key in indx_encontrados:
			ips_code = indx_encontrados[key]
			u_with_ips = {**u, "ips_code": ips_code}
			encontrados.append(u_with_ips)
		else:
			no_encontrados.append(u)
			errors.append({
				"row": u["row_idx"] + 2,
				"column": f"{tipo_col} / {num_col}",
				"original": f"{u['tipo_id']} {u['numero_id']}",
				"corrected": f"Usuaria no encontrada en la base de afiliados. Tipo: {u['tipo_id']}, Número: {u['numero_id']}",
				"status": "error",
			})
	
	# Obtener nombres de IPS desde ct_ips
	ips_codes = list({e["ips_code"] for e in encontrados if e.get("ips_code")})
	try:
		nombres_ips = obtener_nombres_ips(ips_codes) if ips_codes else {}
	except Exception:
		nombres_ips = {}
	
	# Construir grupos de IPS
	ips_groups = {}
	for u in encontrados:
		ips_code = u.get("ips_code")
		if ips_code:
			ips_name = nombres_ips.get(str(ips_code).strip(), f"IPS {ips_code}")
			if ips_name not in ips_groups:
				ips_groups[ips_name] = []
			ips_groups[ips_name].append({
				"tipo_id": u["tipo_id"],
				"numero_id": u["numero_id"],
				"nombre1": u["nombre1"],
				"nombre2": u["nombre2"],
				"apellido1": u["apellido1"],
				"apellido2": u["apellido2"],
			})
	
	return {
		"success": True,
		"encontrados": len(encontrados),
		"no_encontrados": len(no_encontrados),
		"errors": errors,
		"valid_users": [
			{"tipo_id": e["tipo_id"], "numero_id": e["numero_id"], "ips_code": e.get("ips_code"),
			 "nombre1": e["nombre1"], "nombre2": e["nombre2"], "apellido1": e["apellido1"], "apellido2": e["apellido2"]}
			for e in encontrados
		],
		"ips_groups": ips_groups,
	}


# ─── Gestión de data (ver, editar, crear registros de gestantes) ───────────

@app.get("/ips")
async def listar_ips(current_user: User = Depends(get_current_user)):
	"""Devuelve lista de IPS primarias para dropdown. Admin ve todas; prestador solo ve su IPS."""
	ensure_db_ready()
	db = SessionLocal()
	try:
		ips_filtro = None
		if current_user.role != "admin":
			prestador = db.query(Prestador).filter(Prestador.user_id == current_user.id).first()
			if prestador and prestador.ips:
				ips_filtro = str(prestador.ips).strip().upper()

		from sqlalchemy import text as sa_text
		if ips_filtro:
			rows = db.execute(sa_text(
				'SELECT DISTINCT "NOMBRE_DE_LA_IPS_PRIMARIA" FROM gestantes '
				'WHERE UPPER("NOMBRE_DE_LA_IPS_PRIMARIA") = :ips AND "NOMBRE_DE_LA_IPS_PRIMARIA" IS NOT NULL '
				'ORDER BY "NOMBRE_DE_LA_IPS_PRIMARIA"'
			), {"ips": ips_filtro}).fetchall()
		else:
			rows = db.execute(sa_text(
				'SELECT DISTINCT "NOMBRE_DE_LA_IPS_PRIMARIA" FROM gestantes '
				'WHERE "NOMBRE_DE_LA_IPS_PRIMARIA" IS NOT NULL AND "NOMBRE_DE_LA_IPS_PRIMARIA" != \'\' '
				'ORDER BY "NOMBRE_DE_LA_IPS_PRIMARIA"'
			)).fetchall()

		ips_list = [str(r[0]).strip() for r in rows if r[0]]
		return {"ips": ips_list}
	except Exception as e:
		return {"ips": [], "error": str(e)[:200]}
	finally:
		db.close()


@app.get("/data/gestantes/diagnostico")
async def diagnosticar_gestantes(current_user: User = Depends(get_current_user)):
	"""Diagnostico: muestra valores unicos de NOMBRE_DE_LA_IPS_PRIMARIA y un sample de registros."""
	ensure_db_ready()
	db = SessionLocal()
	try:
		from sqlalchemy import text as sa_text
		# Valores unicos de IPS
		rows_ips = db.execute(sa_text('''
			SELECT "NOMBRE_DE_LA_IPS_PRIMARIA", COUNT(*) as total
			FROM gestantes
			WHERE "NOMBRE_DE_LA_IPS_PRIMARIA" IS NOT NULL
			GROUP BY "NOMBRE_DE_LA_IPS_PRIMARIA"
			ORDER BY COUNT(*) DESC
			LIMIT 30
		''')).fetchall()
		ips_vals = [{"valor": str(r[0]), "total": int(r[1])} for r in rows_ips]

		# Sample de primeros 5 registros con todas sus columnas
		sample_rows = db.execute(sa_text('SELECT * FROM gestantes LIMIT 5')).fetchall()
		sample_cols = [c.name for c in db.execute(sa_text('SELECT * FROM gestantes WHERE 1=0')).cursor.description]
		sample = []
		for row in sample_rows:
			sample.append({col: str(row[i])[:80] for i, col in enumerate(sample_cols)})

		# Columnas 20-35 del sample (rango donde cae IPS)
		return {"ips_valores": ips_vals, "sample_columns": sample_cols[20:35], "sample_data": sample}
	except Exception as e:
		return {"error": str(e)[:300]}
	finally:
		db.close()


@app.delete("/data/gestantes/clean")
async def clean_gestantes(current_user: User = Depends(get_current_user)):
	"""Elimina todos los registros de la tabla gestantes."""
	ensure_db_ready()
	db = SessionLocal()
	try:
		from sqlalchemy import text as sa_text
		count_result = db.execute(sa_text('SELECT COUNT(*) FROM gestantes')).scalar()
		db.execute(sa_text('DELETE FROM gestantes'))
		db.commit()
		return {"ok": True, "eliminados": int(count_result or 0)}
	except Exception as e:
		db.rollback()
		return {"error": str(e)[:300]}
	finally:
		db.close()


@app.get("/data/gestantes/ips-grupos")
async def listar_ips_grupos(current_user: User = Depends(get_current_user)):
	"""Devuelve lista de IPS primarias con conteo de registros.
	Admin ve todas; prestador solo ve su IPS."""
	ensure_db_ready()
	db = SessionLocal()
	try:
		# Para prestadores, filtrar por su IPS
		ips_filtro = None
		if current_user.role != "admin":
			prestador = db.query(Prestador).filter(Prestador.user_id == current_user.id).first()
			if prestador and prestador.ips:
				ips_filtro = str(prestador.ips).strip().upper()

		# Valores de IPS que son claramente invalidos (confundidos con otra columna)
		IPS_INVALIDOS = {"NO", "SI", "N/A", "NA", "SIN IPS", "S/N", "-", "0", "NO APLICA"}

		sql = '''SELECT "NOMBRE_DE_LA_IPS_PRIMARIA", COUNT(*) as total
				 FROM gestantes
				 WHERE "NOMBRE_DE_LA_IPS_PRIMARIA" IS NOT NULL
				   AND "NOMBRE_DE_LA_IPS_PRIMARIA" != ''
				   AND UPPER(TRIM("NOMBRE_DE_LA_IPS_PRIMARIA")) NOT IN :invalidos
				 GROUP BY "NOMBRE_DE_LA_IPS_PRIMARIA"
				 ORDER BY "NOMBRE_DE_LA_IPS_PRIMARIA"'''
		
		if ips_filtro:
			# Prestador: solo ver su IPS
			sql = '''SELECT "NOMBRE_DE_LA_IPS_PRIMARIA", COUNT(*) as total
					 FROM gestantes
					 WHERE UPPER("NOMBRE_DE_LA_IPS_PRIMARIA") = UPPER(:ips)
					 GROUP BY "NOMBRE_DE_LA_IPS_PRIMARIA"
					 ORDER BY "NOMBRE_DE_LA_IPS_PRIMARIA"'''

		from sqlalchemy import text as sa_text
		params = {"ips": ips_filtro, "invalidos": tuple(IPS_INVALIDOS)} if ips_filtro else {"invalidos": tuple(IPS_INVALIDOS)}
		rows = db.execute(sa_text(sql), params).fetchall()
		ips_list = [{"nombre": str(r[0]).strip(), "total": int(r[1])} for r in rows]
		return {"ips": ips_list}
	except Exception as e:
		return {"error": str(e)[:300], "ips": []}
	finally:
		db.close()


@app.post("/data/gestantes/populate")
async def populate_gestantes_from_cargues(current_user: User = Depends(require_admin)):
	"""Lee el ultimo cargue y puebla la tabla gestantes.
	El correctedText NO tiene headers (generado con header=False).
	TODAS las lineas son datos — no se salta ninguna."""
	ensure_db_ready()
	db = SessionLocal()
	try:
		crear_tabla_gestantes()

		from sqlalchemy import text as sa_text
		import base64, gzip, io as _io
		import pandas as _pd

		cargues = db.query(Cargue).filter(Cargue.template_key == "gestante").order_by(Cargue.id.desc()).limit(1).all()
		if not cargues:
			return {"error": "No hay cargues de gestantes", "insertadas": 0}

		cargue = cargues[0]
		db_cols = GESTANTE_COLUMNS
		try:
			real_cols = [c.name for c in db.execute(sa_text('SELECT * FROM gestantes WHERE 1=0')).cursor.description]
		except Exception as e:
			return {"error": f"No se pudo leer columnas de gestantes: {str(e)[:200]}", "insertadas": 0}

		texto = cargue.corrected_text or cargue.raw_text or ""
		if not texto:
			return {"error": f"Cargue {cargue.id} vacio", "insertadas": 0}

		if cargue.compressed:
			try:
				texto = gzip.decompress(base64.b64decode(texto)).decode("utf-8", errors="replace")
			except Exception as e:
				return {"error": f"Cargue {cargue.id}: error descomprimiendo: {str(e)[:200]}", "insertadas": 0}

		# ── Leer correctedText como DataFrame (sin headers, lines may vary in length) ──
		try:
			meta = get_template_by_key("gestante")
			tmpl = meta["template"]
			tmpl_names = [t["name"] for t in tmpl]
		except Exception:
			tmpl_names = []

		# Leer SIN headers. correctedText NO tiene fila de headers.
		df = _pd.read_csv(_io.StringIO(texto), sep='|', header=None, dtype=str, engine='python', keep_default_na=False)
		df = df.fillna('').astype(str)

		# Asignar nombres de columnas del template si la cantidad coincide
		has_template_names = len(df.columns) == len(tmpl_names)
		if has_template_names:
			df.columns = tmpl_names

		# ── Mapear nombre de template → nombre de columna DB ──
		def _norm(s):
			import unicodedata
			s = str(s).strip()
			s = unicodedata.normalize('NFKD', s).encode('ascii', 'ignore').decode('ascii')
			s = s.upper()
			s = ''.join(c for c in s if c.isalnum() or c in (' ', '_'))
			s = s.strip()
			s = '_'.join(s.split())
			return s

		norm_to_db = {}
		for col in db_cols:
			norm_to_db[_norm(col)] = col

		def _find_db_col(tmpl_name):
			n = _norm(tmpl_name)
			if n in norm_to_db:
				return norm_to_db[n]
			for db_norm, db_col in norm_to_db.items():
				if len(n) >= 3 and len(db_norm) >= 3:
					if n in db_norm or db_norm in n:
						return db_col
			return None

		# Construir mapping: si tenemos nombres de template, usarlos; si no, usar indice
		tmpl_to_db = {}
		if has_template_names:
			for cname in df.columns:
				db_match = _find_db_col(cname)
				if db_match:
					tmpl_to_db[cname] = db_match
		else:
			# Fallback: mapear por indice. Position i in CSV → template[i] → DB col
			for i in range(min(len(df.columns), len(tmpl_names), len(db_cols))):
				db_match = _find_db_col(tmpl_names[i]) if i < len(tmpl_names) else None
				if db_match:
					tmpl_to_db[i] = db_match

		mapped_count = len(tmpl_to_db)

		# ── Mapeo de nombres de IPS a prestadores ──
		ips_to_prestador = {}
		prestadores = db.query(Prestador).all()
		for prest in prestadores:
			ips_key = str(prest.ips).strip().upper() if prest.ips else ""
			if ips_key:
				ips_to_prestador[ips_key] = prest.id
			if prest.nombre:
				name_key = str(prest.nombre).strip().upper()
				if name_key not in ips_to_prestador:
					ips_to_prestador[name_key] = prest.id

		# ── Insertar filas ──
		total_insertadas = 0
		total_errores = []
		mapped_count = len(tmpl_to_db)

		for idx, row in df.iterrows():
			registro = {}
			for tmpl_name, val in row.items():
				db_col = tmpl_to_db.get(tmpl_name)
				if db_col:
					registro[db_col] = str(val).strip()

			ips_primaria = registro.get("NOMBRE_DE_LA_IPS_PRIMARIA", "").strip().upper()
			prestador_id = ips_to_prestador.get(ips_primaria)

			registro["prestador_id"] = prestador_id if prestador_id else (cargue.prestador_id if cargue.prestador_id else None)
			registro["user_id"] = cargue.user_id if cargue.user_id else None
			registro["mes"] = cargue.mes or None
			registro["original_filename"] = cargue.original_filename or None

			try:
				cols_validas = [c for c in real_cols if c in registro and c != "id"]
				if not cols_validas:
					continue
				placeholders = ", ".join(f':{c}' for c in cols_validas)
				col_names = ", ".join(f'"{c}"' for c in cols_validas)
				params_ins = {c: registro.get(c) if registro.get(c) != "" else None for c in cols_validas}
				db.execute(sa_text(f'INSERT INTO gestantes ({col_names}) VALUES ({placeholders})'), params_ins)
				total_insertadas += 1
			except Exception as e:
				total_errores.append(f"Fila {idx+1}: {str(e)[:100]}")
				if len(total_errores) >= 10:
					break

		db.commit()
		return {
			"ok": True,
			"cargue_id": cargue.id,
			"total_lineas": len(df),
			"insertadas": total_insertadas,
			"headers_mapeados": mapped_count,
			"total_template_cols": len(tmpl_names),
			"errores": total_errores[:10],
			"debug": {
				"total_lineas": len(df),
				"db_cols_count": len(db_cols),
				"tmpl_cols_count": len(tmpl_names),
				"mapped_count": mapped_count,
				"sample_mapping": list(tmpl_to_db.items())[:10],
			},
		}
	except Exception as e:
		db.rollback()
		return {"error": str(e)[:300], "insertadas": 0}
	finally:
		db.close()


@app.post("/data/gestantes/clean-repopulate")
async def clean_and_repopulate(current_user: User = Depends(require_admin)):
	"""Limpia la tabla gestantes y re-puebla desde el ultimo cargue.
	Misma logica que populate pero primero borra todo."""
	ensure_db_ready()
	db = SessionLocal()
	try:
		from sqlalchemy import text as sa_text
		import base64, gzip, io as _io
		import pandas as _pd

		cargues = db.query(Cargue).filter(Cargue.template_key == "gestante").order_by(Cargue.id.desc()).limit(1).all()
		if not cargues:
			return {"error": "No hay cargues", "insertadas": 0}

		cargue = cargues[0]
		db_cols = GESTANTE_COLUMNS

		texto = cargue.corrected_text or cargue.raw_text or ""
		if not texto:
			return {"error": f"Cargue {cargue.id} vacio", "insertadas": 0}

		if cargue.compressed:
			try:
				texto = gzip.decompress(base64.b64decode(texto)).decode("utf-8", errors="replace")
			except Exception as e:
				return {"error": f"Error descomprimiendo: {str(e)[:200]}", "insertadas": 0}

		# Leer como DataFrame (sin headers)
		try:
			meta = get_template_by_key("gestante")
			tmpl = meta["template"]
			tmpl_names = [t["name"] for t in tmpl]
		except Exception:
			tmpl_names = []

		df = _pd.read_csv(_io.StringIO(texto), sep='|', header=None, dtype=str, engine='python', keep_default_na=False)
		df = df.fillna('').astype(str)

		has_template_names = len(df.columns) == len(tmpl_names)
		if has_template_names:
			df.columns = tmpl_names

		def _norm(s):
			import unicodedata
			s = str(s).strip()
			s = unicodedata.normalize('NFKD', s).encode('ascii', 'ignore').decode('ascii')
			s = s.upper()
			s = ''.join(c for c in s if c.isalnum() or c in (' ', '_'))
			s = s.strip()
			s = '_'.join(s.split())
			return s

		norm_to_db = {}
		for col in db_cols:
			norm_to_db[_norm(col)] = col

		def _find_db_col(tmpl_name):
			n = _norm(tmpl_name)
			if n in norm_to_db:
				return norm_to_db[n]
			for db_norm, db_col in norm_to_db.items():
				if len(n) >= 3 and len(db_norm) >= 3:
					if n in db_norm or db_norm in n:
						return db_col
			return None

		tmpl_to_db = {}
		if has_template_names:
			for cname in df.columns:
				db_match = _find_db_col(cname)
				if db_match:
					tmpl_to_db[cname] = db_match
		else:
			for i in range(min(len(df.columns), len(db_cols))):
				tmpl_to_db[i] = db_cols[i]

		# Limpiar tabla
		db.execute(sa_text('DELETE FROM gestantes'))
		db.flush()

		# Mapeo IPS a prestadores
		ips_to_prestador = {}
		prestadores = db.query(Prestador).all()
		for prest in prestadores:
			if prest.ips:
				ips_to_prestador[str(prest.ips).strip().upper()] = prest.id
			if prest.nombre:
				name_key = str(prest.nombre).strip().upper()
				if name_key not in ips_to_prestador:
					ips_to_prestador[name_key] = prest.id

		try:
			real_cols = [c.name for c in db.execute(sa_text('SELECT * FROM gestantes WHERE 1=0')).cursor.description]
		except:
			real_cols = db_cols

		total_insertadas = 0
		errores = []

		for idx, row in df.iterrows():
			registro = {}
			for tmpl_name, val in row.items():
				db_col = tmpl_to_db.get(tmpl_name)
				if db_col:
					registro[db_col] = str(val).strip()

			ips_primaria = registro.get("NOMBRE_DE_LA_IPS_PRIMARIA", "").strip().upper()
			prestador_id = ips_to_prestador.get(ips_primaria)

			registro["prestador_id"] = prestador_id if prestador_id else (cargue.prestador_id if cargue.prestador_id else None)
			registro["user_id"] = cargue.user_id if cargue.user_id else None
			registro["mes"] = cargue.mes or None
			registro["original_filename"] = cargue.original_filename or None

			try:
				cols_validas = [c for c in real_cols if c in registro and c != "id"]
				if not cols_validas:
					continue
				placeholders = ", ".join(f':{c}' for c in cols_validas)
				col_names = ", ".join(f'"{c}"' for c in cols_validas)
				params_ins = {c: registro.get(c) if registro.get(c) != "" else None for c in cols_validas}
				db.execute(sa_text(f'INSERT INTO gestantes ({col_names}) VALUES ({placeholders})'), params_ins)
				total_insertadas += 1
			except Exception as e:
				errores.append(f"Fila {idx+1}: {str(e)[:100]}")
				if len(errores) >= 10:
					break

		db.commit()

		return {
			"ok": True,
			"insertadas": total_insertadas,
			"errores": len(errores),
			"muestra_errores": errores[:5],
			"headers_mapeados": len(tmpl_to_db),
			"total_template_cols": len(tmpl_names),
		}
	except Exception as e:
		db.rollback()
		return {"error": str(e)[:400], "insertadas": 0}
	finally:
		db.close()


@app.get("/data/gestantes")
async def listar_gestantes(
	current_user: User = Depends(get_current_user),
	page: int = 1,
	page_size: int = 50,
	search: str = "",
	ips: str = "",
):
	"""Lista registros de gestantes. Admin ve todos, prestador solo los de su IPS."""
	ensure_db_ready()
	db = SessionLocal()
	try:
		# Obtener IPS del prestador actual
		ips_filtro = None
		if current_user.role != "admin":
			prestador = db.query(Prestador).filter(Prestador.user_id == current_user.id).first()
			if prestador and prestador.ips:
				ips_filtro = str(prestador.ips).strip().upper()

		# Si se pasa el param ips explicitamente, usarlo (para admin que quiere filtrar por IPS)
		if ips and current_user.role == "admin":
			ips_filtro = str(ips).strip().upper()

		offset = (max(1, page) - 1) * page_size
		from sqlalchemy import text

		# Contar total
		count_sql = 'SELECT COUNT(*) FROM gestantes WHERE 1=1'
		count_params = {}
		if ips_filtro:
			count_sql += ' AND UPPER("NOMBRE_DE_LA_IPS_PRIMARIA") = :ips'
			count_params["ips"] = ips_filtro
		if search:
			count_sql += ' AND ("NO_DE_IDENTIFICACION" ILIKE :q OR "APELLIDO_1" ILIKE :q OR "NOMBRE_1" ILIKE :q)'
			count_params["q"] = f"%{search}%"

		total = db.execute(text(count_sql), count_params).scalar() or 0

		# Obtener registros
		query_sql = 'SELECT * FROM gestantes WHERE 1=1'
		if ips_filtro:
			query_sql += ' AND UPPER("NOMBRE_DE_LA_IPS_PRIMARIA") = :ips'
		if search:
			query_sql += ' AND ("NO_DE_IDENTIFICACION" ILIKE :q OR "APELLIDO_1" ILIKE :q OR "NOMBRE_1" ILIKE :q)'
		query_sql += ' ORDER BY id DESC LIMIT :limit OFFSET :offset'
		params = {"limit": page_size, "offset": offset}
		if ips_filtro:
			params["ips"] = ips_filtro
		if search:
			params["q"] = f"%{search}%"

		rows = db.execute(text(query_sql), params).fetchall()
		columnas = [c.name for c in db.execute(text('SELECT * FROM gestantes WHERE 1=0')).cursor.description]
		registros = []
		for row in rows:
			registros.append(dict(zip(columnas, [str(v) if v is not None else "" for v in row])))

		return {"registros": registros, "total": total, "page": page, "page_size": page_size}
	except Exception as e:
		return {"error": str(e)[:300], "registros": [], "total": 0}
	finally:
		db.close()


def _get_prestador_ips_name(db, current_user):
	"""Obtiene el nombre de la IPS del prestador actual. Retorna None si es admin o no tiene IPS."""
	if current_user.role == "admin":
		return None
	prestador = db.query(Prestador).filter(Prestador.user_id == current_user.id).first()
	if not prestador or not prestador.ips:
		return None
	ips_code = str(prestador.ips).strip()
	try:
		row = db.execute(text('SELECT razon_social FROM ct_ips WHERE ips = :code'), {"code": ips_code}).fetchone()
		if row:
			return str(row[0]).strip().upper()
	except Exception:
		pass
	return None


def _check_gestante_ips(db, current_user, registro_id):
	"""Verifica que la gestante pertenezca a la IPS del prestador. Lanza 403 si no."""
	ips_nombre = _get_prestador_ips_name(db, current_user)
	if ips_nombre is None:
		return
	row = db.execute(text('SELECT "NOMBRE_DE_LA_IPS_PRIMARIA" FROM gestantes WHERE id = :id'), {"id": registro_id}).fetchone()
	if not row:
		return
	ips_gestante = str(row[0] or "").strip().upper()
	if ips_gestante and ips_gestante != ips_nombre and ips_gestante != "NA":
		raise HTTPException(status_code=403, detail="No autorizado: esta gestante pertenece a otra IPS")


@app.get("/data/gestantes/{registro_id}")
async def obtener_gestante(registro_id: int, current_user: User = Depends(get_current_user)):
	"""Obtiene un registro de gestante por ID."""
	ensure_db_ready()
	db = SessionLocal()
	try:
		from sqlalchemy import text
		_check_gestante_ips(db, current_user, registro_id)
		row = db.execute(text('SELECT * FROM gestantes WHERE id = :id'), {"id": registro_id}).fetchone()
		if not row:
			raise HTTPException(status_code=404, detail="Registro no encontrado")
		columnas = [c.name for c in db.execute(text('SELECT * FROM gestantes WHERE 1=0')).cursor.description]
		registro = dict(zip(columnas, [str(v) if v is not None else "" for v in row]))
		return registro
	except HTTPException:
		raise
	except Exception as e:
		raise HTTPException(status_code=500, detail=str(e)[:300])
	finally:
		db.close()


@app.put("/data/gestantes/{registro_id}")
async def actualizar_gestante(registro_id: int, payload: dict, current_user: User = Depends(get_current_user)):
	"""Actualiza un registro de gestante. Valida contra el instructivo. Registra auditoria."""
	ensure_db_ready()
	db = SessionLocal()
	try:
		from sqlalchemy import text
		_check_gestante_ips(db, current_user, registro_id)
		# Verificar que existe y obtener valores actuales
		existing = db.execute(text('SELECT * FROM gestantes WHERE id = :id'), {"id": registro_id}).fetchone()
		if not existing:
			raise HTTPException(status_code=404, detail="Registro no encontrado")

		# Obtener columnas para comparar
		columnas = [c.name for c in db.execute(text('SELECT * FROM gestantes WHERE 1=0')).cursor.description]
		current_data = dict(zip(columnas, [str(v) if v is not None else "" for v in existing]))

		# Validar campos contra el instructivo
		try:
			from .gestante_config import RAW_FIELDS
		except ImportError:
			from gestante_config import RAW_FIELDS

		errores = []
		for i, (col_name, col_type) in enumerate(RAW_FIELDS):
			if col_name in payload:
				val = str(payload[col_name]).strip()
				if col_type == "SET" and val and val != "NA":
					try:
						from .gestante_config import get_gestante_template
					except ImportError:
						from gestante_config import get_gestante_template
					tmpl = get_gestante_template()
					for t in tmpl:
						if t["name"] == col_name and "allowed" in t:
							if val not in t["allowed"]:
								errores.append(f"{col_name}: '{val}' no es una opcion valida")
							break

		if errores:
			raise HTTPException(status_code=400, detail="; ".join(errores[:10]))

		# Detectar cambios y registrar auditoria
		audit_entries = []
		for key, val in payload.items():
			new_val = str(val) if val is not None else ""
			old_val = current_data.get(key, "")
			if new_val != old_val:
				audit_entries.append({
					"user_id": current_user.id,
					"username": current_user.username,
					"gestante_id": registro_id,
					"action": "UPDATE",
					"field_name": key,
					"old_value": old_val[:500] if old_val else "",
					"new_value": new_val[:500] if new_val else "",
				})

		# Actualizar
		set_parts = []
		params = {"id": registro_id}
		for key, val in payload.items():
			set_parts.append(f'"{key}" = :{key}')
			params[key] = str(val) if val is not None else ""

		if set_parts:
			update_sql = f'UPDATE gestantes SET {", ".join(set_parts)} WHERE id = :id'
			db.execute(text(update_sql), params)

			# Insertar registros de auditoria
			for entry in audit_entries:
				db.execute(text(
					'INSERT INTO audit_logs (user_id, username, gestante_id, action, field_name, old_value, new_value) '
					'VALUES (:user_id, :username, :gestante_id, :action, :field_name, :old_value, :new_value)'
				), entry)

			db.commit()

		return {"success": True, "audit_count": len(audit_entries)}
	except HTTPException:
		raise
	except Exception as e:
		db.rollback()
		raise HTTPException(status_code=500, detail=str(e)[:300])
	finally:
		db.close()


@app.post("/data/gestantes")
async def crear_gestante(payload: dict, current_user: User = Depends(get_current_user)):
	"""Crea un registro individual de gestante (cargue de uno en uno)."""
	ensure_db_ready()
	db = SessionLocal()
	try:
		# Obtener IPS del prestador
		ips_prestador = ""
		if current_user.role != "admin":
			prestador = db.query(Prestador).filter(Prestador.user_id == current_user.id).first()
			if prestador and prestador.ips:
				ips_prestador = str(prestador.ips).strip()

		# Si el prestador tiene IPS, forzar que la gestante sea de esa IPS
		if ips_prestador and "NOMBRE_DE_LA_IPS_PRIMARIA" in payload:
			ips_gestante = str(payload["NOMBRE_DE_LA_IPS_PRIMARIA"]).strip().upper()
			# Buscar nombre real de la IPS
			ips_nombre_real = None
			try:
				from sqlalchemy import text as sa_text
				row = db.execute(sa_text('SELECT "razon_social" FROM "administrativo"."ct_ips" WHERE "ips" = :cod LIMIT 1'), {"cod": ips_prestador}).fetchone()
				if row:
					ips_nombre_real = str(row[0]).strip().upper()
			except Exception:
				pass
			if ips_nombre_real and ips_gestante and ips_gestante != ips_nombre_real:
				raise HTTPException(status_code=400, detail=f"No se puede crear: la gestante pertenece a otra IPS ({ips_gestante}). Tu IPS es: {ips_nombre_real}")

		from sqlalchemy import text
		columnas = [c.name for c in db.execute(text('SELECT * FROM gestantes WHERE 1=0')).cursor.description]
		# Solo usar columnas que existen en la tabla y que vienen en el payload
		cols_validas = [c for c in columnas if c in payload and c not in ("id", "created_at")]
		if not cols_validas:
			raise HTTPException(status_code=400, detail="No se enviaron campos válidos para guardar")

		placeholders = ", ".join(f':{c}' for c in cols_validas)
		col_names = ", ".join(f'"{c}"' for c in cols_validas)
		params = {c: str(payload.get(c, "")) for c in cols_validas}

		insert_sql = f'INSERT INTO gestantes ({col_names}) VALUES ({placeholders})'
		db.execute(text(insert_sql), params)
		db.commit()
		return {"success": True}
	except HTTPException:
		raise
	except Exception as e:
		db.rollback()
		raise HTTPException(status_code=500, detail=str(e)[:300])
	finally:
		db.close()


@app.delete("/data/gestantes/{registro_id}")
async def eliminar_gestante(registro_id: int, current_user: User = Depends(get_current_user)):
	"""Elimina un registro de gestante."""
	if current_user.role != "admin":
		raise HTTPException(status_code=403, detail="Solo el admin puede eliminar registros")
	ensure_db_ready()
	db = SessionLocal()
	try:
		from sqlalchemy import text
		existing = db.execute(text('SELECT id FROM gestantes WHERE id = :id'), {"id": registro_id}).fetchone()
		if not existing:
			raise HTTPException(status_code=404, detail="Registro no encontrado")
		db.execute(text('DELETE FROM gestantes WHERE id = :id'), {"id": registro_id})
		db.commit()
		return {"success": True}
	except HTTPException:
		raise
	except Exception as e:
		db.rollback()
		raise HTTPException(status_code=500, detail=str(e)[:300])
	finally:
		db.close()


@app.get("/data/gestantes/{registro_id}/audit")
async def obtener_auditoria(registro_id: int, current_user: User = Depends(get_current_user)):
	"""Obtiene el historial de auditoria de una gestante."""
	ensure_db_ready()
	db = SessionLocal()
	try:
		from sqlalchemy import text
		rows = db.execute(text(
			'SELECT * FROM audit_logs WHERE gestante_id = :gid ORDER BY created_at DESC LIMIT 100'
		), {"gid": registro_id}).fetchall()
		columnas = [c.name for c in db.execute(text('SELECT * FROM audit_logs WHERE 1=0')).cursor.description]
		logs = [dict(zip(columnas, [str(v) if v is not None else "" for v in row])) for row in rows]
		return {"logs": logs}
	except Exception as e:
		return {"logs": [], "error": str(e)[:200]}
	finally:
		db.close()


@app.post("/data/gestantes/caso-cerrado/auto-fill")
async def auto_fill_caso_cerrado(current_user: User = Depends(require_admin)):
	"""Auto-llena el campo CASO_CERRADO para gestantes que cumplen los criterios."""
	ensure_db_ready()
	db = SessionLocal()
	try:
		from sqlalchemy import text
		from datetime import datetime, date

		# Obtener el mes de reporte actual
		fecha_reporte = datetime.now()
		mes_reporte = fecha_reporte.strftime("%Y-%m")

		# Criterios:
		# 1. Al menos 10 meses entre el mes de la gestante y el mes de reporte
		# 2. ULTIMO_CONTROL_PRENATAL < mes de reporte
		# 3. FECHA (aborto) no es comodín AND ULTIMO_CONTROL_PRENATAL < FECHA
		#    O FECHA_DE_PARTO no es comodín AND ULTIMO_CONTROL_PRENATAL < FECHA_DE_PARTO

		# Buscar gestantes que cumplen los criterios
		query = text(f'''
			UPDATE gestantes SET CASO_CERRADO = TRUE
			WHERE CASO_CERRADO = FALSE
			AND (SELECT strftime('%Y-%m', created_at) FROM gestantes WHERE id = gestantes.id) <= ?
			AND ULTIMO_CONTROL_PRENATAL IS NOT NULL AND ULTIMO_CONTROL_PRENATAL != ''
			AND (
				(FECHA != '0' AND FECHA != '' AND ULTIMO_CONTROL_PRENATAL < FECHA)
				OR (FECHA_DE_PARTO != '0' AND FECHA_DE_PARTO != '' AND ULTIMO_CONTROL_PRENATAL < FECHA_DE_PARTO)
			)
		''')

		# Calcular fecha límite (10 meses antes del reporte)
		from datetime import timedelta
		fecha_limite = fecha_reporte - timedelta(days=300)  # ~10 meses
		mes_limite = fecha_limite.strftime("%Y-%m")

		# SQLite version
		from sqlalchemy import text as _text
		rows = db.execute(_text(f'''
			SELECT COUNT(*) FROM gestantes
			WHERE CASO_CERRADO = FALSE
			AND (STRFTIME('%Y-%m', created_at) <= '{mes_limite}' OR mes <= '{mes_limite}')
			AND ULTIMO_CONTROL_PRENATAL IS NOT NULL AND ULTIMO_CONTROL_PRENATAL != ''
			AND (
				(FECHA != '0' AND FECHA != '' AND ULTIMO_CONTROL_PRENATAL < FECHA)
				OR (FECHA_DE_PARTO != '0' AND FECHA_DE_PARTO != '' AND ULTIMO_CONTROL_PRENATAL < FECHA_DE_PARTO)
			)
		''')).scalar()

		# Actualizar
		db.execute(_text(f'''
			UPDATE gestantes SET CASO_CERRADO = TRUE
			WHERE CASO_CERRADO = FALSE
			AND (STRFTIME('%Y-%m', created_at) <= '{mes_limite}' OR mes <= '{mes_limite}')
			AND ULTIMO_CONTROL_PRENATAL IS NOT NULL AND ULTIMO_CONTROL_PRENATAL != ''
			AND (
				(FECHA != '0' AND FECHA != '' AND ULTIMO_CONTROL_PRENATAL < FECHA)
				OR (FECHA_DE_PARTO != '0' AND FECHA_DE_PARTO != '' AND ULTIMO_CONTROL_PRENATAL < FECHA_DE_PARTO)
			)
		'''))
		db.commit()

		return {
			"success": True,
			"total_caso_cerrado": rows,
			"mes_reporte": mes_reporte,
			"criterios": {
				"mes_limite": mes_limite,
				"ultimo_control_prenatal": "debe ser anterior al mes de reporte",
				"fecha_aborto": "no puede ser comodín (0) y debe ser posterior al último control prenatal",
				"fecha_parto": "no puede ser comodín (0) y debe ser posterior al último control prenatal",
			}
		}
	except Exception as e:
		db.rollback()
		raise HTTPException(status_code=500, detail=str(e)[:300])
	finally:
		db.close()


@app.get("/data/gestantes/caso-cerrado")
async def listar_caso_cerrado(
	current_user: User = Depends(get_current_user),
	page: int = 1,
	page_size: int = 50,
	search: str = "",
):
	"""Lista gestantes marcadas como Caso Cerrado."""
	ensure_db_ready()
	db = SessionLocal()
	try:
		from sqlalchemy import text

		offset = (max(1, page) - 1) * page_size
		params = {"limit": page_size, "offset": offset}

		count_sql = 'SELECT COUNT(*) FROM gestantes WHERE CASO_CERRADO = TRUE AND 1=1'
		query_sql = 'SELECT * FROM gestantes WHERE CASO_CERRADO = TRUE AND 1=1'

		if search:
			count_sql += ' AND ("NO_DE_IDENTIFICACION" ILIKE :q OR "APELLIDO_1" ILIKE :q OR "NOMBRE_1" ILIKE :q)'
			query_sql += ' AND ("NO_DE_IDENTIFICACION" ILIKE :q OR "APELLIDO_1" ILIKE :q OR "NOMBRE_1" ILIKE :q)'
			params["q"] = f"%{search}%"

		count_sql += ' LIMIT 1'
		total = db.execute(text(count_sql), params).scalar() or 0

		query_sql += f' ORDER BY id DESC LIMIT :limit OFFSET :offset'
		rows = db.execute(text(query_sql), params).fetchall()
		columnas = [c.name for c in db.execute(text('SELECT * FROM gestantes WHERE 1=0')).cursor.description]
		registros = []
		for row in rows:
			registros.append(dict(zip(columnas, [str(v) if v is not None else "" for v in row])))

		return {"registros": registros, "total": total, "page": page, "page_size": page_size}
	except Exception as e:
		return {"error": str(e)[:300], "registros": [], "total": 0}
	finally:
		db.close()


if __name__ == "__main__":
	import uvicorn
	uvicorn.run(app, host="0.0.0.0", port=8000)
