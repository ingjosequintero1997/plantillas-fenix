from __future__ import annotations

import io

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


VERDE = "6BC06B"
VERDE_OSCURO = "1A5E1A"
VERDE_CLARO = "EAF6EC"
GRIS_CLARO = "F5F5F4"
AMARILLO = "F59E0B"
ROJO = "DC2626"

def _estilo_encabezado(cell):
    cell.font = Font(bold=True, color="FFFFFF", size=10)
    cell.fill = PatternFill("solid", fgColor=VERDE_OSCURO)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

def _borde(cell):
    thin = Side(style="thin", color="D1D5DB")
    cell.border = Border(top=thin, bottom=thin, left=thin, right=thin)

def _colorear_resultado(cell, valor):
    if valor is None:
        cell.fill = PatternFill("solid", fgColor="FEE2E2")
        cell.font = Font(bold=True, color=ROJO)
    elif valor >= 95:
        cell.fill = PatternFill("solid", fgColor=VERDE_CLARO)
        cell.font = Font(bold=True, color=VERDE_OSCURO)
    elif valor >= 80:
        cell.fill = PatternFill("solid", fgColor="FEF3C7")
        cell.font = Font(bold=True, color="92400E")
    else:
        cell.fill = PatternFill("solid", fgColor="FEE2E2")
        cell.font = Font(bold=True, color=ROJO)

def _fmt(v):
    if v is None:
        return "#DIV/0!"
    return round(v, 2)

def _titulo_seccion(ws, row, texto):
    ws.cell(row=row, column=1, value=texto)
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
    c = ws.cell(row=row, column=1)
    c.font = Font(bold=True, size=11, color=VERDE_OSCURO)
    c.fill = PatternFill("solid", fgColor=GRIS_CLARO)
    c.alignment = Alignment(horizontal="left", vertical="center")
    return row + 1

def build_indicadores_excel(pare, descriptivos=None, nombre_cargue=""):
    """Genera un Excel con los indicadores PARE MM (departamental + municipal)."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Indicadores PARE MM"
    ws.sheet_view.showGridLines = False

    widths = [72, 14, 14, 12, 14]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    row = 1
    # Titulo principal
    ws.cell(row=row, column=1, value="COHORTE DE GESTANTES PARE MM")
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
    c = ws.cell(row=row, column=1)
    c.font = Font(bold=True, size=16, color="FFFFFF")
    c.fill = PatternFill("solid", fgColor=VERDE_OSCURO)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[row].height = 30
    row += 1

    ws.cell(row=row, column=1, value=f"Nivel Departamental - Fecha referencia: {pare.get('fecha_referencia', '')} - Total gestantes: {pare.get('total_gestantes', 0)}")
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
    c = ws.cell(row=row, column=1)
    c.font = Font(size=10, color=VERDE_OSCURO)
    c.fill = PatternFill("solid", fgColor=VERDE_CLARO)
    c.alignment = Alignment(horizontal="left", vertical="center")
    row += 2

    # Encabezados tabla departamental
    headers = ["Indicador", "Numerador (a)", "Denominador (b)", "Coef.", "Resultado %"]
    for ci, h in enumerate(headers, 1):
        cell = ws.cell(row=row, column=ci, value=h)
        _estilo_encabezado(cell)
        _borde(cell)
    ws.row_dimensions[row].height = 24
    row += 1

    for ind in pare.get("lista", []):
        ws.cell(row=row, column=1, value=ind.get("label", "")).alignment = Alignment(wrap_text=True, vertical="center")
        ws.cell(row=row, column=2, value=ind.get("numerador", 0)).alignment = Alignment(horizontal="center")
        ws.cell(row=row, column=3, value=ind.get("denominador", 0)).alignment = Alignment(horizontal="center")
        ws.cell(row=row, column=4, value=ind.get("coeficiente", 100)).alignment = Alignment(horizontal="center")
        rcell = ws.cell(row=row, column=5, value=_fmt(ind.get("resultado")))
        rcell.alignment = Alignment(horizontal="center")
        _colorear_resultado(rcell, ind.get("resultado"))
        for ci in range(1, 6):
            _borde(ws.cell(row=row, column=ci))
        row += 1

    # Seccion por municipio
    por_municipio = pare.get("por_municipio", [])
    if por_municipio:
        row += 1
        row = _titulo_seccion(ws, row, "INDICADORES POR MUNICIPIO")
        row += 1
        for m in por_municipio:
            # Sub-titulo del municipio
            ws.cell(row=row, column=1, value=f"{m.get('municipio', '')} - {m.get('total', 0)} gestantes")
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
            c = ws.cell(row=row, column=1)
            c.font = Font(bold=True, size=10, color=VERDE_OSCURO)
            c.fill = PatternFill("solid", fgColor=VERDE_CLARO)
            row += 1
            # Encabezados
            for ci, h in enumerate(headers, 1):
                cell = ws.cell(row=row, column=ci, value=h)
                _estilo_encabezado(cell)
                _borde(cell)
            row += 1
            for ind in m.get("indicadores", []):
                ws.cell(row=row, column=1, value=ind.get("label", "")).alignment = Alignment(wrap_text=True, vertical="center")
                ws.cell(row=row, column=2, value=ind.get("numerador", 0)).alignment = Alignment(horizontal="center")
                ws.cell(row=row, column=3, value=ind.get("denominador", 0)).alignment = Alignment(horizontal="center")
                ws.cell(row=row, column=4, value=ind.get("coeficiente", 100)).alignment = Alignment(horizontal="center")
                rcell = ws.cell(row=row, column=5, value=_fmt(ind.get("resultado")))
                rcell.alignment = Alignment(horizontal="center")
                _colorear_resultado(rcell, ind.get("resultado"))
                for ci in range(1, 6):
                    _borde(ws.cell(row=row, column=ci))
                row += 1
            row += 1

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf