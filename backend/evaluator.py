from __future__ import annotations
import io
import re
from datetime import datetime, timedelta
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.utils import get_column_letter

# ── Helpers ──────────────────────────────────────────────────────────────────

def _to_num(val: str):
    try:
        return float(val.replace(",", "."))
    except (ValueError, AttributeError):
        return None


def _to_int(val: str):
    try:
        return int(float(val.replace(",", ".")))
    except (ValueError, AttributeError):
        return None


def _parse_date(val: str):
    if not val or val.strip().upper() in ("SIN DATO", "1900-01-01", ""):
        return None
    try:
        return pd.to_datetime(val.strip(), dayfirst=False, errors="coerce")
    except Exception:
        return None


def _is_yes(val: str) -> bool:
    v = val.strip().upper() if val else ""
    return v in ("SI", "SÍ", "YES", "1", "X", "DX CONFIRMADO")


def _six_months_ago(ref: datetime | None = None) -> datetime:
    ref = ref or datetime.now()
    return ref - timedelta(days=183)


def _age_in_range(val: str, lo: int = 18, hi: int = 69) -> bool:
    age = _to_int(val)
    return age is not None and lo <= age <= hi


def _find_col(df: pd.DataFrame, pattern: str) -> str | None:
    pat = re.sub(r"\s+", "", pattern.upper())
    for col in df.columns:
        if re.sub(r"\s+", "", col.upper()) == pat:
            return col
    for col in df.columns:
        if pat in re.sub(r"\s+", "", col.upper()):
            return col
    return None


def _col_val(row: pd.Series, col: str | None):
    if col is None:
        return ""
    v = row.get(col)
    return str(v).strip() if pd.notna(v) else ""


# ── RCV evaluation ───────────────────────────────────────────────────────────

def _evaluate_rcv(df: pd.DataFrame, today: datetime | None = None) -> tuple[pd.DataFrame, pd.DataFrame]:
    """
    Returns (indicators_df, patients_df) for the RCV template.
    patients_df contains all patient data plus evaluation columns.
    """
    t = today or datetime.now()
    six_months = _six_months_ago(t)

    # Find relevant columns
    col_dm      = _find_col(df, "DX CONFIRMADO DM")
    col_hta     = _find_col(df, "DX CONFIRMADO HTA")
    col_hba1c   = _find_col(df, "REPORTE DE HEMOGLOBINA GLICOSILADA")
    col_hba1c_f = _find_col(df, "FECHA DE REPORTE DE HEMOGLOBINA GLICOSILADA")
    col_sys     = _find_col(df, "ULTIMA TENSION ARTERIAL SISTOLICA")
    col_dias    = _find_col(df, "ULTIMA TENSION ARTERIAL DIASTOLICA")
    col_pa_f    = _find_col(df, "FECHA DE LA ULTIMA TOMA DE PRESION ARTERIAL")
    col_edad    = _find_col(df, "EDAD")
    col_reg     = _find_col(df, "REGIMEN DE AFILIACION")

    # Evaluar cada paciente
    rows = []
    for _, row in df.iterrows():
        r = {}

        # Datos básicos
        for c in df.columns:
            r[c] = _col_val(row, c)

        # Indicador 1: DM controlado (HbA1c < 7%)
        dm_yes = _is_yes(_col_val(row, col_dm))
        hba1c_val = _to_num(_col_val(row, col_hba1c))
        hba1c_date = _parse_date(_col_val(row, col_hba1c_f))
        dm_controlled = (
            dm_yes
            and hba1c_val is not None
            and hba1c_val < 7.0
            and hba1c_date is not None
            and hba1c_date >= six_months
        )
        r["_DM_CONTROLADO"] = "SI" if dm_controlled else "NO"
        r["_TIENE_DM"] = "SI" if dm_yes else "NO"

        # Indicador 2: PA < 140/90
        sys = _to_int(_col_val(row, col_sys))
        dias = _to_int(_col_val(row, col_dias))
        pa_date = _parse_date(_col_val(row, col_pa_f))
        pa_140_90 = (
            sys is not None and dias is not None
            and sys < 140 and dias < 90
            and pa_date is not None and pa_date >= six_months
        )
        r["_PA_140_90"] = "SI" if pa_140_90 else "NO"

        # Indicador 3: PA < 150/90
        pa_150_90 = (
            sys is not None and dias is not None
            and sys < 150 and dias < 90
            and pa_date is not None and pa_date >= six_months
        )
        r["_PA_150_90"] = "SI" if pa_150_90 else "NO"

        # Indicador 4: Captación HTA 18-69 subsidiado
        edad    = _col_val(row, col_edad)
        reg     = _col_val(row, col_reg)
        hta_yes = _is_yes(_col_val(row, col_hta))
        in_range = _age_in_range(edad)
        subsidiado = reg.strip().upper() == "SUBSIDIADO"
        r["_HTA_CAPTADO"] = "SI" if (in_range and subsidiado and hta_yes) else "NO"
        r["_POBLACION_HTA"] = "SI" if (in_range and subsidiado) else "NO"

        # Indicador 5: Captación DM 18-69 subsidiado
        r["_DM_CAPTADO"] = "SI" if (in_range and subsidiado and dm_yes) else "NO"
        r["_POBLACION_DM"] = "SI" if (in_range and subsidiado) else "NO"

        rows.append(r)

    patients = pd.DataFrame(rows)

    # ── Cálculo de indicadores agregados ──
    total_pob = len(patients)

    def _count(flag_col: str) -> int:
        return int((patients[flag_col] == "SI").sum())

    def _pct(n: int, d: int) -> float:
        return round(n / d * 100, 2) if d else 0.0

    # Indicador 1
    n1 = _count("_DM_CONTROLADO")
    d1 = _count("_TIENE_DM")
    # Indicador 2
    n2 = _count("_PA_140_90")
    d2 = total_pob
    # Indicador 3
    n3 = _count("_PA_150_90")
    d3 = total_pob
    # Indicador 4
    n4 = _count("_HTA_CAPTADO")
    d4 = _count("_POBLACION_HTA")
    # Indicador 5
    n5 = _count("_DM_CAPTADO")
    d5 = _count("_POBLACION_DM")

    indicators = pd.DataFrame([
        {
            "INDICADOR": "Diabéticos controlados",
            "DEFINICIÓN": "Pacientes con DM con HbA1c < 7% en últimos 6 meses",
            "NUMERADOR": n1,
            "DENOMINADOR": d1,
            "POBLACIÓN": f"Solo pacientes con DM ({d1} de {total_pob})",
            "META": "Bueno > 50% | Aceptable 30-50% | Crítico < 30%",
            "CUMPLIMIENTO": f"{_pct(n1, d1)}%",
        },
        {
            "INDICADOR": "Control PA < 140/90",
            "DEFINICIÓN": "Pacientes con PA < 140/90 mmHg en último semestre",
            "NUMERADOR": n2,
            "DENOMINADOR": d2,
            "POBLACIÓN": f"Todos los pacientes ({d2})",
            "META": "Bueno > 60% | Aceptable 40-60% | Crítico < 40%",
            "CUMPLIMIENTO": f"{_pct(n2, d2)}%",
        },
        {
            "INDICADOR": "Control PA < 150/90",
            "DEFINICIÓN": "Pacientes con PA < 150/90 mmHg en último semestre",
            "NUMERADOR": n3,
            "DENOMINADOR": d3,
            "POBLACIÓN": f"Todos los pacientes ({d3})",
            "META": "Bueno > 60% | Aceptable 40-60% | Crítico < 40%",
            "CUMPLIMIENTO": f"{_pct(n3, d3)}%",
        },
        {
            "INDICADOR": "Captación HTA 18-69 subsidiado",
            "DEFINICIÓN": "Pacientes 18-69 años régimen subsidiado con Dx HTA",
            "NUMERADOR": n4,
            "DENOMINADOR": d4,
            "POBLACIÓN": f"Solo subsidiados 18-69 años ({d4} de {total_pob})",
            "META": "> 60%",
            "CUMPLIMIENTO": f"{_pct(n4, d4)}%",
        },
        {
            "INDICADOR": "Captación DM 18-69 subsidiado",
            "DEFINICIÓN": "Pacientes 18-69 años régimen subsidiado con Dx DM",
            "NUMERADOR": n5,
            "DENOMINADOR": d5,
            "POBLACIÓN": f"Solo subsidiados 18-69 años ({d5} de {total_pob})",
            "META": "> 60%",
            "CUMPLIMIENTO": f"{_pct(n5, d5)}%",
        },
    ])

    return indicators, patients


# ── Generic fallback ─────────────────────────────────────────────────────────

def _evaluate_generic(df: pd.DataFrame) -> tuple[pd.DataFrame, pd.DataFrame]:
    indicators = pd.DataFrame([
        {"INDICADOR": "No hay indicadores definidos para esta plantilla",
         "DEFINICIÓN": "", "NUMERADOR": "", "DENOMINADOR": "",
         "META": "", "CUMPLIMIENTO": ""}
    ])
    return indicators, df


# ── Public API ───────────────────────────────────────────────────────────────

def evaluate(df: pd.DataFrame, template_key: str, today: datetime | None = None) -> tuple[pd.DataFrame, pd.DataFrame]:
    """
    Evaluate indicators for the given corrected data.
    Only RCV template has defined indicators.
    """
    if template_key == "rcv":
        return _evaluate_rcv(df, today)
    available = {"rcv"}
    raise ValueError(
        f"La plantilla '{template_key}' no tiene indicadores definidos. "
        f"Plantillas con indicadores: {', '.join(sorted(available))}"
    )


def _compliance_level(value_str: str, meta: str) -> str:
    """Return 'bueno', 'aceptable', 'critico', or 'neutral'."""
    try:
        v = float(str(value_str).replace("%", "").replace(",", "."))
    except (ValueError, AttributeError):
        return "neutral"
    # Meta con tres niveles: Bueno > X | Aceptable X-Y | Crítico < Y
    if "50" in meta and "30" in meta:
        if v > 50: return "bueno"
        if v >= 30: return "aceptable"
        return "critico"
    if "60" in meta and "40" in meta:
        if v > 60: return "bueno"
        if v >= 40: return "aceptable"
        return "critico"
    # Meta simple: > X%
    if "60" in meta:
        return "bueno" if v > 60 else "critico"
    return "neutral"


# ── Excel: write-optimized generation ────────────────────────────────────────

_EXCEL_HEADER_FILL = PatternFill("solid", fgColor="1B5E20")
_EXCEL_HEADER_FONT = Font(bold=True, color="FFFFFF", size=10, name="Calibri")
_EXCEL_DATA_FONT = Font(size=10, name="Calibri")
_EXCEL_TITLE_FONT = Font(bold=True, size=16, name="Calibri", color="1B5E20")
_EXCEL_SUBTITLE_FONT = Font(size=11, name="Calibri", color="475569")
_EXCEL_THIN_BORDER = Border(
    left=Side(style="thin", color="CBD5E1"),
    right=Side(style="thin", color="CBD5E1"),
    top=Side(style="thin", color="CBD5E1"),
    bottom=Side(style="thin", color="CBD5E1"),
)
_EXCEL_SI_FONT = Font(size=10, name="Calibri", bold=True, color="166534")
_EXCEL_NO_FONT = Font(size=10, name="Calibri", color="94A3B8")


def _clean_display(name: str) -> str:
    n = name.replace("_", " ").strip()
    return n.title() if len(n) < 30 else n


def _find_id_col(cols: list[str], *patterns: str) -> str:
    for pat in patterns:
        pn = re.sub(r"\s+", "", pat.upper())
        for c in cols:
            if re.sub(r"\s+", "", c.upper()) == pn:
                return c
            if pn in re.sub(r"\s+", "", c.upper()):
                return c
    return cols[0]


def _write_wo_sheet(ws, headers: list[str], rows: list[list],
                    flag_col_idx: int | None = None, fill=None):
    """Write a sheet in write-only mode using append. Very fast."""
    from openpyxl.cell.cell import Cell

    tb = _EXCEL_THIN_BORDER
    al = Alignment(horizontal="center", vertical="center", wrap_text=True)
    hfill = fill or _EXCEL_HEADER_FILL

    # Header row — styled Cell objects
    hcells = []
    for h in headers:
        c = Cell(ws, value=h)
        c.font = _EXCEL_HEADER_FONT; c.fill = hfill; c.border = tb; c.alignment = al
        hcells.append(c)
    ws.append(hcells)

    # Data rows — plain values for speed, Cell only for the flag column
    df = _EXCEL_DATA_FONT
    si_font = _EXCEL_SI_FONT
    no_font = _EXCEL_NO_FONT
    for vals in rows:
        if flag_col_idx is not None:
            row = []
            for ci, v in enumerate(vals):
                if ci == flag_col_idx:
                    c = Cell(ws, value=v)
                    c.border = tb
                    vs = str(v).strip()
                    if vs == "SI":
                        c.font = si_font
                    elif vs == "NO":
                        c.font = no_font
                    else:
                        c.font = df
                    row.append(c)
                else:
                    row.append(v)
            ws.append(row)
        else:
            ws.append(vals)

    # Column widths
    for ci, h in enumerate(headers, start=1):
        w = len(str(h)) + 3
        ws.column_dimensions[get_column_letter(ci)].width = min(max(w, 10), 50)


def _indicator_sheet_data(patients: pd.DataFrame, id_cols: list[str],
                          flag_col: str, denom_col: str | None,
                          cumple: bool):
    """Return (headers, rows) for a cumple/no-cumple indicator sheet."""
    if cumple:
        mask = patients.get(flag_col, "") == "SI"
    elif denom_col:
        mask = (patients.get(denom_col, "") == "SI") & (patients.get(flag_col, "") != "SI")
    else:
        mask = patients.get(flag_col, "") != "SI"
    subset = patients[mask]
    cols = id_cols + [flag_col]
    rows = [[row.get(c, "") for c in cols] for _, row in subset.iterrows()]
    return rows, cols


def build_evaluation_excel(indicators: pd.DataFrame, patients: pd.DataFrame) -> io.BytesIO:
    """Generate Excel with Dashboard, full Pacientes sheet, and per-indicator
    cumple/no-cumple sheets (minimal columns for speed). Uses write-only mode."""
    from openpyxl.cell.cell import Cell

    wb = Workbook(write_only=True)

    eval_cols = [c for c in patients.columns if c.startswith("_")]
    data_cols = [c for c in patients.columns if not c.startswith("_")]

    # Identify key patient columns
    doc_col = _find_id_col(data_cols, "DOCUMENTO", "DOC", "IDENTIFICACION", "NUMERO DE DOCUMENTO")
    name_col = _find_id_col(data_cols, "NOMBRE", "NOMBRE DEL PACIENTE", "PACIENTE", "APELLIDOS")
    age_col = _find_id_col(data_cols, "EDAD", "AGE")
    id_cols = [doc_col, name_col, age_col]

    display_cols = data_cols + eval_cols
    patient_rows = [[row.get(c, "") for c in display_cols] for _, row in patients.iterrows()]

    level_fills = {
        "bueno": PatternFill("solid", fgColor="DCFCE7"),
        "aceptable": PatternFill("solid", fgColor="FEF3C7"),
        "critico": PatternFill("solid", fgColor="FEE2E2"),
        "neutral": PatternFill("solid", fgColor="F8FAFC"),
    }
    level_colors = {
        "bueno": "166534", "aceptable": "92400E", "critico": "991B1B", "neutral": "475569",
    }
    level_labels = {
        "bueno": "✓ Cumple meta", "aceptable": "! En alerta", "critico": "✗ Crítico",
    }
    tb = _EXCEL_THIN_BORDER

    # ── Sheet 1: Dashboard ──
    ws = wb.create_sheet(title="Dashboard")

    c = Cell(ws, value="Evaluación de Indicadores — Riesgo Cardiovascular")
    c.font = _EXCEL_TITLE_FONT
    ws.append([c])

    c = Cell(ws, value=f"Total: {len(patients)} pacientes  |  {len(indicators)} indicadores  |  {datetime.now().strftime('%Y-%m-%d %H:%M')}")
    c.font = _EXCEL_SUBTITLE_FONT
    ws.append([c])
    ws.append([])

    ind_headers = ["Indicador", "Numerador", "Denominador", "Cumplimiento", "Estado", "Meta"]
    hcells = []
    al = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for h in ind_headers:
        c = Cell(ws, value=h)
        c.font = _EXCEL_HEADER_FONT; c.fill = _EXCEL_HEADER_FILL; c.border = tb; c.alignment = al
        hcells.append(c)
    ws.append(hcells)

    for _, ind in indicators.iterrows():
        level = _compliance_level(ind["CUMPLIMIENTO"], ind["META"])
        fill = level_fills.get(level, level_fills["neutral"])
        color = level_colors.get(level, "475569")
        vals = [
            ind["INDICADOR"], ind["NUMERADOR"], ind["DENOMINADOR"],
            ind["CUMPLIMIENTO"], level_labels.get(level, ""), ind["META"],
        ]
        row = []
        for ci, v in enumerate(vals, start=1):
            c = Cell(ws, value=v)
            sz = 14 if ci == 4 else 10
            c.font = Font(size=sz, name="Calibri", bold=(ci == 4), color=color)
            c.fill = fill; c.border = tb
            row.append(c)
        ws.append(row)

    for col, w in zip("ABCDEF", [42, 14, 14, 16, 18, 50]):
        ws.column_dimensions[col].width = w

    # ── Sheet 2: Pacientes (full columns) ──
    ws_p = wb.create_sheet(title="Pacientes")
    hcells = []
    for h in display_cols:
        c = Cell(ws_p, value=h)
        c.font = _EXCEL_HEADER_FONT; c.fill = _EXCEL_HEADER_FILL; c.border = tb
        hcells.append(c)
    ws_p.append(hcells)

    # Build eval column index set for styling
    eval_name_set = set(eval_cols)
    for vals in patient_rows:
        row = []
        for ci, v in enumerate(vals):
            col_name = display_cols[ci]
            if col_name in eval_name_set:
                c = Cell(ws_p, value=v)
                c.border = tb
                vs = str(v).strip()
                if vs == "SI":
                    c.font = _EXCEL_SI_FONT
                elif vs == "NO":
                    c.font = _EXCEL_NO_FONT
                else:
                    c.font = _EXCEL_DATA_FONT
                row.append(c)
            else:
                row.append(v)
        ws_p.append(row)

    for ci, h in enumerate(display_cols, start=1):
        w = len(str(h)) + 3
        ws_p.column_dimensions[get_column_letter(ci)].width = min(max(w, 10), 50)

    # ── Per-indicator sheets: cumple + no cumple (minimal columns) ──
    indicator_defs = [
        ("_DM_CONTROLADO", "_TIENE_DM", "DM Controlado"),
        ("_PA_140_90",     None,        "PA <140-90"),
        ("_PA_150_90",     None,        "PA <150-90"),
        ("_HTA_CAPTADO",   "_POBLACION_HTA", "Captación HTA"),
        ("_DM_CAPTADO",    "_POBLACION_DM",  "Captación DM"),
    ]

    green_fill = PatternFill("solid", fgColor="1B5E20")
    red_fill = PatternFill("solid", fgColor="991B1B")

    for flag_col, denom_col, label in indicator_defs:
        for cumple, prefix, hfill in [
            (True, "", green_fill),
            (False, "No ", red_fill),
        ]:
            rows, cols = _indicator_sheet_data(patients, id_cols, flag_col, denom_col, cumple)
            if not rows:
                continue
            title = f"{prefix}{label}"[:31]
            ws_ind = wb.create_sheet(title=title)

            headers = [_clean_display(c) for c in cols]
            flag_idx = len(id_cols)  # flag is the last column
            _write_wo_sheet(ws_ind, headers, rows, flag_col_idx=flag_idx, fill=hfill)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf
